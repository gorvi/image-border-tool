"""
主窗口模块
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, colorchooser, simpledialog
from PIL import Image, ImageTk, ImageDraw
import json
import random
import os
import sys
import subprocess
from datetime import datetime

from auth_manager import auth  # [AUTH] 导入授权管理器

from canvas_widget import CanvasWidget
from image_processor import ImageProcessor, CompositeImage
from constants import (SIZE_PRESETS, BORDER_STYLES, STICKER_LIST, COLORS, 
                      BORDER_STYLES_WITH_PREVIEW, BORDER_CATEGORIES, 
                      BORDER_COLORS, BORDER_STYLE_NAMES,
                      BORDER_SHAPES, BORDER_LINE_STYLES, DEFAULT_BACKGROUNDS,
                      BORDER_PATTERNS, BACKGROUND_PATTERNS, DEFAULT_BORDER_CONFIG,
                      QUICK_COLORS)
from color_picker import ColorPicker
from color_wheel_picker import ColorWheelPicker
from PIL import Image, ImageTk, ImageDraw, ImageFont


class Tooltip:
    """鼠标悬停提示工具类"""
    def __init__(self, widget, text, delay=500):
        self.widget = widget
        self.text = text
        self.delay = delay
        self.tooltip_window = None
        self.after_id = None
        
        widget.bind('<Enter>', self._on_enter)
        widget.bind('<Leave>', self._on_leave)
    
    def _on_enter(self, event):
        self.after_id = self.widget.after(self.delay, self._show_tooltip)
    
    def _on_leave(self, event):
        if self.after_id:
            self.widget.after_cancel(self.after_id)
            self.after_id = None
        self._hide_tooltip()
    
    def _show_tooltip(self):
        if self.tooltip_window:
            return
        
        x = self.widget.winfo_rootx() + 20
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 5
        
        self.tooltip_window = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        tw.attributes('-topmost', True)
        
        label = tk.Label(tw, text=self.text, bg='#333333', fg='#FFFFFF',
                        font=('SF Pro Text', 10), padx=8, pady=4,
                        relief='solid', borderwidth=1)
        label.pack()
    
    def _hide_tooltip(self):
        if self.tooltip_window:
            self.tooltip_window.destroy()
            self.tooltip_window = None


class MainWindow(tk.Tk):
    """主窗口类"""
    
    def __init__(self):
        super().__init__()
        
        self.title('图片套版工具')
        
        # 获取屏幕尺寸并设置窗口大小（屏幕的80%）
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        window_width = int(screen_width * 0.85)
        window_height = int(screen_height * 0.85)
        
        # 居中显示
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        
        self.geometry(f'{window_width}x{window_height}+{x}+{y}')
        self.minsize(1200, 700)  # 最小窗口尺寸
        self.configure(bg=COLORS['bg'])
        
        # 初始化变量
        self.image_processor = ImageProcessor()
        self.current_size_preset = SIZE_PRESETS[3]  # 默认小红书3:4
        self.current_border = BORDER_STYLES[0]  # 默认无边框
        self.batch_images = []  # 批量图片列表
        self.sticker_images = {}  # 缓存贴纸图片
        self.border_preview_images = {}  # 缓存边框预览图
        self.sticker_photo_refs = []  # 保持图片引用，防止被垃圾回收
        self.border_photo_refs = []  # 保持边框图片引用
        
        # 历史记录系统
        self.history_stack = []  # 历史记录栈
        self.history_index = -1  # 当前历史位置
        self.max_history = 30  # 最大历史记录数
        
        # 边框选择状态（旧版）
        self.selected_border_category = 'modern'
        self.selected_border_color = 'black'
        
        # 自定义边框配置
        # 边框配置 - 使用默认值
        self.border_config = {
            'shape': DEFAULT_BORDER_CONFIG['shape'],
            'line_style': DEFAULT_BORDER_CONFIG['line_style'],
            'width': DEFAULT_BORDER_CONFIG['width'],
            'radius': DEFAULT_BORDER_CONFIG['radius'],
            'color': DEFAULT_BORDER_CONFIG['color'],
            'pattern': DEFAULT_BORDER_CONFIG['pattern'],
            'pattern_color': DEFAULT_BORDER_CONFIG['pattern_color'],
            'pattern_size': DEFAULT_BORDER_CONFIG['pattern_size'],
        }
        
        # 背景配置
        self.background_color = '#FFFFFF'
        self.background_pattern = 'none'
        self.background_pattern_color = '#E0E0E0'
        self.background_pattern_size = 10
        self.background_image = None
        
        # 颜色方块引用
        self.bg_color_canvases = {}
        self.border_color_canvas = None
        
        # 加载资源
        self.load_sticker_images()
        self.load_border_preview_images()
        
        # 历史记录
        self.history = []
        self.history_index = -1
        
        # 预设主题列表
        self.preset_themes = []
        
        # 滚动控制
        self._active_scroll_widget = None
        
        # 批量处理配置
        self.batch_input_dir = ''  # 输入目录
        self.batch_output_dir = ''  # 输出目录
        self.processed_images = set()  # 已处理的图片集合
        # self.batch_regenerate_all = tk.BooleanVar(value=False) # 已废弃
        
        # 批量随机化选项
        # 批量随机化选项
        self.batch_random_color = tk.BooleanVar(value=True)
        self.batch_random_style = tk.BooleanVar(value=True)
        self.batch_random_pattern = tk.BooleanVar(value=True)
        self.batch_random_highlight = tk.BooleanVar(value=True) # NEW
        self.batch_random_font_style = tk.BooleanVar(value=True) # 随机字体样式
        self.batch_match_canvas = tk.BooleanVar(value=True) # 参考画布位置
        
        # 文字层配置
        self.text_layers = []  # 文字层列表
        self.current_text_config = {
            'content': '',
            'font_size': 48,
            'color': '#FFFFFF',
            'font_family': 'pingfang',
            'align': 'left',
            'position': 'top',
            'margin': 20,
            'shadow': {'enabled': True, 'color': '#000000', 'offset': (2, 2), 'blur': 4},
            'stroke': {'enabled': False, 'color': '#000000', 'width': 2},
        }
        
        # 批量文字配置
        self.batch_text_dir = ''  # 文本目录
        self.batch_use_text_dir = tk.BooleanVar(value=False)  # 使用文本目录
        
        # 自动高亮定时器
        self._highlight_timer = None
        
        # 加载用户设置
        self.load_settings()
        
         
        # [AUTH] 初始化后检查授权
        self.check_auth_at_startup()
        self.create_auth_menu()
        
        # [UI] 创建界面
        self.create_widgets()

    def create_auth_menu(self):
        """创建授权菜单"""
        menubar = tk.Menu(self)
        self.config(menu=menubar)
        
        # 帮助菜单
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="帮助", menu=help_menu)
        help_menu.add_command(label="软件激活 / 授权信息", command=self.show_activation_dialog)
        help_menu.add_command(label="用量统计", command=self.show_usage_dialog)
        help_menu.add_command(label="关于", command=lambda: messagebox.showinfo("关于", "图片批量套版工具 v1.0"))

    def check_auth_at_startup(self):
        """启动时检查授权"""
        status = auth.get_status()
        title_suffix = ""
        
        if status['status'] == 'limited':
            messagebox.showwarning("今日额度耗尽", f"{status['msg']}\n请明天再来或激活解除限制。")
            title_suffix = " [免费版 - 今日额度耗尽]"
        elif status['status'] == 'trial':
             # 试用期提示
             title_suffix = f" [全功能体验版 - {status['msg']}]"
        elif status['status'] == 'free':
             title_suffix = f" [免费版 - {status['msg']}]"
             
        if title_suffix:
            self.title(f"{self.title().split(' [')[0]}{title_suffix}")

    def show_activation_dialog(self):
        """显示激活对话框"""
        info = auth.get_activation_info()
        status_msg = info['status']['msg']
        
        dialog = tk.Toplevel(self)
        dialog.title("软件激活")
        dialog.geometry("500x350")
        dialog.resizable(False, False)
        
        # 居中
        dialog.transient(self)
        dialog.grab_set()
        
        padding = 20
        
        # 标题
        tk.Label(dialog, text="软件授权激活", font=("Arial", 16, "bold")).pack(pady=padding)
        
        # 状态
        status_frame = tk.Frame(dialog)
        status_frame.pack(fill=tk.X, padx=padding)
        tk.Label(status_frame, text=f"当前状态: {status_msg}", fg="red" if info['status']['status']!='activated' else "green").pack(anchor='w')
        
        # 机器码区域
        code_frame = tk.LabelFrame(dialog, text="您的机器码 (请复制发给管理员)", pady=10)
        code_frame.pack(fill=tk.X, padx=padding, pady=10)
        
        entry_machine = tk.Entry(code_frame, font=("Arial", 12), justify='center')
        entry_machine.pack(fill=tk.X, padx=10)
        entry_machine.insert(0, info['machine_code'])
        entry_machine.config(state='readonly') # 只读
        
        # 激活码输入
        input_frame = tk.LabelFrame(dialog, text="输入激活码", pady=10)
        input_frame.pack(fill=tk.X, padx=padding, pady=10)
        
        entry_key = tk.Entry(input_frame, font=("Arial", 12), justify='center')
        entry_key.pack(fill=tk.X, padx=10)
        
        def do_activate():
            code = entry_key.get()
            if auth.validate_activation_code(code):
                messagebox.showinfo("激活成功", "感谢您的支持！软件已永久激活。")
                dialog.destroy()
                self.title("图片批量套版工具 [永久激活版]") # 刷新标题
            else:
                messagebox.showerror("激活失败", "激活码错误，请检查是否对应本机机器码。")

        tk.Button(dialog, text="立即激活", command=do_activate, bg="#007AFF", fg="black", font=("Arial", 12, "bold"), height=2).pack(fill=tk.X, padx=padding, pady=10)

    def show_usage_dialog(self):
        """显示用量统计"""
        stats = auth.get_usage_stats()
        status = auth.get_status()
        
        msg = (f"📊 用量统计\n\n"
               f"累计导出总数: {stats['total_count']} 张\n"
               f"今日导出数量: {stats['daily_count']} 张\n"
               f"软件安装日期: {stats['install_date']}\n\n"
               f"当前账户状态: {status['msg']}")
               
        messagebox.showinfo("用量统计", msg)

    def create_widgets(self):
        """创建界面组件 - 毛玻璃风格"""
        # 主容器 - 使用 PanedWindow 实现可调整大小
        self.paned_window = tk.PanedWindow(self, orient=tk.HORIZONTAL, 
                                          bg=COLORS['bg'], sashwidth=4, sashpad=0,
                                          showhandle=False, borderwidth=0)
        # 状态变量 - 记录当前拖拽的sash索引
        self.dragging_sash_index = None
        
        self.paned_window.pack(fill=tk.BOTH, expand=True)
        
        # 绑定鼠标事件处理拖拽限制
        # ButtonPress: 检测点中了哪个 sash
        self.paned_window.bind('<ButtonPress-1>', self.start_sash_drag, add='+')
        # B1-Motion: 拦截拖拽，实施限制
        self.paned_window.bind('<B1-Motion>', self.on_sash_drag)
        # ButtonRelease: 结束拖拽
        self.paned_window.bind('<ButtonRelease-1>', self.end_sash_drag, add='+')
        
        # 延迟绑定窗口大小改变事件
        self.after(1000, self.bind_configure_limit)
        
        # 左侧面板容器
        self.left_container = tk.Frame(self.paned_window, bg=COLORS['bg'])
        # self.left_container.bind('<Configure>', self.on_panel_resize) # 移除容易导致闪烁的 Configure 绑定
        
        self.left_panel = self.create_left_panel(self.left_container)
        self.left_panel.pack(fill=tk.BOTH, expand=True, padx=(8, 0), pady=8)
        
        self.left_panel_visible = True
        
        # 将左侧容器添加到 PanedWindow (设置最小宽度 260)
        self.paned_window.add(self.left_container, minsize=260, width=280)

        # 中间画布区域
        self.center_panel = self.create_center_panel(self.paned_window)
        # 设置 stretch='always' 确保中间区域优先占用空间
        self.paned_window.add(self.center_panel, stretch='always', minsize=360)
        
        # 绑定文字交互回调
        if hasattr(self, 'canvas_widget'):
            self.canvas_widget.set_text_callback(self.on_text_transform)
        
        # 右侧面板
        self.right_panel = self.create_right_panel(self.paned_window)
        # 初始宽度设小一点，限制最小宽度
        self.paned_window.add(self.right_panel, minsize=260, width=280)
        
        # 延迟应用默认边框（等待画布初始化完成）
        self.after(200, self.apply_default_border)
        
        # 绑定快捷键
        self.bind('<Command-z>', lambda e: self.undo())
        self.bind('<Command-Shift-Z>', lambda e: self.redo())
        self.bind('<Command-s>', lambda e: self.export_image())
        self.bind('<Configure>', self.on_window_resize)
    
    def load_settings(self):
        """加载用户设置"""
        import json
        settings_path = os.path.join(os.path.dirname(__file__), 'settings.json')
        try:
            if os.path.exists(settings_path):
                with open(settings_path, 'r', encoding='utf-8') as f:
                    settings = json.load(f)
                    self.batch_input_dir = settings.get('batch_input_dir', '')
                    self.batch_output_dir = settings.get('batch_output_dir', '')
                    self.batch_text_dir = settings.get('batch_text_dir', '') # NOW SAVED
                    self.processed_images = set(settings.get('processed_images', []))
                    self.preset_themes = settings.get('preset_themes', [])
                    print(f"✓ 已加载设置: 输入={self.batch_input_dir}, 输出={self.batch_output_dir}, 预设={len(self.preset_themes)}个")
        except Exception as e:
            print(f"加载设置失败: {e}")
    
    def save_settings(self):
        """保存用户设置"""
        import json
        settings_path = os.path.join(os.path.dirname(__file__), 'settings.json')
        try:
            settings = {
                'batch_input_dir': self.batch_input_dir,
                'batch_input_dir': self.batch_input_dir,
                'batch_output_dir': self.batch_output_dir,
                'batch_text_dir': self.batch_text_dir, # NOW SAVED
                'processed_images': list(self.processed_images),
                'preset_themes': self.preset_themes
            }
            with open(settings_path, 'w', encoding='utf-8') as f:
                json.dump(settings, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"保存设置失败: {e}")
    
    def batch_log(self, message):
        """输出日志到批量处理日志框"""
        if hasattr(self, 'batch_log_text'):
            self.batch_log_text.config(state=tk.NORMAL)
            self.batch_log_text.insert(tk.END, f"{message}\n")
            self.batch_log_text.see(tk.END)  # 自动滚动到底部
            self.batch_log_text.config(state=tk.DISABLED)
            self.update_idletasks()  # 强制更新UI
    
    def on_window_resize(self, event):
        """窗口大小改变时调整画布"""
        if event.widget == self:
            # 延迟调整，避免频繁触发
            if hasattr(self, 'resize_timer'):
                self.after_cancel(self.resize_timer)
            self.resize_timer = self.after(100, self.adjust_canvas_display)
    
    def adjust_canvas_display(self):
        """自适应调整画布显示，并限制侧边栏宽度"""
        try:
            # 1. 强制限制侧边栏宽度 (最大 1/4)
            if hasattr(self, 'paned_window'):
                total_width = self.paned_window.winfo_width()
                if total_width > 100:
                    max_side = int(total_width * 0.25)
                    
                    # 检查左侧 sash (index 0)
                    try:
                        sash0_x, sash0_y = self.paned_window.sash_coord(0)
                        if sash0_x > max_side:
                            self.paned_window.sash_place(0, max_side, sash0_y)
                            # print(f"Limit Left: {sash0_x} -> {max_side}")
                    except Exception:
                        pass
                    
                    # 检查右侧 sash (index 1)
                    try:
                        sash1_x, sash1_y = self.paned_window.sash_coord(1)
                        right_width = total_width - sash1_x
                        if right_width > max_side:
                            target_x = total_width - max_side
                            self.paned_window.sash_place(1, target_x, sash1_y)
                            # print(f"Limit Right: {right_width} -> {max_side}")
                    except Exception:
                        pass

            if hasattr(self, 'canvas_widget') and hasattr(self, 'center_panel'):
                # 获取中间面板实际大小
                self.center_panel.update_idletasks()
                panel_width = self.center_panel.winfo_width()
                panel_height = self.center_panel.winfo_height()
                
                # 预留一点边距
                if panel_width > 40 and panel_height > 40:
                    available_width = panel_width - 40
                    available_height = panel_height - 40
                    
                    # 获取当前预设比例
                    preset = self.current_size_preset
                    ratio = preset['width'] / preset['height']
                    
                    # 计算保持比例的尺寸
                    if ratio > available_width / available_height:
                        # 宽图，以宽度为准
                        new_width = available_width
                        new_height = int(new_width / ratio)
                    else:
                        # 高图或方图，以高度为准
                        new_height = available_height
                        new_width = int(new_height * ratio)
                    
                    # 只有当尺寸发生显著变化时才调整
                    current_width = self.canvas_widget.width
                    current_height = self.canvas_widget.height
                    
                    if abs(new_width - current_width) > 5 or abs(new_height - current_height) > 5:
                        self.canvas_widget.resize_canvas(new_width, new_height)
                        # 如果有当前图片，重新显示
                        if hasattr(self, 'image_processor') and self.image_processor.current_image:
                            self.canvas_widget.display_image(self.image_processor.current_image)
                        
                        # 重新应用背景图案
                        if hasattr(self, 'background_pattern') and self.background_pattern != 'none':
                             self.canvas_widget.set_background_pattern(
                                self.background_pattern, 
                                self.background_color, 
                                self.background_pattern_color,
                                self.background_pattern_size
                             )
                        
                        # 重新应用边框 (必须在背景图案之后，否则会被覆盖)
                        # 总是重新应用边框以确保它在最上层
                        self.canvas_widget.apply_custom_border(self.border_config)
        except Exception as e:
            print(f"Resize error: {e}")
            pass
        
    def next_tab(self, event=None):
        """切换到下一个标签页"""
        if hasattr(self, 'notebook'):
            current_index = self.notebook.index(self.notebook.select())
            total_tabs = self.notebook.index('end')
            next_index = (current_index + 1) % total_tabs
            self.notebook.select(next_index)
            return "break" # 防止默认行为

    def create_widgets(self):
        """创建界面组件 - 毛玻璃风格"""
        # 主容器 - 使用 PanedWindow 实现可调整大小
        self.paned_window = tk.PanedWindow(self, orient=tk.HORIZONTAL, 
                                          bg=COLORS['bg'], sashwidth=4, sashpad=0,
                                          showhandle=False, borderwidth=0)
        # 状态变量 - 记录当前拖拽的sash索引
        self.dragging_sash_index = None
        
        self.paned_window.pack(fill=tk.BOTH, expand=True)
        
        # 绑定鼠标事件处理拖拽限制
        # ButtonPress: 检测点中了哪个 sash
        self.paned_window.bind('<ButtonPress-1>', self.start_sash_drag, add='+')
        # B1-Motion: 拦截拖拽，实施限制
        self.paned_window.bind('<B1-Motion>', self.on_sash_drag)
        # ButtonRelease: 结束拖拽
        self.paned_window.bind('<ButtonRelease-1>', self.end_sash_drag, add='+')
        
        # 延迟绑定窗口大小改变事件
        self.after(1000, self.bind_configure_limit)
        
        # 左侧面板容器
        self.left_container = tk.Frame(self.paned_window, bg=COLORS['bg'])
        # self.left_container.bind('<Configure>', self.on_panel_resize) # 移除容易导致闪烁的 Configure 绑定
        
        self.left_panel = self.create_left_panel(self.left_container)
        self.left_panel.pack(fill=tk.BOTH, expand=True, padx=(8, 0), pady=8)
        
        self.left_panel_visible = True
        
        # 将左侧容器添加到 PanedWindow (设置最小宽度 260)
        self.paned_window.add(self.left_container, minsize=260, width=280)

        # 中间画布区域
        self.center_panel = self.create_center_panel(self.paned_window)
        # 设置 stretch='always' 确保中间区域优先占用空间
        self.paned_window.add(self.center_panel, stretch='always', minsize=360)
        
        # 绑定文字交互回调
        if hasattr(self, 'canvas_widget'):
            self.canvas_widget.set_text_callback(self.on_text_transform)
        
        # 右侧面板
        self.right_panel = self.create_right_panel(self.paned_window)
        # 初始宽度设小一点，限制最小宽度
        self.paned_window.add(self.right_panel, minsize=260, width=280)
        
        # 延迟应用默认边框（等待画布初始化完成）
        self.after(200, self.apply_default_border)
        
    def bind_configure_limit(self):
        """延迟绑定窗口调整事件"""
        if hasattr(self, 'paned_window'):
             self.paned_window.bind('<Configure>', lambda e: self.on_sash_drag(e, configure=True), add='+')
    
    def start_sash_drag(self, event):
        """开始拖拽：判断点中了哪个sash"""
        if not hasattr(self, 'paned_window'): return
        
        try:
            # 简单的距离判断：Sash 宽度约4px，增加一点容错
            click_x = event.x
            
            # 检查 Sash 0 (左侧)
            try:
                sash0_x, _ = self.paned_window.sash_coord(0)
                if abs(click_x - sash0_x) < 10:
                    self.dragging_sash_index = 0
                    return
            except: pass
            
            # 检查 Sash 1 (右侧)
            try:
                sash1_x, _ = self.paned_window.sash_coord(1)
                if abs(click_x - sash1_x) < 10:
                    self.dragging_sash_index = 1
                    return
            except: pass
            
        except Exception as e:
            print(f"Drag start error: {e}")
            
    def end_sash_drag(self, event):
        """结束拖拽"""
        self.dragging_sash_index = None
        self.adjust_canvas_display()

    def on_sash_drag(self, event=None, configure=False):
        """处理拖拽过程中的限制"""
        if not hasattr(self, 'paned_window'): return
        if not self.winfo_viewable(): return
        
        try:
            total_width = self.paned_window.winfo_width()
            if total_width < 200: return
            
            MIN_SIDE = 260
            # 左侧最大 30%，右侧最大 35%
            max_left = int(total_width * 0.30) 
            max_right = int(total_width * 0.35)
            
            # 如果是窗口调整事件(configure=True)，检查所有 sash 并在越界时修正
            if configure or self.dragging_sash_index is None:
                # 检查所有并修正（此时不 return break，仅修正）
                try:
                    sash0_x, sash0_y = self.paned_window.sash_coord(0)
                    limit = max(MIN_SIDE, max_left)
                    if sash0_x > limit:
                        self.paned_window.sash_place(0, limit, sash0_y)
                except: pass
                
                try:
                    sash1_x, sash1_y = self.paned_window.sash_coord(1)
                    limit_right_panel = max(MIN_SIDE, max_right)
                    limit_x = total_width - limit_right_panel
                    if sash1_x < limit_x:
                         self.paned_window.sash_place(1, limit_x, sash1_y)
                except: pass
                return

            # 如果是主动拖拽 (dragging_sash_index valid)
            # 我们直接控制 sash 位置并拦截事件 (return 'break') 防止冲突
            
            if self.dragging_sash_index == 0:
                # 左侧
                # 目标位置受限于：最小宽度 ~ 最大宽度
                # 注意：sash_place 0 设置的是左侧面板宽度
                target_limit = max(MIN_SIDE, max_left)
                
                # 鼠标位置限制
                new_x = max(MIN_SIDE, min(event.x, target_limit))
                
                self.paned_window.sash_place(0, new_x, 0)
                return "break" # 拦截，防止系统覆盖
                
            elif self.dragging_sash_index == 1:
                # 右侧
                # sash_place 1 设置的是 (左+中) 的宽度
                # 右侧面板宽度 = Total - new_x
                # 限制：RightWidth <= max_right AND RightWidth >= MIN_SIDE
                # 所以: Total - new_x <= max_right  => new_x >= Total - max_right (Left bound)
                #       Total - new_x >= MIN_SIDE   => new_x <= Total - MIN_SIDE (Right bound)
                
                actual_max_right = max(MIN_SIDE, max_right)
                
                left_bound = total_width - actual_max_right
                right_bound = total_width - MIN_SIDE
                
                new_x = max(left_bound, min(event.x, right_bound))
                
                self.paned_window.sash_place(1, new_x, 0)
                return "break" # 拦截，防止系统覆盖

        except Exception as e:
            # print(f"Drag error: {e}")
            pass

    def apply_default_border(self):
        """应用默认边框"""
        self.canvas_widget.apply_custom_border(self.border_config)
        print("✓ 默认边框已应用")
    
    def bind_mousewheel(self, content_widget, scroll_widget=None):
        """绑定鼠标滚轮事件
        
        Args:
            content_widget: 内容widget,鼠标悬停在这里时触发滚动
            scroll_widget: 实际执行滚动的widget(有yview_scroll方法),如果为None则使用content_widget
        """
        target = scroll_widget if scroll_widget else content_widget
        
        def on_mousewheel(event):
            try:
                delta = event.delta
                if abs(delta) > 100:
                    delta = delta // 120
                target.yview_scroll(-delta, "units")
            except:
                pass
            return "break"
        
        # 递归绑定到content_widget及其所有子控件
        def bind_all(w):
            try:
                w.bind('<MouseWheel>', on_mousewheel)
                for child in w.winfo_children():
                    bind_all(child)
            except:
                pass
        
        bind_all(content_widget)
    
    def load_sticker_images(self):
        """加载贴纸PNG图片"""
        assets_dir = os.path.join(os.path.dirname(__file__), 'assets', 'stickers')
        
        for sticker in STICKER_LIST:
            if 'file' in sticker:
                img_path = os.path.join(assets_dir, sticker['file'])
                if os.path.exists(img_path):
                    try:
                        # 加载并调整大小
                        img = Image.open(img_path).convert('RGBA')
                        img = img.resize((32, 32), Image.Resampling.LANCZOS)
                        photo = ImageTk.PhotoImage(img)
                        self.sticker_images[sticker['id']] = photo
                        self.sticker_photo_refs.append(photo)  # 保持引用
                    except Exception as e:
                        print(f"加载贴纸失败 {sticker['file']}: {e}")
    
    def load_border_preview_images(self):
        """加载边框预览图"""
        frames_dir = os.path.join(os.path.dirname(__file__), 'assets', 'borders', 'frames')
        
        if not os.path.exists(frames_dir):
            return
        
        # 加载所有边框预览图
        for filename in os.listdir(frames_dir):
            if filename.endswith('.png'):
                try:
                    img_path = os.path.join(frames_dir, filename)
                    img = Image.open(img_path).convert('RGBA')
                    # 缩小到缩略图尺寸
                    img.thumbnail((60, 60), Image.Resampling.LANCZOS)
                    photo = ImageTk.PhotoImage(img)
                    border_id = filename.replace('.png', '')
                    self.border_preview_images[border_id] = photo
                    self.border_photo_refs.append(photo)
                except Exception as e:
                    print(f"加载边框预览失败 {filename}: {e}")
    
    def create_left_panel(self, parent):
        """创建左侧面板 - 现代深色风格"""
        panel = tk.Frame(
            parent, 
            bg=COLORS['panel_bg'], 
            width=150,
            relief=tk.FLAT,
            bd=0,
            highlightthickness=0
        )
        
        # 尺寸选择 - 紧凑风格
        size_label = tk.Label(
            panel,
            text='尺寸',
            font=('SF Pro Text', 10),
            bg=COLORS['panel_bg'],
            fg=COLORS['text_secondary'],
            anchor='w'
        )
        size_label.pack(fill=tk.X, padx=8, pady=(12, 4))
        
        size_frame = tk.Frame(panel, bg=COLORS['panel_bg'])
        size_frame.pack(fill=tk.X, padx=4, pady=(0, 6))
        
        # 保存尺寸按钮引用 (使用Label替代Button，解决macOS颜色不显示问题)
        self.size_preset_buttons = {}
        
        for preset in SIZE_PRESETS:
            is_selected = preset['id'] == self.current_size_preset['id']
            # 使用Label而非Button，macOS上Button的fg颜色不生效
            btn = tk.Label(
                size_frame,
                text=f"{preset['name']}\n{preset['width']}×{preset['height']}",
                bg=COLORS['selected_bg'] if is_selected else COLORS['bg_tertiary'],
                fg=COLORS['text_bright'] if is_selected else COLORS['text_primary'],
                font=('SF Pro Text', 9),
                pady=6,
                padx=8,
                cursor='hand2',
                anchor='center'
            )
            btn.pack(fill=tk.X, padx=2, pady=1)
            btn.bind('<Button-1>', lambda e, p=preset: self.select_size_preset(p))
            self.size_preset_buttons[preset['id']] = btn
            
            # 鼠标悬停效果
            def on_enter(e, b=btn, pid=preset['id']):
                if self.current_size_preset['id'] != pid:
                    b.config(bg=COLORS['hover'])
            def on_leave(e, b=btn, pid=preset['id']):
                if self.current_size_preset['id'] != pid:
                    b.config(bg=COLORS['bg_tertiary'])
            btn.bind('<Enter>', on_enter)
            btn.bind('<Leave>', on_leave)
        
        # 分隔线
        separator1 = tk.Frame(panel, bg=COLORS['separator'], height=1)
        separator1.pack(fill=tk.X, padx=8, pady=8)
        
        # 上传图片 - 紧凑风格
        upload_label = tk.Label(
            panel,
            text='操作',
            font=('SF Pro Text', 10),
            bg=COLORS['panel_bg'],
            fg=COLORS['text_secondary'],
            anchor='w'
        )
        upload_label.pack(fill=tk.X, padx=8, pady=(0, 4))
        
        upload_frame = tk.Frame(panel, bg=COLORS['panel_bg'])
        upload_frame.pack(fill=tk.X, padx=4, pady=(0, 6))
        
        # 使用Label替代Button
        upload_btn = tk.Label(
            upload_frame,
            text='📁 上传图片',
            bg=COLORS['btn_primary'],
            fg=COLORS['text_bright'],
            font=('SF Pro Text', 10, 'bold'),
            pady=8,
            padx=8,
            cursor='hand2'
        )
        upload_btn.pack(fill=tk.X, padx=2, pady=2)
        upload_btn.bind('<Button-1>', lambda e: self.upload_image())
        upload_btn.bind('<Enter>', lambda e: upload_btn.config(bg=COLORS['accent_hover']))
        upload_btn.bind('<Leave>', lambda e: upload_btn.config(bg=COLORS['btn_primary']))
        
        reset_btn = tk.Label(
            upload_frame,
            text='🔄 重置画布',
            bg=COLORS['bg_tertiary'],
            fg=COLORS['text_primary'],
            font=('SF Pro Text', 10),
            pady=8,
            padx=8,
            cursor='hand2'
        )
        reset_btn.pack(fill=tk.X, padx=2, pady=2)
        reset_btn.bind('<Button-1>', lambda e: self.reset_image())
        reset_btn.bind('<Enter>', lambda e: reset_btn.config(bg=COLORS['hover']))
        reset_btn.bind('<Leave>', lambda e: reset_btn.config(bg=COLORS['bg_tertiary']))
        
        # 分隔线
        separator2 = tk.Frame(panel, bg=COLORS['separator'], height=1)
        separator2.pack(fill=tk.X, padx=8, pady=8)
        
        # 预设主题区域
        theme_label = tk.Label(
            panel,
            text='预设主题',
            font=('SF Pro Text', 10),
            bg=COLORS['panel_bg'],
            fg=COLORS['text_secondary'],
            anchor='w'
        )
        theme_label.pack(fill=tk.X, padx=8, pady=(0, 4))
        
        # 预设主题网格 (2列4行)
        self.left_preset_grid = tk.Frame(panel, bg=COLORS['panel_bg'])
        self.left_preset_grid.pack(fill=tk.X, padx=4, pady=(0, 6))
        
        # 初始化预设主题按钮
        self.update_left_preset_display()
        
        # 自动保存预设勾选框
        self.auto_save_preset_var = tk.BooleanVar(value=True)  # 默认勾选
        auto_save_check = tk.Checkbutton(
            panel,
            text='导出时自动保存预设',
            variable=self.auto_save_preset_var,
            font=('SF Pro Text', 9),
            bg=COLORS['panel_bg'],
            fg=COLORS['text_secondary'],
            selectcolor=COLORS['bg_tertiary'],
            activebackground=COLORS['panel_bg'],
            activeforeground=COLORS['text_primary'],
            cursor='hand2'
        )
        auto_save_check.pack(fill=tk.X, padx=8, pady=(0, 8))
        
        return panel
    
    def create_center_panel(self, parent):
        """创建中间画布面板 - 现代深色风格"""
        panel = tk.Frame(
            parent,
            bg=COLORS['bg'],
            relief=tk.FLAT,
            bd=0,
            highlightthickness=0
        )
        
        # 工具栏 - 现代深色风格
        toolbar = tk.Frame(panel, bg=COLORS['bg_secondary'], height=44)
        toolbar.pack(fill=tk.X, padx=0, pady=0)
        
        # 左侧按钮 - 使用Label替代Button
        left_buttons = tk.Frame(toolbar, bg=COLORS['bg_secondary'])
        left_buttons.pack(side=tk.LEFT, padx=8, pady=6)
        
        btn_undo = tk.Label(
            left_buttons,
            text='↶ 撤销',
            font=('SF Pro Text', 10),
            bg=COLORS['bg_tertiary'],
            fg=COLORS['text_primary'],
            padx=12,
            pady=6,
            cursor='hand2'
        )
        btn_undo.pack(side=tk.LEFT, padx=(0, 4))
        btn_undo.bind('<Button-1>', lambda e: self.undo())
        btn_undo.bind('<Enter>', lambda e: btn_undo.config(bg=COLORS['hover']))
        btn_undo.bind('<Leave>', lambda e: btn_undo.config(bg=COLORS['bg_tertiary']))
        
        btn_redo = tk.Label(
            left_buttons,
            text='↷ 重做',
            font=('SF Pro Text', 10),
            bg=COLORS['bg_tertiary'],
            fg=COLORS['text_primary'],
            padx=12,
            pady=6,
            cursor='hand2'
        )
        btn_redo.pack(side=tk.LEFT, padx=(0, 4))
        btn_redo.bind('<Button-1>', lambda e: self.redo())
        btn_redo.bind('<Enter>', lambda e: btn_redo.config(bg=COLORS['hover']))
        btn_redo.bind('<Leave>', lambda e: btn_redo.config(bg=COLORS['bg_tertiary']))
        
        btn_delete = tk.Label(
            left_buttons,
            text='🗑️ 删除',
            font=('SF Pro Text', 10),
            bg=COLORS['bg_tertiary'],
            fg=COLORS['danger'],
            padx=12,
            pady=6,
            cursor='hand2'
        )
        btn_delete.pack(side=tk.LEFT)
        btn_delete.bind('<Button-1>', lambda e: self.delete_selected_sticker())
        btn_delete.bind('<Enter>', lambda e: btn_delete.config(bg=COLORS['hover']))
        btn_delete.bind('<Leave>', lambda e: btn_delete.config(bg=COLORS['bg_tertiary']))
        
        # 右侧导出按钮 - 使用Label替代Button
        right_buttons = tk.Frame(toolbar, bg=COLORS['bg_secondary'])
        right_buttons.pack(side=tk.RIGHT, padx=8, pady=6)
        
        btn_export = tk.Label(
            right_buttons,
            text='💾 导出',
            bg=COLORS['success'],
            fg=COLORS['text_bright'],
            font=('SF Pro Text', 10, 'bold'),
            padx=14,
            pady=6,
            cursor='hand2'
        )
        btn_export.pack(side=tk.RIGHT)
        btn_export.bind('<Button-1>', lambda e: self.export_image())
        btn_export.bind('<Enter>', lambda e: btn_export.config(bg='#28A745'))
        btn_export.bind('<Leave>', lambda e: btn_export.config(bg=COLORS['success']))
        
        # 画布 - 按比例计算显示尺寸（占可用空间的90%）
        preset = self.current_size_preset
        # 获取窗口尺寸，计算可用空间（减去左右面板）
        window_width = self.winfo_screenwidth() * 0.85  # 窗口宽度
        window_height = self.winfo_screenheight() * 0.85  # 窗口高度
        # 左面板约120px，右面板约320px，工具栏约50px，边距约40px
        available_width = int((window_width - 120 - 320 - 40) * 0.9)
        available_height = int((window_height - 50 - 40) * 0.9)
        
        ratio = preset['width'] / preset['height']
        
        if ratio > available_width / available_height:
            # 宽图，以宽度为准
            display_width = available_width
            display_height = int(display_width / ratio)
        else:
            # 高图或方图，以高度为准
            display_height = available_height
            display_width = int(display_height * ratio)
        
        self.canvas_widget = CanvasWidget(
            panel,
            width=display_width,
            height=display_height
        )
        self.canvas_widget.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        return panel
    
    def create_right_panel(self, parent):
        """创建右侧面板 - 现代深色风格"""
        panel = tk.Frame(
            parent,
            bg=COLORS['panel_bg'],
            width=240,
            relief=tk.FLAT,
            bd=0,
            highlightthickness=0
        )
        
        # 配置 ttk 标签页样式 (备用)
        style = ttk.Style()
        style.theme_use('default')
        
        # --- 自定义两行标签页实现 ---
        # 标签页定义: (id, emoji, name)
        self.tab_definitions = [
            # 第一行 (Row 0)
            ('background', '🎨', '背景'),
            ('border', '🖼️', '边框'),
            ('sticker', '✨', '贴纸'),
            ('text', '🔤', '文字'),
            # 第二行 (Row 1)
            ('basic', '📐', '编辑'),
            ('batch', '⚡', '批量'),
            ('layer', '📚', '图层'),
            ('history', '📝', '记录'),
        ]
        
        # 标签页容器
        self.tab_header_frame = tk.Frame(panel, bg=COLORS['panel_bg'])
        self.tab_header_frame.pack(fill=tk.X, padx=0, pady=0)
        
        # 两行标签按钮
        self.tab_row_frames = [
            tk.Frame(self.tab_header_frame, bg=COLORS['panel_bg']),
            tk.Frame(self.tab_header_frame, bg=COLORS['panel_bg'])
        ]
        
        self.tab_buttons = {}
        self.tab_frames = {}
        self.current_tab_id = 'background'
        self.current_active_row = 0
        
        # 创建标签按钮
        for i, (tab_id, emoji, name) in enumerate(self.tab_definitions):
            row = i // 4  # 0-3 在第一行, 4-7 在第二行
            
            btn = tk.Label(
                self.tab_row_frames[row],
                text=f'{emoji} {name}',
                font=('SF Pro Text', 9),
                bg=COLORS['bg_tertiary'],
                fg=COLORS['text_secondary'],
                padx=6, pady=4,
                cursor='hand2'
            )
            btn.pack(side=tk.LEFT, padx=1, pady=2)
            btn.bind('<Button-1>', lambda e, tid=tab_id: self.switch_tab(tid))
            self.tab_buttons[tab_id] = btn
        
        # 内容容器
        self.tab_content_frame = tk.Frame(panel, bg=COLORS['panel_bg'])
        self.tab_content_frame.pack(fill=tk.BOTH, expand=True)
        
        # 创建各标签页内容Frame
        for tab_id, _, _ in self.tab_definitions:
            frame = tk.Frame(self.tab_content_frame, bg=COLORS['panel_bg'])
            self.tab_frames[tab_id] = frame
        
        # 初始化各标签页内容
        self.create_background_tab(self.tab_frames['background'])
        self.create_border_tab(self.tab_frames['border'])
        self.create_sticker_tab(self.tab_frames['sticker'])
        self.create_text_tab(self.tab_frames['text'])
        self.create_basic_tools_tab(self.tab_frames['basic'])
        self.create_batch_tab(self.tab_frames['batch'])
        self.create_layer_tab(self.tab_frames['layer'])
        self.create_history_tab(self.tab_frames['history'])
        
        # 初始显示
        self._update_tab_rows()
        self.switch_tab('background')
        
        # 绑定 Tab 键切换标签
        self.bind('<Tab>', self.next_tab)
        
        return panel
    
    def switch_tab(self, tab_id):
        """切换标签页"""
        if tab_id == self.current_tab_id:
            return
        
        # 更新当前标签页
        self.current_tab_id = tab_id
        
        # 判断激活的是哪一行
        tab_index = [t[0] for t in self.tab_definitions].index(tab_id)
        new_active_row = tab_index // 4
        
        # 如果激活行变化，需要交换行顺序
        if new_active_row != self.current_active_row:
            self.current_active_row = new_active_row
            self._update_tab_rows()
        
        # 更新按钮样式
        for tid, btn in self.tab_buttons.items():
            if tid == tab_id:
                btn.config(bg=COLORS['panel_bg'], fg=COLORS['accent'])
            else:
                btn.config(bg=COLORS['bg_tertiary'], fg=COLORS['text_secondary'])
        
        # 隐藏所有内容，显示当前内容
        for tid, frame in self.tab_frames.items():
            frame.pack_forget()
        self.tab_frames[tab_id].pack(fill=tk.BOTH, expand=True)
        
        # 如果是历史记录Tab，刷新显示
        if tab_id == 'history':
            self.update_history_display()
    
    def _update_tab_rows(self):
        """更新标签行顺序：激活行在下面"""
        for row_frame in self.tab_row_frames:
            row_frame.pack_forget()
        
        if self.current_active_row == 0:
            # Row 1 在上，Row 0 在下
            self.tab_row_frames[1].pack(fill=tk.X)
            self.tab_row_frames[0].pack(fill=tk.X)
        else:
            # Row 0 在上，Row 1 在下
            self.tab_row_frames[0].pack(fill=tk.X)
            self.tab_row_frames[1].pack(fill=tk.X)
    
    def create_basic_tools_tab(self, parent):
        """基础工具标签页 - 现代风格"""
        # 创建滚动区域
        scroll_canvas = tk.Canvas(parent, bg=COLORS['panel_bg'], highlightthickness=0)
        scrollbar = tk.Scrollbar(parent, orient='vertical', command=scroll_canvas.yview)
        scroll_frame = tk.Frame(scroll_canvas, bg=COLORS['panel_bg'])
        
        scroll_frame.bind('<Configure>', lambda e: scroll_canvas.configure(scrollregion=scroll_canvas.bbox('all')))
        scroll_canvas.create_window((0, 0), window=scroll_frame, anchor='nw')
        scroll_canvas.configure(yscrollcommand=scrollbar.set)
        
        scroll_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 绑定鼠标滚轮
        self.bind_mousewheel(scroll_frame, scroll_canvas)
        
        # 图片操作标题
        tk.Label(
            scroll_frame, text='🔄 变换操作', font=('SF Pro Display', 13, 'bold'),
            bg=COLORS['panel_bg'], fg=COLORS['text_primary'], anchor='w'
        ).pack(fill=tk.X, padx=16, pady=(16, 8))
        
        transform_frame = tk.Frame(scroll_frame, bg=COLORS['panel_bg'])
        transform_frame.pack(fill=tk.X, padx=12, pady=(0, 12))
        
        transforms = [
            ('↺ 逆时针90°', lambda: self.apply_transform('rotate', -90)),
            ('↻ 顺时针90°', lambda: self.apply_transform('rotate', 90)),
            ('⇄ 水平翻转', lambda: self.apply_transform('flip_h')),
            ('⇅ 垂直翻转', lambda: self.apply_transform('flip_v')),
        ]
        
        for text, command in transforms:
            btn = tk.Label(
                transform_frame, text=text, bg=COLORS['bg_tertiary'], fg=COLORS['text_primary'],
                font=('SF Pro Text', 11), pady=10, cursor='hand2'
            )
            btn.pack(fill=tk.X, padx=4, pady=2)
            btn.bind('<Button-1>', lambda e, c=command: c())
            btn.bind('<Enter>', lambda e, b=btn: b.config(bg=COLORS['hover']))
            btn.bind('<Leave>', lambda e, b=btn: b.config(bg=COLORS['bg_tertiary']))
        
        # 滤镜标题
        tk.Label(
            scroll_frame, text='🎨 滤镜效果', font=('SF Pro Display', 13, 'bold'),
            bg=COLORS['panel_bg'], fg=COLORS['text_primary'], anchor='w'
        ).pack(fill=tk.X, padx=16, pady=(12, 8))
        
        filter_frame = tk.Frame(scroll_frame, bg=COLORS['panel_bg'])
        filter_frame.pack(fill=tk.X, padx=12, pady=(0, 12))
        
        filters = [
            ('🖤 黑白', 'grayscale'),
            ('🔆 锐化', 'sharpen'),
            ('🌫️ 模糊', 'blur'),
            ('✨ 平滑', 'smooth'),
            ('📐 轮廓', 'contour'),
            ('🗿 浮雕', 'emboss'),
        ]
        
        filter_grid = tk.Frame(filter_frame, bg=COLORS['panel_bg'])
        filter_grid.pack(fill=tk.X)
        
        # 初始化滤镜状态和按钮引用
        self.active_filters = set()  # 当前激活的滤镜
        self.filter_buttons = {}  # 按钮引用
        self.filter_base_texts = {}  # 原始文本
        
        for idx, (text, filter_type) in enumerate(filters):
            row, col = idx // 2, idx % 2
            btn = tk.Label(
                filter_grid, text=text, bg=COLORS['bg_tertiary'], fg=COLORS['text_primary'],
                font=('SF Pro Text', 10), pady=8, padx=8, cursor='hand2', width=10
            )
            btn.grid(row=row, column=col, padx=2, pady=2, sticky='ew')
            btn.bind('<Button-1>', lambda e, f=filter_type, b=btn: self.toggle_filter(f, b))
            btn.bind('<Enter>', lambda e, b=btn: b.config(bg=COLORS['hover']) if b.cget('bg') != COLORS['accent'] else None)
            btn.bind('<Leave>', lambda e, b=btn, f=filter_type: b.config(bg=COLORS['accent'] if f in self.active_filters else COLORS['bg_tertiary']))
            
            self.filter_buttons[filter_type] = btn
            self.filter_base_texts[filter_type] = text
        
        filter_grid.columnconfigure(0, weight=1)
        filter_grid.columnconfigure(1, weight=1)
        
        # 调整标题
        tk.Label(
            scroll_frame, text='⚡ 图片调整', font=('SF Pro Display', 13, 'bold'),
            bg=COLORS['panel_bg'], fg=COLORS['text_primary'], anchor='w'
        ).pack(fill=tk.X, padx=16, pady=(12, 8))
        
        adjust_frame = tk.Frame(scroll_frame, bg=COLORS['panel_bg'])
        adjust_frame.pack(fill=tk.X, padx=16, pady=(0, 12))
        
        # 亮度行
        brightness_row = tk.Frame(adjust_frame, bg=COLORS['panel_bg'])
        brightness_row.pack(fill=tk.X, pady=(0, 8))
        
        tk.Label(brightness_row, text='亮度', font=('SF Pro Text', 10), width=6,
            bg=COLORS['panel_bg'], fg=COLORS['text_secondary'], anchor='w').pack(side=tk.LEFT)
        
        self.brightness_scale = tk.Scale(
            brightness_row, from_=0.2, to=2.0, resolution=0.1, orient=tk.HORIZONTAL,
            bg=COLORS['bg_tertiary'], fg=COLORS['text_primary'], highlightthickness=0,
            troughcolor=COLORS['separator'], length=150, showvalue=True
        )
        self.brightness_scale.set(1.0)
        self.brightness_scale.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.brightness_scale.bind('<ButtonRelease-1>', lambda e: self.apply_adjustment('brightness'))
        
        brightness_reset = tk.Label(brightness_row, text='⟲', font=('SF Pro Text', 12),
            bg=COLORS['bg_tertiary'], fg=COLORS['text_primary'], padx=6, pady=2, cursor='hand2')
        brightness_reset.pack(side=tk.RIGHT, padx=(4, 0))
        brightness_reset.bind('<Button-1>', lambda e: self.reset_single_adjustment('brightness'))
        brightness_reset.bind('<Enter>', lambda e: brightness_reset.config(bg=COLORS['hover']))
        brightness_reset.bind('<Leave>', lambda e: brightness_reset.config(bg=COLORS['bg_tertiary']))
        
        # 对比度行
        contrast_row = tk.Frame(adjust_frame, bg=COLORS['panel_bg'])
        contrast_row.pack(fill=tk.X, pady=(0, 8))
        
        tk.Label(contrast_row, text='对比度', font=('SF Pro Text', 10), width=6,
            bg=COLORS['panel_bg'], fg=COLORS['text_secondary'], anchor='w').pack(side=tk.LEFT)
        
        self.contrast_scale = tk.Scale(
            contrast_row, from_=0.2, to=2.0, resolution=0.1, orient=tk.HORIZONTAL,
            bg=COLORS['bg_tertiary'], fg=COLORS['text_primary'], highlightthickness=0,
            troughcolor=COLORS['separator'], length=150, showvalue=True
        )
        self.contrast_scale.set(1.0)
        self.contrast_scale.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.contrast_scale.bind('<ButtonRelease-1>', lambda e: self.apply_adjustment('contrast'))
        
        contrast_reset = tk.Label(contrast_row, text='⟲', font=('SF Pro Text', 12),
            bg=COLORS['bg_tertiary'], fg=COLORS['text_primary'], padx=6, pady=2, cursor='hand2')
        contrast_reset.pack(side=tk.RIGHT, padx=(4, 0))
        contrast_reset.bind('<Button-1>', lambda e: self.reset_single_adjustment('contrast'))
        contrast_reset.bind('<Enter>', lambda e: contrast_reset.config(bg=COLORS['hover']))
        contrast_reset.bind('<Leave>', lambda e: contrast_reset.config(bg=COLORS['bg_tertiary']))
        
        # 饱和度行
        saturation_row = tk.Frame(adjust_frame, bg=COLORS['panel_bg'])
        saturation_row.pack(fill=tk.X, pady=(0, 8))
        
        tk.Label(saturation_row, text='饱和度', font=('SF Pro Text', 10), width=6,
            bg=COLORS['panel_bg'], fg=COLORS['text_secondary'], anchor='w').pack(side=tk.LEFT)
        
        self.saturation_scale = tk.Scale(
            saturation_row, from_=0.0, to=2.0, resolution=0.1, orient=tk.HORIZONTAL,
            bg=COLORS['bg_tertiary'], fg=COLORS['text_primary'], highlightthickness=0,
            troughcolor=COLORS['separator'], length=150, showvalue=True
        )
        self.saturation_scale.set(1.0)
        self.saturation_scale.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.saturation_scale.bind('<ButtonRelease-1>', lambda e: self.apply_adjustment('saturation'))
        
        saturation_reset = tk.Label(saturation_row, text='⟲', font=('SF Pro Text', 12),
            bg=COLORS['bg_tertiary'], fg=COLORS['text_primary'], padx=6, pady=2, cursor='hand2')
        saturation_reset.pack(side=tk.RIGHT, padx=(4, 0))
        saturation_reset.bind('<Button-1>', lambda e: self.reset_single_adjustment('saturation'))
        saturation_reset.bind('<Enter>', lambda e: saturation_reset.config(bg=COLORS['hover']))
        saturation_reset.bind('<Leave>', lambda e: saturation_reset.config(bg=COLORS['bg_tertiary']))
        
        # 重置按钮
        reset_btn = tk.Label(
            scroll_frame, text='🔄 重置图片', bg=COLORS['accent'], fg='white',
            font=('SF Pro Text', 11, 'bold'), pady=10, cursor='hand2'
        )
        reset_btn.pack(fill=tk.X, padx=16, pady=(8, 16))
        reset_btn.bind('<Button-1>', lambda e: self.reset_image_and_sliders())
        reset_btn.bind('<Enter>', lambda e: reset_btn.config(bg='#0066CC'))
        reset_btn.bind('<Leave>', lambda e: reset_btn.config(bg=COLORS['accent']))
    
    def create_decoration_tab(self, parent):
        """装饰标签页 - 现代风格"""
        # 贴纸部分
        sticker_label = tk.Label(
            parent,
            text='贴纸',
            font=('SF Pro Display', 13, 'bold'),
            bg=COLORS['panel_bg'],
            fg=COLORS['text_primary'],
            anchor='w'
        )
        sticker_label.pack(fill=tk.X, padx=16, pady=(16, 8))
        
        # 贴纸网格
        sticker_grid = tk.Frame(parent, bg=COLORS['panel_bg'])
        sticker_grid.pack(fill=tk.X, padx=12, pady=(0, 16))
        
        for idx, sticker in enumerate(STICKER_LIST):
            row = idx // 4
            col = idx % 4
            
            # 使用Label替代Button
            if sticker['id'] in self.sticker_images:
                # 使用PNG图片
                btn = tk.Label(
                    sticker_grid,
                    image=self.sticker_images[sticker['id']],
                    bg=COLORS['bg_tertiary'],
                    cursor='hand2'
                )
            else:
                # 使用emoji
                btn = tk.Label(
                    sticker_grid,
                    text=sticker['emoji'],
                    font=('Apple Color Emoji', 28),
                    bg=COLORS['bg_tertiary'],
                    width=2,
                    height=1,
                    cursor='hand2'
                )
            btn.grid(row=row, column=col, padx=4, pady=4)
            btn.bind('<Button-1>', lambda e, s=sticker: self.add_sticker(s))
            btn.bind('<Enter>', lambda e, b=btn: b.config(bg=COLORS['hover']))
            btn.bind('<Leave>', lambda e, b=btn: b.config(bg=COLORS['bg_tertiary']))
        
        # 分隔线
        separator = tk.Frame(parent, bg=COLORS['separator'], height=1)
        separator.pack(fill=tk.X, padx=16, pady=12)
        
        # 边框部分
        border_label = tk.Label(
            parent,
            text='边框',
            font=('SF Pro Display', 13, 'bold'),
            bg=COLORS['panel_bg'],
            fg=COLORS['text_primary'],
            anchor='w'
        )
        border_label.pack(fill=tk.X, padx=16, pady=(8, 8))
        
        # 边框分类选择
        category_frame = tk.Frame(parent, bg=COLORS['panel_bg'])
        category_frame.pack(fill=tk.X, padx=16, pady=(0, 8))
        
        self.border_category_buttons = {}
        for category_id, category_info in BORDER_CATEGORIES.items():
            is_selected = category_id == self.selected_border_category
            btn = tk.Label(
                category_frame,
                text=category_info['name'],
                bg=COLORS['selected_bg'] if is_selected else COLORS['bg_tertiary'],
                fg=COLORS['text_bright'] if is_selected else COLORS['text_primary'],
                font=('SF Pro Text', 10, 'bold' if is_selected else 'normal'),
                padx=12,
                pady=8,
                cursor='hand2'
            )
            btn.pack(side=tk.LEFT, padx=2)
            btn.bind('<Button-1>', lambda e, c=category_id: self.select_border_category(c))
            self.border_category_buttons[category_id] = btn
        
        # 颜色选择
        color_label = tk.Label(
            parent,
            text='颜色',
            font=('SF Pro Text', 11),
            bg=COLORS['panel_bg'],
            fg=COLORS['text_secondary'],
            anchor='w'
        )
        color_label.pack(fill=tk.X, padx=16, pady=(8, 4))
        
        color_grid = tk.Frame(parent, bg=COLORS['panel_bg'])
        color_grid.pack(fill=tk.X, padx=16, pady=(0, 8))
        
        self.border_color_buttons = {}
        current_category = BORDER_CATEGORIES[self.selected_border_category]
        for idx, color_id in enumerate(current_category['colors']):
            if color_id in BORDER_COLORS:
                color_info = BORDER_COLORS[color_id]
                row = idx // 6
                col = idx % 6
                
                # 使用Canvas替代Button显示颜色
                btn = tk.Canvas(
                    color_grid,
                    bg=color_info['preview'],
                    width=28,
                    height=20,
                    highlightthickness=2 if color_id == self.selected_border_color else 1,
                    highlightbackground=COLORS['accent'] if color_id == self.selected_border_color else COLORS['separator'],
                    cursor='hand2'
                )
                btn.grid(row=row, column=col, padx=3, pady=3)
                btn.bind('<Button-1>', lambda e, c=color_id: self.select_border_color(c))
                self.border_color_buttons[color_id] = btn
        
        # 边框样式网格（滚动）
        style_container = tk.Frame(parent, bg=COLORS['panel_bg'])
        style_container.pack(fill=tk.BOTH, expand=True, padx=12, pady=(0, 12))
        
        style_canvas = tk.Canvas(
            style_container,
            bg=COLORS['panel_bg'],
            highlightthickness=0,
            bd=0
        )
        style_scrollbar = tk.Scrollbar(
            style_container,
            orient='vertical',
            command=style_canvas.yview
        )
        self.border_style_frame = tk.Frame(style_canvas, bg=COLORS['panel_bg'])
        
        style_canvas.create_window((0, 0), window=self.border_style_frame, anchor='nw')
        style_canvas.configure(yscrollcommand=style_scrollbar.set)
        
        # 初始化边框样式显示
        self.update_border_styles_display()
        
        self.border_style_frame.update_idletasks()
        style_canvas.config(scrollregion=style_canvas.bbox('all'))
        
        style_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        style_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    
    def select_border_category(self, category_id):
        """选择边框分类"""
        self.selected_border_category = category_id
        
        # 更新分类按钮样式 (Label)
        for cat_id, btn in self.border_category_buttons.items():
            if cat_id == category_id:
                btn.config(
                    bg=COLORS['selected_bg'],
                    fg=COLORS['text_bright'],
                    font=('SF Pro Text', 10, 'bold')
                )
            else:
                btn.config(
                    bg=COLORS['bg_tertiary'],
                    fg=COLORS['text_primary'],
                    font=('SF Pro Text', 10)
                )
        
        # 更新边框样式显示
        self.update_border_styles_display()
    
    def select_border_color(self, color_id):
        """选择边框颜色"""
        self.selected_border_color = color_id
        
        # 同步更新 border_config 的颜色 (从 BORDER_COLORS 获取实际颜色值)
        from constants import BORDER_COLORS
        if color_id in BORDER_COLORS:
            self.border_config['color'] = BORDER_COLORS[color_id]['hex']
        
        # 更新颜色按钮样式
        for c_id, btn in self.border_color_buttons.items():
            if c_id == color_id:
                btn.config(highlightthickness=2, highlightbackground=COLORS['accent'])
            else:
                btn.config(highlightthickness=1, highlightbackground=COLORS['separator'])
        
        # 更新边框样式显示
        self.update_border_styles_display()
    
    def update_border_styles_display(self):
        """更新边框样式显示"""
        # 清空现有内容
        for widget in self.border_style_frame.winfo_children():
            widget.destroy()
        
        # 获取当前分类的样式
        category = BORDER_CATEGORIES[self.selected_border_category]
        
        for style in category['styles']:
            # 构建边框ID
            border_id = f"{self.selected_border_category}_{style}_{self.selected_border_color}"
            
            # 创建边框按钮
            btn_frame = tk.Frame(self.border_style_frame, bg='white', highlightthickness=1, highlightbackground=COLORS['separator'])
            btn_frame.pack(fill=tk.X, padx=4, pady=4)
            
            # 如果有预览图，显示预览图
            if border_id in self.border_preview_images:
                img_label = tk.Label(
                    btn_frame,
                    image=self.border_preview_images[border_id],
                    bg='white',
                    cursor='hand2'
                )
                img_label.pack(side=tk.LEFT, padx=8, pady=8)
                img_label.bind('<Button-1>', lambda e, bid=border_id: self.apply_border(bid))
            
            # 显示边框名称
            name = BORDER_STYLE_NAMES.get(style, style)
            color_name = BORDER_COLORS.get(self.selected_border_color, {}).get('name', '')
            full_name = f"{name} - {color_name}"
            
            name_label = tk.Label(
                btn_frame,
                text=full_name,
                font=('SF Pro Text', 11),
                bg='white',
                fg=COLORS['text_primary'],
                anchor='w',
                cursor='hand2'
            )
            name_label.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=8, pady=12)
            name_label.bind('<Button-1>', lambda e, bid=border_id: self.apply_border(bid))
            
            # 悬停效果
            def on_enter(e, frame=btn_frame):
                frame.config(bg=COLORS['hover'])
                for child in frame.winfo_children():
                    child.config(bg=COLORS['hover'])
            
            def on_leave(e, frame=btn_frame):
                frame.config(bg='white')
                for child in frame.winfo_children():
                    child.config(bg='white')
            
            btn_frame.bind('<Enter>', on_enter)
            btn_frame.bind('<Leave>', on_leave)
            for child in btn_frame.winfo_children():
                child.bind('<Enter>', on_enter)
                child.bind('<Leave>', on_leave)
    
    def apply_border(self, border_id):
        """应用边框 - 直接生效"""
        # 解析border_id: category_style_color
        parts = border_id.split('_')
        if len(parts) >= 3:
            category = parts[0]
            style = parts[1]
            color = parts[2]
            
            # 加载边框图片
            frames_dir = os.path.join(os.path.dirname(__file__), 'assets', 'borders', 'frames')
            border_path = os.path.join(frames_dir, f"{border_id}.png")
            
            if os.path.exists(border_path):
                try:
                    # 应用到画布
                    border_img = Image.open(border_path).convert('RGBA')
                    self.canvas_widget.apply_border_image(border_img)
                    print(f"✓ 边框已应用: {border_id}")
                except Exception as e:
                    print(f"应用边框失败: {e}")
            else:
                print(f"边框文件不存在: {border_path}")
    
    def create_text_tab(self, parent):
        """文字编辑标签页"""
        from image_processor import TextLayer
        
        # 滚动区域
        scroll_canvas = tk.Canvas(parent, bg=COLORS['panel_bg'], highlightthickness=0)
        scrollbar = tk.Scrollbar(parent, orient='vertical', command=scroll_canvas.yview)
        text_frame = tk.Frame(scroll_canvas, bg=COLORS['panel_bg'])
        
        text_frame.bind('<Configure>', lambda e: scroll_canvas.configure(scrollregion=scroll_canvas.bbox('all')))
        scroll_canvas.create_window((0, 0), window=text_frame, anchor='nw')
        scroll_canvas.configure(yscrollcommand=scrollbar.set)
        
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        scroll_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # 绑定滚轮
        self.bind_mousewheel(text_frame, scroll_canvas)
        
        # 1. 文字内容输入 (可调整大小)
        tk.Label(text_frame, text='📝 文字内容', font=('SF Pro Display', 12, 'bold'),
                 bg=COLORS['panel_bg'], fg=COLORS['text_primary']).pack(fill=tk.X, padx=12, pady=(12, 4))
        
        # 文本框容器
        text_entry_container = tk.Frame(text_frame, bg=COLORS['panel_bg'])
        text_entry_container.pack(anchor='w', padx=12, pady=(0, 8))
        
        self.text_content_entry = tk.Text(text_entry_container, height=4, width=24, font=('SF Pro Text', 10),
                                          bg=COLORS['bg_secondary'], fg=COLORS['text_primary'],
                                          insertbackground=COLORS['text_primary'],
                                          wrap=tk.WORD, highlightthickness=1, 
                                          highlightbackground=COLORS['separator'])
        self.text_content_entry.pack(side=tk.TOP, anchor='w')
        # 实时预览：每次按键更新画布
        self.text_content_entry.bind('<KeyRelease>', lambda e: self._on_text_preview())
        # 高亮检测：仅在换行或移出时触发
        self.text_content_entry.bind('<Return>', lambda e: self._on_detect_keywords())
        self.text_content_entry.bind('<FocusOut>', lambda e: self._on_detect_keywords())
        self._keyword_detect_job = None  # 用于防抖
        
        # 调整大小的手柄
        resize_handle = tk.Label(text_entry_container, text='⋮⋮', font=('SF Pro Text', 8),
                                 bg=COLORS['bg_tertiary'], fg=COLORS['text_secondary'],
                                 cursor='bottom_right_corner', padx=2, pady=0)
        resize_handle.pack(side=tk.RIGHT, anchor='se')
        
        # 拖拽调整大小
        def on_resize_drag(event):
            # 获取文本框当前位置
            entry_x = self.text_content_entry.winfo_x()
            entry_y = self.text_content_entry.winfo_y()
            # 计算新尺寸(相对于文本框左上角)
            new_w = max(15, (event.x_root - self.text_content_entry.winfo_rootx()) // 8)  # 字符宽度
            new_h = max(2, (event.y_root - self.text_content_entry.winfo_rooty()) // 16)   # 行高
            self.text_content_entry.config(width=new_w, height=new_h)
        
        resize_handle.bind('<B1-Motion>', on_resize_drag)
        
        # 字符计数器
        self.char_count_label = tk.Label(text_entry_container, text='0 / 150', font=('SF Pro Text', 9),
                                         bg=COLORS['panel_bg'], fg=COLORS['text_secondary'])
        self.char_count_label.pack(anchor='e', padx=4)
        
        # 关键词高亮 + 清除文字 (移到文字框下方)
        text_actions_frame = tk.Frame(text_entry_container, bg=COLORS['panel_bg'])
        text_actions_frame.pack(fill=tk.X, pady=(4, 0))
        
        self.highlight_enabled_var = tk.BooleanVar(value=True)
        tk.Checkbutton(text_actions_frame, text='🔍 自动高亮', variable=self.highlight_enabled_var,
                      bg=COLORS['panel_bg'], fg=COLORS['text_primary'],
                      selectcolor=COLORS['accent'], activebackground=COLORS['panel_bg'],
                      font=('SF Pro Text', 9),
                      command=self._on_highlight_toggle).pack(side=tk.LEFT)
        
        # 高亮颜色 (默认随机)
        self.highlight_color_var = tk.StringVar(value='random')
        # 用户要求删除切换颜色的方块，默认使用随机多巴胺/马卡龙色
        tk.Label(text_actions_frame, text='(随机糖果色)', font=('SF Pro Text', 9),
                bg=COLORS['panel_bg'], fg=COLORS['text_secondary']).pack(side=tk.LEFT, padx=2)

        
        # 存储自动检测的关键词 (内部使用)
        self._auto_keywords = []
        
        
        def _on_setting_release(action_name):
            self._auto_apply_text()
            self.save_history(action_name)
        
        # 2. 字体设置
        font_frame = tk.Frame(text_frame, bg=COLORS['panel_bg'])
        font_frame.pack(fill=tk.X, padx=12, pady=4)
        
        tk.Label(font_frame, text='字体:', font=('SF Pro Text', 10),
                 bg=COLORS['panel_bg'], fg=COLORS['text_secondary']).pack(side=tk.LEFT)
        
        from image_processor import TextLayer
        from tkinter import ttk
        
        font_map = TextLayer.FONT_NAMES
        font_values = list(font_map.values())
        default_font_name = font_map.get('pingfang', '苹方 (默认)')
        
        self.font_family_var = tk.StringVar(value=default_font_name)
        
        #样式调整
        style = ttk.Style()
        style.theme_use('default') 
        style.configure("TCombobox", fieldbackground=COLORS['bg_secondary'], background=COLORS['bg_secondary'], foreground='#333333')
        
        font_combo = ttk.Combobox(font_frame, textvariable=self.font_family_var, values=font_values, 
                                  state="readonly", width=12)
        font_combo.pack(side=tk.LEFT, padx=4)
        
        def on_font_change(event):
            self._auto_apply_text()
            self.save_history("切换字体")
            
        font_combo.bind('<<ComboboxSelected>>', on_font_change)
        
        # 3. 字号设置
        size_frame = tk.Frame(text_frame, bg=COLORS['panel_bg'])
        size_frame.pack(fill=tk.X, padx=12, pady=4)
        
        tk.Label(size_frame, text='字号:', font=('SF Pro Text', 10),
                 bg=COLORS['panel_bg'], fg=COLORS['text_secondary']).pack(side=tk.LEFT)
        
        self.font_size_var = tk.IntVar(value=48)
        self.font_size_scale = tk.Scale(size_frame, from_=12, to=120, orient=tk.HORIZONTAL,
                             variable=self.font_size_var, bg=COLORS['panel_bg'], 
                             fg=COLORS['text_primary'], highlightthickness=0,
                             troughcolor=COLORS['bg_secondary'], length=100,
                             command=lambda v: self.update_text_preview())
        self.font_size_scale.pack(side=tk.LEFT, padx=(8, 0))
        self.font_size_scale.bind('<ButtonRelease-1>', lambda e: _on_setting_release("设置字号"))
        
        self.font_size_label = tk.Label(size_frame, text='48', font=('SF Pro Text', 10),
                                        bg=COLORS['panel_bg'], fg=COLORS['text_primary'], width=4)
        self.font_size_label.pack(side=tk.LEFT)
        
        # 4. 颜色设置 - 扩展版
        color_section = tk.Frame(text_frame, bg=COLORS['panel_bg'])
        color_section.pack(fill=tk.X, padx=12, pady=4)
        
        tk.Label(color_section, text='颜色:', font=('SF Pro Text', 10),
                 bg=COLORS['panel_bg'], fg=COLORS['text_secondary']).pack(anchor='w')
        
        self.text_color_var = tk.StringVar(value='#333333')
        
        def _set_text_color_with_history(color):
            self.set_text_color(color)
            self.save_history("设置文字颜色")
        
        # 基础色
        basic_frame = tk.Frame(color_section, bg=COLORS['panel_bg'])
        basic_frame.pack(anchor='w', pady=2)
        basic_colors = ['#333333', '#000000', '#FFFFFF', '#FF2D55', '#FF9500', '#FFCC00', '#34C759', '#007AFF', '#5856D6']
        for c in basic_colors:
            cb = tk.Canvas(basic_frame, width=18, height=18, bg=c, highlightthickness=1,
                          highlightbackground=COLORS['separator'], cursor='hand2')
            cb.pack(side=tk.LEFT, padx=1)
            cb.bind('<Button-1>', lambda e, color=c: _set_text_color_with_history(color))
        
        # 马卡龙色
        from constants import MACARON_COLORS, DOPAMINE_COLORS
        macaron_frame = tk.Frame(color_section, bg=COLORS['panel_bg'])
        macaron_frame.pack(anchor='w', pady=2)
        for c in MACARON_COLORS[:9]:
            cb = tk.Canvas(macaron_frame, width=18, height=18, bg=c, highlightthickness=1,
                          highlightbackground=COLORS['separator'], cursor='hand2')
            cb.pack(side=tk.LEFT, padx=1)
            cb.bind('<Button-1>', lambda e, color=c: _set_text_color_with_history(color))
        
        # 多巴胺色
        dopamine_frame = tk.Frame(color_section, bg=COLORS['panel_bg'])
        dopamine_frame.pack(anchor='w', pady=2)
        for c in DOPAMINE_COLORS[:9]:
            cb = tk.Canvas(dopamine_frame, width=18, height=18, bg=c, highlightthickness=1,
                          highlightbackground=COLORS['separator'], cursor='hand2')
            cb.pack(side=tk.LEFT, padx=1)
            cb.bind('<Button-1>', lambda e, color=c: _set_text_color_with_history(color))
        
        # 自定义颜色按钮
        custom_btn_frame = tk.Frame(color_section, bg=COLORS['panel_bg'])
        custom_btn_frame.pack(anchor='w', pady=4)
        
        self.text_color_preview = tk.Canvas(custom_btn_frame, width=24, height=24, 
                                            bg='#333333', highlightthickness=1,
                                            highlightbackground=COLORS['separator'])
        self.text_color_preview.pack(side=tk.LEFT)
        
        custom_btn = tk.Label(custom_btn_frame, text='🎨 自定义', font=('SF Pro Text', 9),
                             bg=COLORS['bg_tertiary'], fg=COLORS['text_primary'], 
                             padx=6, pady=2, cursor='hand2')
        custom_btn.pack(side=tk.LEFT, padx=4)
        custom_btn.bind('<Button-1>', lambda e: self.open_text_color_picker())
        
        # 5. 对齐设置 (图标按钮)
        align_frame = tk.Frame(text_frame, bg=COLORS['panel_bg'])
        align_frame.pack(fill=tk.X, padx=12, pady=4)
        
        tk.Label(align_frame, text='对齐:', font=('SF Pro Text', 10),
                 bg=COLORS['panel_bg'], fg=COLORS['text_secondary']).pack(side=tk.LEFT)
        
        self.text_align_var = tk.StringVar(value='left')
        # 使用图标: ☰ (左对齐), ☰ (居中), ☰ (右对齐)
        align_icons = [('⬚≡', 'left'), ('≡', 'center'), ('≡⬚', 'right')]
        for icon, val in align_icons:
            btn = tk.Label(align_frame, text=icon, font=('SF Pro Text', 14),
                          bg=COLORS['bg_tertiary'], fg=COLORS['text_primary'],
                          padx=8, pady=2, cursor='hand2')
            btn.pack(side=tk.LEFT, padx=2)
            btn.bind('<Button-1>', lambda e, v=val: self._set_align_with_history(v))
        
        # 6. 位置设置 (图标按钮)
        pos_frame = tk.Frame(text_frame, bg=COLORS['panel_bg'])
        pos_frame.pack(fill=tk.X, padx=12, pady=4)
        
        tk.Label(pos_frame, text='位置:', font=('SF Pro Text', 10),
                 bg=COLORS['panel_bg'], fg=COLORS['text_secondary']).pack(side=tk.LEFT)
        
        self.text_position_var = tk.StringVar(value='top')
        # 使用图标: ⬆ (顶部), ⬌ (居中), ⬇ (底部)
        pos_icons = [('⬆', 'top'), ('⬌', 'center'), ('⬇', 'bottom')]
        for icon, val in pos_icons:
            btn = tk.Label(pos_frame, text=icon, font=('SF Pro Text', 14),
                          bg=COLORS['bg_tertiary'], fg=COLORS['text_primary'],
                          padx=8, pady=2, cursor='hand2')
            btn.pack(side=tk.LEFT, padx=2)
            btn.bind('<Button-1>', lambda e, v=val: self._set_position_with_history(v))
        
        # 6.1 文字样式 (加粗/斜体/下划线)
        style_frame = tk.Frame(text_frame, bg=COLORS['panel_bg'])
        style_frame.pack(fill=tk.X, padx=12, pady=4)
        
        tk.Label(style_frame, text='样式:', font=('SF Pro Text', 10),
                 bg=COLORS['panel_bg'], fg=COLORS['text_secondary']).pack(side=tk.LEFT)
        
        self.text_bold_var = tk.BooleanVar(value=False)
        self.text_italic_var = tk.BooleanVar(value=False)
        self.text_underline_var = tk.BooleanVar(value=False)
        
        # B = Bold, I = Italic, U = Underline
        style_btns = [('B', self.text_bold_var, 'bold'), 
                      ('I', self.text_italic_var, 'italic'), 
                      ('U̲', self.text_underline_var, 'underline')]
        for icon, var, name in style_btns:
            btn = tk.Checkbutton(style_frame, text=icon, variable=var,
                                font=('SF Pro Text', 12, 'bold' if name == 'bold' else 'italic' if name == 'italic' else 'normal'),
                                bg=COLORS['bg_tertiary'], fg=COLORS['text_primary'],
                                selectcolor=COLORS['accent'], activebackground=COLORS['bg_tertiary'],
                                indicatoron=False, padx=8, pady=2,
                                command=lambda: self._apply_style_with_history("切换文字样式"))
            btn.pack(side=tk.LEFT, padx=2)
            
        # 首行缩进 (New)
        self.text_indent_var = tk.BooleanVar(value=True) # 默认开启
        indent_cb = tk.Checkbutton(style_frame, text="首行缩进", variable=self.text_indent_var,
                                  font=('SF Pro Text', 10), bg=COLORS['panel_bg'], fg=COLORS['text_primary'],
                                  selectcolor=COLORS['accent'], activebackground=COLORS['panel_bg'],
                                  command=lambda: self._apply_style_with_history("切换缩进"))
        indent_cb.pack(side=tk.LEFT, padx=8)
        
        # 7. 边距设置
        margin_frame = tk.Frame(text_frame, bg=COLORS['panel_bg'])
        margin_frame.pack(fill=tk.X, padx=12, pady=4)
        
        tk.Label(margin_frame, text='边距:', font=('SF Pro Text', 10),
                 bg=COLORS['panel_bg'], fg=COLORS['text_secondary']).pack(side=tk.LEFT)
        
        self.text_margin_var = tk.IntVar(value=20)
        margin_scale = tk.Scale(margin_frame, from_=0, to=100, orient=tk.HORIZONTAL,
                               variable=self.text_margin_var, bg=COLORS['panel_bg'],
                               fg=COLORS['text_primary'], highlightthickness=0,
                               troughcolor=COLORS['bg_secondary'], length=80,
                               command=lambda v: self.update_text_preview())
        margin_scale.pack(side=tk.LEFT, padx=(8, 0))
        margin_scale.bind('<ButtonRelease-1>', lambda e: _on_setting_release("设置文字边距"))
        
        # 8. 阴影设置 (紧凑布局)
        shadow_frame = tk.Frame(text_frame, bg=COLORS['panel_bg'])
        shadow_frame.pack(fill=tk.X, padx=12, pady=2)
        
        self.text_shadow_var = tk.BooleanVar(value=False)
        tk.Checkbutton(shadow_frame, text='阴影', variable=self.text_shadow_var,
                      bg=COLORS['panel_bg'], fg=COLORS['text_primary'],
                      selectcolor=COLORS['accent'], activebackground=COLORS['panel_bg'],
                      font=('SF Pro Text', 10),
                      command=lambda: self._apply_style_with_history("切换阴影")).pack(side=tk.LEFT)
        
        # 9. 描边设置 (紧凑布局，同一行)
        self.text_stroke_var = tk.BooleanVar(value=False)
        tk.Checkbutton(shadow_frame, text='描边', variable=self.text_stroke_var,
                      bg=COLORS['panel_bg'], fg=COLORS['text_primary'],
                      selectcolor=COLORS['accent'], activebackground=COLORS['panel_bg'],
                      font=('SF Pro Text', 10),
                      command=lambda: self._apply_style_with_history("切换描边")).pack(side=tk.LEFT, padx=(12, 0))
        
        # 描边宽度滑块
        stroke_frame = tk.Frame(text_frame, bg=COLORS['panel_bg'])
        stroke_frame.pack(fill=tk.X, padx=12, pady=2)
        
        tk.Label(stroke_frame, text='宽度:', font=('SF Pro Text', 9),
                 bg=COLORS['panel_bg'], fg=COLORS['text_secondary']).pack(side=tk.LEFT)
        
        self.stroke_width_var = tk.IntVar(value=2)
        stroke_scale = tk.Scale(stroke_frame, from_=1, to=10, orient=tk.HORIZONTAL,
                               variable=self.stroke_width_var, bg=COLORS['panel_bg'],
                               fg=COLORS['text_primary'], highlightthickness=0,
                               troughcolor=COLORS['bg_secondary'], length=60,
                               command=lambda v: self.update_text_preview())
        stroke_scale.pack(side=tk.LEFT, padx=(4, 0))
        stroke_scale.bind('<ButtonRelease-1>', lambda e: _on_setting_release("设置描边宽度"))
        
        # 描边颜色 (同一行，9个颜色)
        self.stroke_color_var = tk.StringVar(value='#000000')
        stroke_colors = ['#000000', '#FFFFFF', '#FF2D55', '#FF9500', '#FFCC00', '#34C759', '#007AFF', '#5856D6', '#AF52DE']
        for c in stroke_colors:
            sc = tk.Canvas(stroke_frame, width=14, height=14, bg=c, highlightthickness=1,
                          highlightbackground=COLORS['separator'], cursor='hand2')
            sc.pack(side=tk.LEFT, padx=1)
            sc.bind('<Button-1>', lambda e, color=c: self._set_stroke_color_with_history(color))
        
        # 清除文字按钮 (放在面板底部，避免误点)
        clear_frame = tk.Frame(text_frame, bg=COLORS['panel_bg'])
        clear_frame.pack(fill=tk.X, padx=12, pady=(16, 4))
        
        clear_btn = tk.Label(clear_frame, text='🗑️ 清除文字', font=('SF Pro Text', 10),
                            bg=COLORS['bg_tertiary'], fg=COLORS['danger'], 
                            padx=10, pady=4, cursor='hand2')
        clear_btn.pack(side=tk.RIGHT)
        clear_btn.bind('<Button-1>', lambda e: self.clear_text_layers())

    
    def _on_text_preview(self):
        """实时预览：每次按键时更新画布（不触发关键词检测）"""
        # 更新字符计数
        if hasattr(self, 'text_content_entry') and hasattr(self, 'char_count_label'):
            content = self.text_content_entry.get('1.0', 'end-1c')
            char_count = len(content)
            max_chars = 150
            
            # 限制最大字符数
            if char_count > max_chars:
                # 截断文本
                self.text_content_entry.delete('1.0', tk.END)
                self.text_content_entry.insert('1.0', content[:max_chars])
                char_count = max_chars
            
            # 更新计数显示
            color = COLORS['danger'] if char_count >= max_chars else COLORS['text_secondary']
            self.char_count_label.config(text=f'{char_count} / {max_chars}', fg=color)
        
        self._auto_apply_text()
        
        # 如果启用了自动高亮，延时触发关键词检测 (Debounce 800ms)
        if hasattr(self, 'highlight_enabled_var') and self.highlight_enabled_var.get():
            if hasattr(self, '_highlight_timer') and self._highlight_timer:
                self.after_cancel(self._highlight_timer)
            self._highlight_timer = self.after(800, self._auto_detect_silent)
    
    def _set_align(self, val):
        """设置对齐方式"""
        self.text_align_var.set(val)
        self._auto_apply_text()
        
    def _set_align_with_history(self, val):
        self._set_align(val)
        self.save_history(f"设置文字对齐")
    
    def _set_position(self, val):
        """设置位置"""
        self.text_position_var.set(val)
        self._auto_apply_text()
        
    def _set_position_with_history(self, val):
        self._set_position(val)
        self.save_history(f"设置文字位置")
        
    def _set_stroke_color(self, color):
        self.stroke_color_var.set(color)
        self.update_text_preview()
    def _set_stroke_color_with_history(self, color):
        self._set_stroke_color(color)
        self.save_history(f"设置描边颜色")
    
    def _set_stroke_color(self, color):
        """设置描边颜色"""
        self.stroke_color_var.set(color)
        self._auto_apply_text()
    
    def _on_detect_keywords(self):
        """仅在换行或移出时触发关键词检测"""
        self._auto_detect_silent()
        self.save_history("编辑文字内容")
    
    def _on_highlight_toggle(self):
        """切换自动高亮"""
        # 如果启用了高亮，先检测关键词
        if self.highlight_enabled_var.get():
            self._auto_detect_silent()
        self.update_text_preview()
        self.save_history("编辑文字内容")
    
    def _auto_detect_silent(self):
        """静默自动检测关键词并自动应用到画布"""
        import re
        if not hasattr(self, 'text_content_entry'):
            return
        
        content = self.text_content_entry.get('1.0', 'end-1c')
        if not content.strip() or len(content.strip()) < 2:
            self._auto_keywords = []
            return
        
        keywords = []
        
        # 使用 jieba 关键词提取
        try:
            import jieba.analyse
            jieba_keywords = jieba.analyse.extract_tags(content, topK=5, withWeight=False)
            keywords.extend(jieba_keywords)
        except:
            pass
        
        # 检测英文单词
        english_words = re.findall(r'[a-zA-Z]{2,}', content)
        for word in english_words:
            if word.lower() not in [k.lower() for k in keywords]:
                keywords.append(word)
        
        # 检测 #标签
        hashtags = re.findall(r'#\w+', content)
        for tag in hashtags:
            cleaned = tag.lstrip('#')
            if cleaned not in keywords:
                keywords.append(cleaned)
        
        # 存储关键词并自动应用到画布
        self._auto_keywords = list(dict.fromkeys(keywords))[:8]
        self._auto_apply_text()
    
    def _on_highlight_toggle(self):
        """高亮开关切换时触发"""
        # 1. 获取当前开关状态
        enabled = self.highlight_enabled_var.get()
        
        # 2. 如果开启，立即执行一次完整的关键词检测 (不使用静默方法，确保拿到结果)
        if enabled:
            content = self.text_content_entry.get('1.0', 'end-1c')
            keywords = []
            if content and len(content.strip()) >= 1:
                # 提取关键词
                try:
                    import jieba.analyse
                    keywords.extend(jieba.analyse.extract_tags(content, topK=5))
                except:
                    pass
                
                # 英文和标签
                import re
                keywords.extend(re.findall(r'[a-zA-Z]{2,}', content))
                keywords.extend([t.lstrip('#') for t in re.findall(r'#\w+', content)])
            
            # 去重并保存
            self._auto_keywords = list(dict.fromkeys(keywords))[:8]
        else:
            # 关闭时清空
            self._auto_keywords = []
            
        # 3. 强制重新应用文字 (直接调用应用方法，不走 preview 的 timer 逻辑)
        print(f"[DEBUG] Toggle Highlight: {enabled}, Keywords: {self._auto_keywords}")
        self._auto_apply_text()
        self.save_history("切换自动高亮")
    
    def _apply_style_with_history(self, action_name="调整文字样式"):
        """应用文字样式并保存历史"""
        self._auto_apply_text()
        self.save_history(action_name)

    def _auto_apply_text(self):
        """自动应用文字到画布"""
        from image_processor import TextLayer
        
        content = self.text_content_entry.get('1.0', 'end-1c').strip() if hasattr(self, 'text_content_entry') else ''
        if not content:
            self.clear_text_layers()
            return

        # 获取字体键名 (反向查找)
        font_name = self.font_family_var.get() if hasattr(self, 'font_family_var') else '苹方 (默认)'
        font_family = 'pingfang'
        found = False
        for k, v in TextLayer.FONT_NAMES.items():
            if v == font_name:
                font_family = k
                found = True
                break
        
        # 检查是否需要保留自定义位置
        custom_pos = None
        if hasattr(self, 'current_text_layer') and self.current_text_layer:
            if getattr(self.current_text_layer, 'position', '') == 'custom':
                custom_pos = (self.current_text_layer.rel_x, self.current_text_layer.rel_y)
        
        # 创建文字层
        text_layer = TextLayer(
            content=content,
            font_size=self.font_size_var.get() if hasattr(self, 'font_size_var') else 48,
            color=self.text_color_var.get() if hasattr(self, 'text_color_var') else '#333333',
            font_family=font_family,
            align=self.text_align_var.get() if hasattr(self, 'text_align_var') else 'left',
            position='custom' if custom_pos else (self.text_position_var.get() if hasattr(self, 'text_position_var') else 'top'),
            margin=self.text_margin_var.get() if hasattr(self, 'text_margin_var') else 20,
            shadow={
                'enabled': self.text_shadow_var.get() if hasattr(self, 'text_shadow_var') else False,
                'color': '#000000',
                'offset': (2, 2),
                'blur': 4
            },
            stroke={
                'enabled': self.text_stroke_var.get() if hasattr(self, 'text_stroke_var') else False,
                'color': self.stroke_color_var.get() if hasattr(self, 'stroke_color_var') else '#000000',
                'width': self.stroke_width_var.get() if hasattr(self, 'stroke_width_var') else 2
            },
            highlight={
                'enabled': self.highlight_enabled_var.get() if hasattr(self, 'highlight_enabled_var') else True,
                'keywords': self._auto_keywords if hasattr(self, '_auto_keywords') else [],
                'color': self.highlight_color_var.get() if hasattr(self, 'highlight_color_var') else '#FFB7B2'
            },
            bold=self.text_bold_var.get() if hasattr(self, 'text_bold_var') else False,
            italic=self.text_italic_var.get() if hasattr(self, 'text_italic_var') else False,
            underline=self.text_underline_var.get() if hasattr(self, 'text_underline_var') else False,
            indent=self.text_indent_var.get() if hasattr(self, 'text_indent_var') else True
        )
        
        # 恢复自定义位置坐标
        if custom_pos:
            text_layer.rel_x, text_layer.rel_y = custom_pos
        
        # 存储并应用
        self.current_text_layer = text_layer
        
        # 预览时不写入 ImageProcessor，而是作为独立 Item 添加到 Canvas
        self.image_processor.clear_text_layers()
        
        # [WYSIWYG FIX] 预览应该模拟导出尺寸，然后缩小显示
        # 直接使用当前选中的预设对象，确保与导出逻辑一致
        preset_width = self.current_size_preset['width']
        preset_height = self.current_size_preset['height']
        
        # 画布显示尺寸
        cw = self.canvas_widget.width if self.canvas_widget.width > 10 else 800
        ch = self.canvas_widget.height if self.canvas_widget.height > 10 else 600
        
        # 计算从画布到导出的缩放比例 (和 batch_export 相同)
        preview_scale = preset_width / cw if cw > 0 else 1.0
        
        # 计算导出尺寸下的边框宽度
        export_border_width = 0
        if hasattr(self, 'border_config') and self.border_config.get('id') != 'none':
            export_border_width = int(self.border_config.get('width', 0) * preview_scale)
            export_border_width += int(10 * preview_scale)  # 额外边距
            
        print(f"[DEBUG] PREVIEW: border_width_raw={self.border_config.get('width')}, export_border_width={export_border_width}")
        
        # 强制刷新关键词 (如果是高亮模式且关键词为空)
        if self.highlight_enabled_var.get() and not self._auto_keywords:
             pass
        
        # [关键] 使用导出尺寸渲染，和导出时完全一致
        text_img, x, y = text_layer.render(preset_width, preset_height, scale=preview_scale, safe_margin_x=export_border_width)
        
        if text_img:
            # 缩小回预览尺寸
            display_scale = cw / preset_width
            new_w = int(text_img.width * display_scale)
            new_h = int(text_img.height * display_scale)
            if new_w > 0 and new_h > 0:
                from PIL import Image
                text_img = text_img.resize((new_w, new_h), Image.Resampling.LANCZOS)
                x = int(x * display_scale)
                y = int(y * display_scale)
            self.canvas_widget.add_text_layer_item(text_img, x, y)
    
    def on_text_transform(self, action, **kwargs):
        """处理文字层的交互变换"""
        if not hasattr(self, 'current_text_layer') or not self.current_text_layer:
            return
            
        if action == 'move':
            # 更新相对坐标
            x, y = kwargs.get('x'), kwargs.get('y')
            cw, ch = self.canvas_widget.width, self.canvas_widget.height
            
            if cw > 0 and ch > 0:
                self.current_text_layer.rel_x = x / cw
                self.current_text_layer.rel_y = y / ch
                # 标记为自定义位置
                self.current_text_layer.position = 'custom'
                
        elif action == 'scale':
            # 更新字号
            factor = kwargs.get('factor', 1.0)
            if hasattr(self, 'font_size_var'):
                current_size = self.font_size_var.get()
                new_size = max(12, min(500, int(current_size * factor)))
                if new_size != current_size:
                    self.font_size_var.set(new_size)
                    # 重新应用文字 (重新渲染)
                    self._auto_apply_text()
    
    def set_text_color(self, color):
        """设置文字颜色"""
        self.text_color_var.set(color)
        # 更新颜色预览
        if hasattr(self, 'text_color_preview'):
            self.text_color_preview.config(bg=color)
        
        # 必须调用 _auto_apply_text 以更新 current_text_layer (用于导出)
        # 并确保重绘
        self._auto_apply_text()
    
    def open_text_color_picker(self):
        """打开自定义颜色选择器"""
        from color_wheel_picker import ColorWheelPicker
        
        def on_color_selected(color):
            self.set_text_color(color)
            self.save_history("设置文字颜色")
        
        picker = ColorWheelPicker(
            self, 
            callback=on_color_selected,
            initial_color=self.text_color_var.get()
        )
    
    def set_highlight_color(self, color):
        """设置高亮颜色"""
        # print(f"[DEBUG] Set highlight color: {color}")
        self.highlight_color_var.set(color)
        # 高亮颜色改变也需要重新应用文字
        self._auto_apply_text()
    
    def auto_detect_keywords(self):
        """自动检测关键字 (使用 jieba 智能提取)"""
        import re
        if not hasattr(self, 'text_content_entry'):
            return
        
        content = self.text_content_entry.get('1.0', 'end-1c')
        if not content.strip():
            return
        
        keywords = []
        
        # 尝试使用 jieba 关键词提取
        try:
            import jieba.analyse
            # 使用 TF-IDF 提取关键词 (最多5个)
            jieba_keywords = jieba.analyse.extract_tags(content, topK=5, withWeight=False)
            keywords.extend(jieba_keywords)
        except ImportError:
            pass  # jieba 未安装，使用备用方案
        except Exception as e:
            print(f"[DEBUG] jieba 关键词提取失败: {e}")
        
        # 备用: 检测英文单词 (中文中的英文通常是品牌/专有名词)
        english_words = re.findall(r'[a-zA-Z]{2,}', content)
        for word in english_words:
            if word.lower() not in [k.lower() for k in keywords]:
                keywords.append(word)
        
        # 检测 #标签
        hashtags = re.findall(r'#\w+', content)
        for tag in hashtags:
            cleaned = tag.lstrip('#')
            if cleaned not in keywords:
                keywords.append(cleaned)
        
        # 去重并更新输入框
        unique_keywords = list(dict.fromkeys(keywords))[:8]  # 最多8个
        if hasattr(self, 'highlight_keywords_entry'):
            self.highlight_keywords_entry.delete(0, 'end')
            self.highlight_keywords_entry.insert(0, ','.join(unique_keywords))
            self.highlight_enabled_var.set(True)
            self._auto_apply_text()
            self.show_toast(f'检测到 {len(unique_keywords)} 个关键词')
    
    def update_text_preview(self):
        """更新文字预览 (实时)"""
        # 更新字号显示
        if hasattr(self, 'font_size_label'):
            self.font_size_label.config(text=str(self.font_size_var.get()))
        
        # 更新配置
        self.current_text_config = {
            'content': self.text_content_entry.get('1.0', tk.END).strip() if hasattr(self, 'text_content_entry') else '',
            'font_size': self.font_size_var.get() if hasattr(self, 'font_size_var') else 48,
            'color': self.text_color_var.get() if hasattr(self, 'text_color_var') else '#FFFFFF',
            'font_family': self.font_family_var.get() if hasattr(self, 'font_family_var') else 'pingfang',
            'align': self.text_align_var.get() if hasattr(self, 'text_align_var') else 'center',
            'position': self.text_position_var.get() if hasattr(self, 'text_position_var') else 'bottom',
            'margin': self.text_margin_var.get() if hasattr(self, 'text_margin_var') else 20,
            'indent': self.text_indent_var.get() if hasattr(self, 'text_indent_var') else True,
            'shadow': {
                'enabled': self.text_shadow_var.get() if hasattr(self, 'text_shadow_var') else True,
                'color': '#000000',
                'offset': (2, 2),
                'blur': 4
            },
            'stroke': {
                'enabled': self.text_stroke_var.get() if hasattr(self, 'text_stroke_var') else False,
                'color': '#000000',
                'width': self.stroke_width_var.get() if hasattr(self, 'stroke_width_var') else 2
            },
            'highlight': {
                'enabled': self.highlight_enabled_var.get() if hasattr(self, 'highlight_enabled_var') else False,
                'keywords': self._auto_keywords if hasattr(self, '_auto_keywords') else [],
                'color': self.highlight_color_var.get() if hasattr(self, 'highlight_color_var') else '#FFB7B2'
            },
            'bold': self.text_bold_var.get() if hasattr(self, 'text_bold_var') else False,
            'italic': self.text_italic_var.get() if hasattr(self, 'text_italic_var') else False,
            'underline': self.text_underline_var.get() if hasattr(self, 'text_underline_var') else False
        }
        
        # 刷新画布预览
        if hasattr(self, 'canvas_widget'):
            self.canvas_widget.set_text_preview(self.current_text_config)
    
    def apply_text_to_canvas(self):
        """应用文字到画布"""
        from image_processor import TextLayer
        
        content = self.text_content_entry.get('1.0', tk.END).strip()
        if not content:
            self.show_toast('请输入文字内容')
            return
        
        # 创建文字层
        # 获取高亮关键字列表
        keywords = []
        if hasattr(self, 'highlight_keywords_entry'):
            kw_text = self.highlight_keywords_entry.get().strip()
            if kw_text:
                keywords = [k.strip() for k in kw_text.split(',') if k.strip()]
        
        text_layer = TextLayer(
            content=content,
            font_size=self.font_size_var.get(),
            color=self.text_color_var.get(),
            font_family=self.font_family_var.get(),
            align=self.text_align_var.get(),
            position=self.text_position_var.get(),
            margin=self.text_margin_var.get(),
            shadow={
                'enabled': self.text_shadow_var.get(),
                'color': '#000000',
                'offset': (2, 2),
                'blur': 4
            },
            stroke={
                'enabled': self.text_stroke_var.get(),
                'color': '#000000',
                'width': self.stroke_width_var.get()
            },
            highlight={
                'enabled': self.highlight_enabled_var.get() if hasattr(self, 'highlight_enabled_var') else False,
                'keywords': keywords,
                'color': self.highlight_color_var.get() if hasattr(self, 'highlight_color_var') else '#FFB7B2'
            },
            bold=self.text_bold_var.get() if hasattr(self, 'text_bold_var') else False,
            italic=self.text_italic_var.get() if hasattr(self, 'text_italic_var') else False,
            underline=self.text_underline_var.get() if hasattr(self, 'text_underline_var') else False,
            indent=self.text_indent_var.get() if hasattr(self, 'text_indent_var') else True
        )
        
        self.text_layers = [text_layer]  # 目前只支持一个文字层
        self.canvas_widget.set_text_layer(text_layer)
        self.save_history('添加文字')
        self.show_toast('文字已应用')
    
    def clear_text_layers(self):
        """清除所有文字层 (带确认)"""
        # 检查是否有内容需要清除
        has_content = False
        if hasattr(self, 'text_content_entry'):
            content = self.text_content_entry.get('1.0', 'end-1c').strip()
            if content:
                has_content = True
        
        if not has_content:
            self.show_toast('没有文字需要清除')
            return
        
        # 第一次确认
        if not messagebox.askyesno('确认清除', '确定要清除所有文字内容吗？'):
            return
        
        # 第二次确认
        if not messagebox.askyesno('再次确认', '此操作不可撤销，确定要清除吗？'):
            return
        
        # 执行清除
        self.text_layers = []
        if hasattr(self, 'canvas_widget'):
            self.canvas_widget.clear_text_layer()
        # 清空输入框
        if hasattr(self, 'text_content_entry'):
            self.text_content_entry.delete('1.0', tk.END)
        # 更新字符计数
        if hasattr(self, 'char_count_label'):
            self.char_count_label.config(text='0 / 150', fg=COLORS['text_secondary'])
        self.show_toast('文字已清除')
    
    def create_batch_tab(self, parent):
        """批量处理标签页 - 现代风格"""
        # 滚动区域
        scroll_canvas = tk.Canvas(parent, bg=COLORS['panel_bg'], highlightthickness=0)
        scrollbar = tk.Scrollbar(parent, orient='vertical', command=scroll_canvas.yview)
        batch_frame = tk.Frame(scroll_canvas, bg=COLORS['panel_bg'])
        
        batch_frame.bind('<Configure>', lambda e: scroll_canvas.configure(scrollregion=scroll_canvas.bbox('all')))
        scroll_canvas.create_window((0, 0), window=batch_frame, anchor='nw')
        scroll_canvas.configure(yscrollcommand=scrollbar.set)
        
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        scroll_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # 1. 输入目录设置
        input_header_frame = tk.Frame(batch_frame, bg=COLORS['panel_bg'])
        input_header_frame.pack(fill=tk.X, padx=12, pady=(12, 4))
        
        tk.Label(input_header_frame, text='📁 输入目录', font=('SF Pro Display', 12, 'bold'),
                 bg=COLORS['panel_bg'], fg=COLORS['text_primary']).pack(side=tk.LEFT)
                 
        input_dir_btn = tk.Label(input_header_frame, text='选择', font=('SF Pro Text', 10),
                                 bg=COLORS['accent'], fg='white', padx=10, pady=4, cursor='hand2')
        input_dir_btn.pack(side=tk.LEFT, padx=(10, 0))
        input_dir_btn.bind('<Button-1>', lambda e: self.select_input_dir())
        
        # 打开目录按钮
        input_open_btn = tk.Label(input_header_frame, text='打开', font=('SF Pro Text', 10),
                                  bg=COLORS['bg_tertiary'], fg=COLORS['text_primary'], padx=10, pady=4, cursor='hand2')
        input_open_btn.pack(side=tk.LEFT, padx=(4, 0))
        input_open_btn.bind('<Button-1>', lambda e: self.open_directory(self.batch_input_dir))
        
        self.input_dir_label = tk.Label(batch_frame, text=self.batch_input_dir or '未设置',
                                        font=('SF Pro Text', 9), bg=COLORS['bg_secondary'],
                                        fg=COLORS['text_secondary'], anchor='w', padx=8, pady=6)
        self.input_dir_label.pack(fill=tk.X, padx=12)
        
        # 2. 输出目录设置
        output_header_frame = tk.Frame(batch_frame, bg=COLORS['panel_bg'])
        output_header_frame.pack(fill=tk.X, padx=12, pady=(16, 4))
        
        tk.Label(output_header_frame, text='📤 输出目录', font=('SF Pro Display', 12, 'bold'),
                 bg=COLORS['panel_bg'], fg=COLORS['text_primary']).pack(side=tk.LEFT)
                 
        output_dir_btn = tk.Label(output_header_frame, text='选择', font=('SF Pro Text', 10),
                                  bg=COLORS['accent'], fg='white', padx=10, pady=4, cursor='hand2')
        output_dir_btn.pack(side=tk.LEFT, padx=(10, 0))
        output_dir_btn.bind('<Button-1>', lambda e: self.select_output_dir())
        
        # 打开目录按钮
        output_open_btn = tk.Label(output_header_frame, text='打开', font=('SF Pro Text', 10),
                                   bg=COLORS['bg_tertiary'], fg=COLORS['text_primary'], padx=10, pady=4, cursor='hand2')
        output_open_btn.pack(side=tk.LEFT, padx=(4, 0))
        output_open_btn.bind('<Button-1>', lambda e: self.open_directory(self.batch_output_dir))
        
        self.output_dir_label = tk.Label(batch_frame, text=self.batch_output_dir or '未设置',
                                         font=('SF Pro Text', 9), bg=COLORS['bg_secondary'],
                                         fg=COLORS['text_secondary'], anchor='w', padx=8, pady=6)
        self.output_dir_label.pack(fill=tk.X, padx=12)
        
        # 3. 操作区域标题 (放在分隔线中间)
        op_title_frame = tk.Frame(batch_frame, bg=COLORS['panel_bg'])
        op_title_frame.pack(fill=tk.X, padx=12, pady=(10, 8))
        
        # 使用 grid 布局实现中间文字两边线条
        op_title_frame.columnconfigure(0, weight=1)
        # column 1 contains label
        op_title_frame.columnconfigure(2, weight=1)
        
        tk.Frame(op_title_frame, height=1, bg=COLORS['separator']).grid(row=0, column=0, sticky='ew')
        tk.Label(op_title_frame, text='⚡ 批量操作', font=('SF Pro Display', 12, 'bold'),
                 bg=COLORS['panel_bg'], fg=COLORS['text_primary']).grid(row=0, column=1, padx=8)
        tk.Frame(op_title_frame, height=1, bg=COLORS['separator']).grid(row=0, column=2, sticky='ew')
        
        # 从目录加载按钮
        load_from_dir_btn = tk.Label(
            batch_frame, text='📂 从输入目录加载图片',
            bg=COLORS['warning'], fg='white',
            font=('SF Pro Text', 11, 'bold'), pady=10, cursor='hand2'
        )
        load_from_dir_btn.pack(anchor='w', padx=12, pady=4, ipadx=10)
        load_from_dir_btn.bind('<Button-1>', lambda e: self.load_from_input_dir())
        
        # 4. 状态和选项区域
        status_frame = tk.Frame(batch_frame, bg=COLORS['panel_bg'])
        status_frame.pack(fill=tk.X, padx=12, pady=12)
        
        # 状态显示
        self.batch_count_label = tk.Label(
            status_frame, text='已加载: 0 张',
            bg=COLORS['panel_bg'], fg=COLORS['text_primary'],
            font=('SF Pro Display', 11, 'bold'), anchor='w'
        )
        self.batch_count_label.pack(fill=tk.X)
        
        self.batch_status_label = tk.Label(
            status_frame, text='待处理: 0 张 | 本次已处理: 0 张',
            bg=COLORS['panel_bg'], fg=COLORS['text_secondary'],
            font=('SF Pro Text', 10), anchor='w'
        )
        self.batch_status_label.pack(fill=tk.X, pady=(2, 8))
        
        # 已移除强制重新处理选项 (默认不覆盖，因文件名已唯一)

        # 参考示例位置选项
        match_canvas_check = tk.Checkbutton(
            status_frame, text='参考示例位置和缩放', variable=self.batch_match_canvas,
            bg=COLORS['panel_bg'], fg=COLORS['text_primary'],
            font=('SF Pro Text', 10), selectcolor=COLORS['bg_secondary'],
            activebackground=COLORS['panel_bg']
        )
        match_canvas_check.pack(anchor='w', pady=(0, 15))
        Tooltip(match_canvas_check, '批量处理时，按照当前画布上图片的位置和缩放比例来放置每张图片')

        # --- 文字目录设置 ---
        text_dir_frame = tk.LabelFrame(batch_frame, text='🔤 批量文字', 
                                       font=('SF Pro Text', 10, 'bold'),
                                       bg=COLORS['panel_bg'], fg=COLORS['text_secondary'],
                                       padx=10, pady=8, bd=1, relief='flat')
        text_dir_frame.pack(fill=tk.X, padx=12, pady=(0, 12))
        
        # 启用文字目录勾选框
        text_dir_check = tk.Checkbutton(text_dir_frame, text='启用批量配文', variable=self.batch_use_text_dir,
                      bg=COLORS['panel_bg'], fg=COLORS['text_primary'], font=('SF Pro Text', 10),
                      selectcolor=COLORS['accent'], activebackground=COLORS['panel_bg'])
        text_dir_check.pack(anchor='w')
        Tooltip(text_dir_check, '勾选后将尝试为每张图片添加文字 (源自Excel文件)；若未找到对应文字，则使用当前编辑器内容')
        
        # 文字目录选择
        text_dir_select_frame = tk.Frame(text_dir_frame, bg=COLORS['panel_bg'])
        text_dir_select_frame.pack(fill=tk.X, pady=(4, 0))
        
        text_dir_btn = tk.Label(text_dir_select_frame, text='选择 Excel 数据表', font=('SF Pro Text', 10),
                               bg=COLORS['bg_tertiary'], fg=COLORS['text_primary'], padx=8, pady=4, cursor='hand2')
        text_dir_btn.pack(side=tk.LEFT)
        text_dir_btn.bind('<Button-1>', lambda e: self.select_excel_file())
        
        # 模板下载按钮
        template_btn = tk.Label(text_dir_select_frame, text='下载模版', font=('SF Pro Text', 10),
                               bg=COLORS['bg_tertiary'], fg=COLORS['accent'], padx=8, pady=4, cursor='hand2')
        template_btn.pack(side=tk.LEFT, padx=(4, 0))
        template_btn.bind('<Button-1>', lambda e: self.download_excel_template())
        
        text_open_btn = tk.Label(text_dir_select_frame, text='打开', font=('SF Pro Text', 10),
                                bg=COLORS['bg_tertiary'], fg=COLORS['text_primary'], padx=8, pady=4, cursor='hand2')
        text_open_btn.pack(side=tk.LEFT, padx=(4, 0))
        text_open_btn.bind('<Button-1>', lambda e: self.open_directory(self.batch_text_dir))
        
        self.text_dir_label = tk.Label(text_dir_frame, text=os.path.basename(self.batch_text_dir) if self.batch_text_dir else '未选择文件',
                                       font=('SF Pro Text', 9), bg=COLORS['bg_secondary'],
                                       fg=COLORS['text_secondary'], anchor='w', padx=8, pady=4)
        self.text_dir_label.pack(fill=tk.X, pady=(4, 0))
        
        tk.Label(text_dir_frame, text='提示: 使用模版配置文案，程序将自动更新读取时间',
                font=('SF Pro Text', 8), bg=COLORS['panel_bg'], fg=COLORS['text_tertiary']
                ).pack(anchor='w', pady=(4, 0))

        # --- 随机化选项区域 ---
        # 必须先定义 random_frame
        random_frame = tk.LabelFrame(batch_frame, text='🎲 随机化选项', 
                                   font=('SF Pro Text', 10, 'bold'),
                                   bg=COLORS['panel_bg'], fg=COLORS['text_secondary'],
                                   padx=10, pady=8, bd=1, relief='flat')
        random_frame.pack(fill=tk.X, padx=12, pady=(0, 12))
        
        # 使用 Grid 布局放置选项
        tk.Checkbutton(random_frame, text='随机边框颜色', variable=self.batch_random_color,
                      bg=COLORS['panel_bg'], fg=COLORS['text_primary'], font=('SF Pro Text', 10),
                      selectcolor=COLORS['bg_secondary'], activebackground=COLORS['panel_bg']
                      ).grid(row=0, column=0, sticky='w', padx=(0, 15))
        
        tk.Checkbutton(random_frame, text='随机线条样式', variable=self.batch_random_style,
                      bg=COLORS['panel_bg'], fg=COLORS['text_primary'], font=('SF Pro Text', 10),
                      selectcolor=COLORS['bg_secondary'], activebackground=COLORS['panel_bg']
                      ).grid(row=0, column=1, sticky='w', padx=0)
                      
        tk.Checkbutton(random_frame, text='随机边框图案', variable=self.batch_random_pattern,
                       bg=COLORS['panel_bg'], fg=COLORS['text_primary'], font=('SF Pro Text', 10),
                       selectcolor=COLORS['bg_secondary'], activebackground=COLORS['panel_bg']
                       ).grid(row=1, column=0, columnspan=2, sticky='w', pady=(5, 0))

        tk.Checkbutton(random_frame, text='随机文字高亮', variable=self.batch_random_highlight,
                       bg=COLORS['panel_bg'], fg=COLORS['text_primary'], font=('SF Pro Text', 10),
                       selectcolor=COLORS['bg_secondary'], activebackground=COLORS['panel_bg']
                       ).grid(row=1, column=1, sticky='w', pady=(5, 0))

        tk.Checkbutton(random_frame, text='随机字体样式', variable=self.batch_random_font_style,
                       bg=COLORS['panel_bg'], fg=COLORS['text_primary'], font=('SF Pro Text', 10),
                       selectcolor=COLORS['bg_secondary'], activebackground=COLORS['panel_bg']
                       ).grid(row=2, column=0, columnspan=2, sticky='w', pady=(5, 0))

        
        # 5. 批量导出按钮
        batch_export_btn = tk.Label(
            batch_frame, text='⚡ 批量生成并导出',
            bg=COLORS['success'], fg='white',
            font=('SF Pro Text', 11, 'bold'), pady=12, cursor='hand2'
        )
        batch_export_btn.pack(anchor='w', padx=12, pady=4, ipadx=10)
        batch_export_btn.bind('<Button-1>', lambda e: self.batch_export())
        
        # 6. 日志输出框
        # 6. 日志输出框
        log_header_frame = tk.Frame(batch_frame, bg=COLORS['panel_bg'])
        log_header_frame.pack(fill=tk.X, padx=12, pady=(20, 4))
        
        tk.Label(log_header_frame, text='📋 处理日志', font=('SF Pro Display', 11, 'bold'),
                 bg=COLORS['panel_bg'], fg=COLORS['text_primary'], anchor='w').pack(side=tk.LEFT)
        
        # 复制日志按钮 (放在标题后面)
        copy_btn = tk.Label(log_header_frame, text='[复制日志]', font=('SF Pro Text', 10),
                           bg=COLORS['panel_bg'], fg=COLORS['accent'], cursor='hand2')
        copy_btn.pack(side=tk.LEFT, padx=(10, 0))
        copy_btn.bind('<Button-1>', lambda e: self.copy_batch_log())
        
        log_frame = tk.Frame(batch_frame, bg=COLORS['bg_secondary'])
        log_frame.pack(fill=tk.BOTH, expand=True, padx=12, pady=(0, 12))
        
        self.batch_log_text = tk.Text(log_frame, height=20, font=('Menlo', 9),
                                       bg=COLORS['bg'], fg=COLORS['text_secondary'],
                                       wrap=tk.WORD, state=tk.DISABLED,
                                       highlightthickness=1, highlightbackground=COLORS['separator'])
        log_scrollbar = tk.Scrollbar(log_frame, command=self.batch_log_text.yview)
        self.batch_log_text.configure(yscrollcommand=log_scrollbar.set)
        
        log_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.batch_log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # 7. 底部说明
        tip_text = "支持格式：JPG, JPEG, PNG, BMP, GIF"
        tk.Label(batch_frame, text=tip_text, font=('SF Pro Text', 9),
                 bg=COLORS['panel_bg'], fg=COLORS['text_secondary'], anchor='w',
                 padx=12, pady=12).pack(fill=tk.X)

    def copy_batch_log(self):
        """复制批量处理日志到剪贴板"""
        if hasattr(self, 'batch_log_text'):
            content = self.batch_log_text.get('1.0', tk.END).strip()
            if content:
                self.title_bar.clipboard_clear()
                self.title_bar.clipboard_append(content)
                self.title_bar.update() # 必须 update 才能写入剪贴板
                self.show_toast("日志已复制到剪贴板")
            else:
                self.show_toast("日志内容为空")

    def update_batch_status_text(self):
        """更新批量处理状态文本"""
        if not hasattr(self, 'batch_images') or not self.batch_images:
            return
            
        # Since we always process provided images (unique filenames), pending is just the total count
        pending = len(self.batch_images)
            
        # 本次已处理保持不变，或者如果不希望跟“重新生成”状态挂钩也可以
        processed_text = getattr(self, 'current_session_processed', 0)
        
        if hasattr(self, 'batch_status_label'):
            self.batch_status_label.config(text=f'待处理: {pending} 张 | 本次已处理: {processed_text} 张')
    
    def select_size_preset(self, preset):
        """选择尺寸预设"""
        old_preset_id = self.current_size_preset['id']
        self.current_size_preset = preset
        self.image_processor.set_canvas_size(preset['width'], preset['height'])
        
        # 计算适合显示区域的画布尺寸（占可用空间的90%）
        window_width = self.winfo_width() or self.winfo_screenwidth() * 0.85
        window_height = self.winfo_height() or self.winfo_screenheight() * 0.85
        # 左面板约120px，右面板约320px，工具栏约50px，边距约40px
        available_width = int((window_width - 120 - 320 - 40) * 0.9)
        available_height = int((window_height - 50 - 40) * 0.9)
        
        ratio = preset['width'] / preset['height']
        
        if ratio > available_width / available_height:
            # 宽图，以宽度为准
            display_width = available_width
            display_height = int(display_width / ratio)
        else:
            # 高图或方图，以高度为准
            display_height = available_height
            display_width = int(display_height * ratio)
        
        self.canvas_widget.resize_canvas(display_width, display_height)
        
        # 更新按钮选中效果 (Label)
        if hasattr(self, 'size_preset_buttons'):
            for pid, btn in self.size_preset_buttons.items():
                if pid == preset['id']:
                    btn.config(bg=COLORS['selected_bg'], fg=COLORS['text_bright'])
                else:
                    btn.config(bg=COLORS['bg_tertiary'], fg=COLORS['text_primary'])
        
        # 如果已有图片，重新调整大小
        if self.image_processor.current_image:
            self.refresh_canvas()
        
        # 重新应用背景颜色
        if hasattr(self, 'background_color') and self.background_color:
            self.canvas_widget.set_background_color(self.background_color)
        
        # 延迟重新应用边框（等待画布更新完成）
        self.after(50, self.reapply_border_after_resize)
        
        print(f"✓ 尺寸设置: {preset['name']} ({preset['width']}×{preset['height']})")
    
    def reapply_border_after_resize(self):
        """尺寸调整后重新应用边框"""
        if hasattr(self, 'border_config') and self.border_config['width'] > 0:
            self.canvas_widget.apply_custom_border(self.border_config)
    
    def apply_transform(self, transform_type, angle=None):
        """应用变换操作"""
        if not self.image_processor.current_image:
            messagebox.showwarning('提示', '请先上传图片！')
            return
        
        if transform_type == 'rotate':
            self.image_processor.rotate_image(angle)
            action_name = f"旋转{abs(angle)}°"
        elif transform_type == 'flip_h':
            self.image_processor.flip_image(horizontal=True)
            action_name = "水平翻转"
        elif transform_type == 'flip_v':
            self.image_processor.flip_image(horizontal=False)
            action_name = "垂直翻转"
        else:
            return
        
        self.image_processor.resize_to_canvas(maintain_ratio=True)
        self.refresh_canvas()
        self.save_history(action_name)
    
    def toggle_filter(self, filter_type, btn):
        """切换滤镜状态"""
        if not self.image_processor.current_image:
            messagebox.showwarning('提示', '请先上传图片！')
            return
        
        filter_names = {
            'grayscale': '黑白', 'sharpen': '锐化', 'blur': '模糊',
            'smooth': '平滑', 'contour': '轮廓', 'emboss': '浮雕'
        }
        
        if filter_type in self.active_filters:
            # 取消滤镜 - 重置图片并重新应用其他活跃滤镜
            self.active_filters.discard(filter_type)
            btn.config(
                text=self.filter_base_texts[filter_type],
                bg=COLORS['bg_tertiary'],
                fg=COLORS['text_primary']  # 恢复原文字颜色
            )
            # 从原始图片开始重新应用所有活跃滤镜
            self.reapply_all_filters()
            self.save_history(f"取消{filter_names.get(filter_type, filter_type)}滤镜")
        else:
            # 应用滤镜
            self.active_filters.add(filter_type)
            btn.config(
                text=self.filter_base_texts[filter_type] + " ✓",
                bg=COLORS['accent'],
                fg='white'  # 白色文字确保可读性
            )
            self.image_processor.apply_filter(filter_type)
            self.refresh_canvas()
            self.save_history(f"应用{filter_names.get(filter_type, filter_type)}滤镜")
    
    def reapply_all_filters(self):
        """重新应用所有活跃滤镜"""
        # 从原始图片开始
        self.image_processor.reset_image()
        if self.image_processor.current_image:
            self.image_processor.resize_to_canvas(maintain_ratio=True)
            # 重新应用所有活跃滤镜
            for f_type in self.active_filters:
                self.image_processor.apply_filter(f_type)
            self.refresh_canvas()
    
    def apply_filter(self, filter_type):
        """应用滤镜（单次应用，不可切换）"""
        if not self.image_processor.current_image:
            messagebox.showwarning('提示', '请先上传图片！')
            return
        
        filter_names = {
            'grayscale': '黑白', 'sharpen': '锐化', 'blur': '模糊',
            'smooth': '平滑', 'contour': '轮廓', 'emboss': '浮雕'
        }
        
        self.image_processor.apply_filter(filter_type)
        self.refresh_canvas()
        self.save_history(f"应用{filter_names.get(filter_type, filter_type)}滤镜")
    
    def apply_adjustment(self, adjust_type):
        """应用图片调整"""
        if not self.image_processor.current_image:
            messagebox.showwarning('提示', '请先上传图片！')
            return
        
        if adjust_type == 'brightness':
            factor = self.brightness_scale.get()
            self.image_processor.adjust_brightness(factor)
            action_name = f"亮度调整({factor})"
        elif adjust_type == 'contrast':
            factor = self.contrast_scale.get()
            self.image_processor.adjust_contrast(factor)
            action_name = f"对比度调整({factor})"
        elif adjust_type == 'saturation':
            factor = self.saturation_scale.get()
            self.image_processor.adjust_saturation(factor)
            action_name = f"饱和度调整({factor})"
        else:
            return
        
        self.refresh_canvas()
        self.save_history(action_name)
    
    def reset_image_and_sliders(self):
        """重置图片和滑块"""
        self.image_processor.reset_image()
        if self.image_processor.current_image:
            self.image_processor.resize_to_canvas(maintain_ratio=True)
            self.refresh_canvas()
            
            # 重置滑块
            if hasattr(self, 'brightness_scale'):
                self.brightness_scale.set(1.0)
            if hasattr(self, 'contrast_scale'):
                self.contrast_scale.set(1.0)
            if hasattr(self, 'saturation_scale'):
                self.saturation_scale.set(1.0)
            
            self.save_history("重置图片")
    
    def reset_single_adjustment(self, adjust_type):
        """重置单个调整滑块"""
        if not self.image_processor.current_image:
            messagebox.showwarning('提示', '请先上传图片！')
            return
        
        if adjust_type == 'brightness' and hasattr(self, 'brightness_scale'):
            self.brightness_scale.set(1.0)
        elif adjust_type == 'contrast' and hasattr(self, 'contrast_scale'):
            self.contrast_scale.set(1.0)
        elif adjust_type == 'saturation' and hasattr(self, 'saturation_scale'):
            self.saturation_scale.set(1.0)
        
        # 从原始图片重新应用所有当前调整值
        self.image_processor.reset_image()
        if self.image_processor.current_image:
            self.image_processor.resize_to_canvas(maintain_ratio=True)
            
            # 重新应用活跃滤镜
            for f_type in getattr(self, 'active_filters', []):
                self.image_processor.apply_filter(f_type)
            
            # 应用当前滑块值
            if hasattr(self, 'brightness_scale') and self.brightness_scale.get() != 1.0:
                self.image_processor.adjust_brightness(self.brightness_scale.get())
            if hasattr(self, 'contrast_scale') and self.contrast_scale.get() != 1.0:
                self.image_processor.adjust_contrast(self.contrast_scale.get())
            if hasattr(self, 'saturation_scale') and self.saturation_scale.get() != 1.0:
                self.image_processor.adjust_saturation(self.saturation_scale.get())
            
            self.refresh_canvas()
            name_map = {'brightness': '亮度', 'contrast': '对比度', 'saturation': '饱和度'}
            self.save_history(f"重置{name_map.get(adjust_type, adjust_type)}")
    
    def upload_image(self):
        """上传图片"""
        file_path = filedialog.askopenfilename(
            title='选择图片',
            filetypes=[
                ('图片文件', '*.jpg *.jpeg *.png *.bmp *.gif'),
                ('所有文件', '*.*')
            ]
        )
        
        if file_path:
            if self.image_processor.load_image(file_path):
                self.image_processor.resize_to_canvas(maintain_ratio=True)
                self.refresh_canvas()
                self.save_history("上传图片")
                # messagebox.showinfo('成功', '图片上传成功！')
            else:
                messagebox.showerror('错误', '图片加载失败！')
    
    def reset_image(self):
        """重置图片"""
        self.image_processor.reset_image()
        if self.image_processor.current_image:
            self.image_processor.resize_to_canvas(maintain_ratio=True)
            self.refresh_canvas()
            self.save_history("重置图片")
    
    def add_sticker(self, sticker):
        """添加贴纸"""
        # 如果有PNG图片文件，优先使用图片
        sticker_path = os.path.join(os.path.dirname(__file__), 'assets', 'stickers', sticker.get('file', ''))
        if os.path.exists(sticker_path):
            try:
                # 加载PNG贴纸，但还是用emoji显示（因为canvas_widget当前使用emoji）
                # TODO: 后续可以支持真正的图片贴纸
                self.canvas_widget.add_sticker(sticker['emoji'], font_size=48)
            except Exception as e:
                print(f"加载贴纸图片失败: {e}")
                self.canvas_widget.add_sticker(sticker['emoji'], font_size=48)
        else:
            self.canvas_widget.add_sticker(sticker['emoji'], font_size=48)
        
        self.save_history("添加贴纸")
        self.update_layer_list()
    
    def rotate_image(self, angle):
        """旋转图片"""
        if self.image_processor.base_image:
            rotated = self.image_processor.base_image.rotate(angle, expand=True)
            self.image_processor.base_image = rotated
            self.image_processor.current_image = rotated
            self.image_processor.resize_to_canvas(maintain_ratio=True)
            self.refresh_canvas()
            self.save_history("旋转图片")
    
    def flip_image(self, direction):
        """翻转图片"""
        if self.image_processor.base_image:
            from PIL import Image as PILImage
            if direction == 'horizontal':
                flipped = self.image_processor.base_image.transpose(PILImage.FLIP_LEFT_RIGHT)
            else:  # vertical
                flipped = self.image_processor.base_image.transpose(PILImage.FLIP_TOP_BOTTOM)
            
            self.image_processor.base_image = flipped
            self.image_processor.current_image = flipped
            self.image_processor.resize_to_canvas(maintain_ratio=True)
            self.refresh_canvas()
            self.save_history("翻转图片")
    
    def select_border(self, border):
        """选择边框"""
        self.current_border = border
        self.canvas_widget.add_border(border)
        self.save_history("选择边框")
    
    def delete_selected_sticker(self):
        """删除选中的贴纸"""
        if self.canvas_widget.delete_selected_sticker():
            self.save_history("删除贴纸")
            self.update_layer_list()
            messagebox.showinfo('成功', '贴纸已删除')
        else:
            messagebox.showwarning('提示', '请先点击选择要删除的贴纸')
    
    def refresh_canvas(self):
        """刷新画布"""
        current_image = self.image_processor.get_current_image()
        if current_image:
            self.canvas_widget.display_image(current_image)
            # 使用 border_config 而非 current_border，确保边框配置一致
            self.canvas_widget.apply_custom_border(self.border_config)
        else:
            # 没有图片时清除画布上的主图片
            self.canvas_widget.clear_main_image()
            
        # 确保顺序生效后再强制定序一次 (处理异步渲染)
        self.after(50, lambda: self.canvas_widget._ensure_layer_order())
        
        # 更新图层列表 (如果已创建)
        if hasattr(self, 'update_layer_list'):
            self.update_layer_list()
    
    def export_image(self):
        """导出图片 (支持无图片导出，仅背景+文字)"""
        # 不再强制要求上传图片
        
        # 生成默认文件名
        default_name = f"tupian_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
        
        file_path = filedialog.asksaveasfilename(
            title='保存图片',
            defaultextension='.png',
            initialfile=default_name,
            filetypes=[
                ('PNG图片', '*.png'),
                ('JPEG图片', '*.jpg'),
                ('所有文件', '*.*')
            ]
        )
        
        if file_path:
            # 获取导出参数
            preset_width = self.current_size_preset['width']
            preset_height = self.current_size_preset['height']
            display_width = self.canvas_widget.width
            display_height = self.canvas_widget.height
            
            # 使用独立的缩放比例（避免比例失真）
            scale_x = preset_width / display_width
            scale_y = preset_height / display_height
            
            print(f"[DEBUG] Export: preset={preset_width}x{preset_height}, display={display_width}x{display_height}")
            print(f"[DEBUG] Export: scale_x={scale_x:.2f}, scale_y={scale_y:.2f}")
            print(f"[DEBUG] Border config: {self.border_config}")
            
            # 1. 创建背景图层
            final_img = Image.new('RGB', (preset_width, preset_height), self.background_color)
            draw = ImageDraw.Draw(final_img)
            
            # 2. 绘制背景图案
            if self.background_pattern and self.background_pattern != 'none':
                # 这里简单重构图案绘制逻辑，或调用专门的 helper
                scaled_pattern_size = int(self.background_pattern_size * max(scale_x, scale_y))
                # 使用临时处理器来绘制图案以免影响主状态
                temp_proc = ImageProcessor()
                temp_proc.current_image = final_img
                temp_proc.draw_background_pattern(
                    self.background_pattern,
                    self.background_pattern_color,
                    scaled_pattern_size
                )
                final_img = temp_proc.current_image
            
            # 3. 绘制主图片
            if self.image_processor.current_image:
                # 首先获取图片在画布上的实际位置
                main_img_id = self.canvas_widget.main_image_id
                if main_img_id:
                    coords = self.canvas_widget.canvas.coords(main_img_id)
                    if coords:
                        cx, cy = coords
                        # 获取图片渲染大小
                        # 注意：Tkinter 里的图片坐标是中心点
                        main_pil = self.image_processor.current_image
                        
                        # 按比例缩放并粘贴 (使用独立的scale_x/scale_y保持比例)
                        scaled_main_w = int(main_pil.width * scale_x)
                        scaled_main_h = int(main_pil.height * scale_y)
                        scaled_main_pil = main_pil.resize((scaled_main_w, scaled_main_h), Image.Resampling.LANCZOS)
                        
                        # 计算粘贴位置
                        paste_x = int(cx * scale_x - scaled_main_w / 2)
                        paste_y = int(cy * scale_y - scaled_main_h / 2)
                        final_img.paste(scaled_main_pil, (paste_x, paste_y), scaled_main_pil if scaled_main_pil.mode == 'RGBA' else None)
            
            # 4. 绘制贴纸
            for sticker in self.canvas_widget.get_stickers():
                scaled_x = int(sticker['x'] * scale_x)
                scaled_y = int(sticker['y'] * scale_y)
                scaled_size = int(sticker['size'] * max(scale_x, scale_y))
                
                print(f"[DEBUG] Sticker: orig=({sticker['x']}, {sticker['y']}), scaled=({scaled_x}, {scaled_y}), size={scaled_size}")
                
                try:
                    # Apple Color Emoji 只支持固定大小，使用 160 像素渲染后缩放
                    base_size = 160
                    font = ImageFont.truetype("/System/Library/Fonts/Apple Color Emoji.ttc", base_size)
                    
                    # 创建临时图层渲染 emoji
                    temp_size = base_size * 2  # 留足够边距
                    emoji_temp = Image.new('RGBA', (temp_size, temp_size), (0, 0, 0, 0))
                    emoji_draw = ImageDraw.Draw(emoji_temp)
                    emoji_draw.text((temp_size // 2, temp_size // 2), sticker['text'], font=font, anchor="mm", embedded_color=True)
                    
                    # 裁剪掉透明边距
                    bbox = emoji_temp.getbbox()
                    if bbox:
                        emoji_cropped = emoji_temp.crop(bbox)
                        # 缩放到目标尺寸
                        emoji_resized = emoji_cropped.resize((scaled_size, scaled_size), Image.Resampling.LANCZOS)
                        
                        # 计算粘贴位置（中心对齐）
                        paste_x = scaled_x - scaled_size // 2
                        paste_y = scaled_y - scaled_size // 2
                        
                        # 合成到最终图片
                        if final_img.mode != 'RGBA':
                            final_img = final_img.convert('RGBA')
                        final_img.paste(emoji_resized, (paste_x, paste_y), emoji_resized)
                except Exception as e:
                    print(f"[DEBUG] Emoji rendering error: {e}")
                    # 降级方案：使用文本
                    sticker_draw = ImageDraw.Draw(final_img)
                    try:
                        font = ImageFont.truetype("/System/Library/Fonts/STHeiti Light.ttc", scaled_size)
                    except:
                        font = ImageFont.load_default()
                    sticker_draw.text((scaled_x, scaled_y), sticker['text'], fill='black', font=font, anchor="mm")
            
            # 4.5 绘制文字层 (NEW)

            
            # 5. 绘制边框 (在最上层)
            from image_processor import CompositeImage
            
            # 使用 border_config 而非 current_border
            border_config = self.border_config.copy()
            print(f"[DEBUG] Exporting with border config: {border_config}")  # 调试
            
            # 只检查 width > 0 即可应用边框（移除对 id 的检查）
            uniform_scale = max(scale_x, scale_y)
            if border_config.get('width', 0) > 0:
                # 缩放边框宽度和圆角
                border_config['width'] = int(border_config.get('width', 10) * uniform_scale)
                if 'radius' in border_config:
                    border_config['radius'] = int(border_config['radius'] * uniform_scale)
                # 缩放图案大小
                if 'pattern_size' in border_config:
                    border_config['pattern_size'] = int(border_config['pattern_size'] * uniform_scale)
                
                composite = CompositeImage(preset_width, preset_height)
                composite.canvas = final_img.copy()
                composite.draw = ImageDraw.Draw(composite.canvas)
                
                if border_config.get('radius', 0) > 0:
                    composite.add_rounded_border(border_config)
                else:
                    composite.add_border(border_config)
                final_img = composite.canvas
                print(f"[DEBUG] Border applied successfully")
            else:
                print(f"[DEBUG] Skipping border - width={border_config.get('width')}")

            # 6. 绘制文字层 (Moved to be AFTER border to avoid being covered)
            # 6. 绘制文字层 (Moved to be AFTER border to avoid being covered)
            if hasattr(self, 'current_text_layer') and self.current_text_layer:
                # 使用 x 轴缩放比例 (假设文字随宽度缩放)
                text_scale = scale_x
                
                # 计算有效边框宽度 (用于文字防遮挡)，与 batch_export 逻辑保持一致
                effective_border_width = 0
                if border_config.get('width', 0) > 0:
                     effective_border_width = border_config.get('width', 0)
                     # 稍微多给一点余量 (也要缩放)
                     effective_border_width += int(10 * text_scale)
                
                # 渲染文字到独立图层
                print(f"[DEBUG] Exporting text layer: {self.current_text_layer.content[:10]}..., scale={text_scale}, safe_margin={effective_border_width}")
                text_img, tx, ty = self.current_text_layer.render(preset_width, preset_height, scale=text_scale, safe_margin_x=effective_border_width)
                
                if text_img:
                    # 合成到最终图片
                    if final_img.mode != 'RGBA':
                        final_img = final_img.convert('RGBA')
                    
                    # 确保 text_img 也是 RGBA
                    if text_img.mode != 'RGBA':
                        text_img = text_img.convert('RGBA')
                        
                    final_img.paste(text_img, (tx, ty), text_img)
            
            # 6. 保存
            try:
                final_img.save(file_path)
                
                # 根据勾选框状态决定是否自动保存预设
                save_msg = f'图片已保存到:\n{file_path}'
                if hasattr(self, 'auto_save_preset_var') and self.auto_save_preset_var.get():
                    self.save_preset_theme(silent=True)
                    save_msg += '\n\n✓ 主题预设已自动保存'
                
                # 询问是否打开目录
                if messagebox.askyesno('导出成功', save_msg + '\n\n是否打开所在目录？'):
                    try:
                        folder_path = os.path.dirname(file_path)
                        self.open_directory(folder_path, select_file=file_path)
                    except Exception as e:
                        print(f"打开目录失败: {e}")
            except Exception as e:
                messagebox.showerror('错误', f'保存失败: {e}')
    
    def select_input_dir(self):
        """选择输入目录"""
        dir_path = filedialog.askdirectory(title='选择输入目录', initialdir=self.batch_input_dir or None)
        if dir_path:
            self.batch_input_dir = dir_path
            self.input_dir_label.config(text=dir_path)
            self.save_settings()
    
    def select_output_dir(self):
        """选择输出目录"""
        dir_path = filedialog.askdirectory(title='选择输出目录', initialdir=self.batch_output_dir or None)
        if dir_path:
            self.batch_output_dir = dir_path
            self.output_dir_label.config(text=dir_path)
            self.save_settings()
    
    def download_excel_template(self):
        """下载 Excel 模板"""
        template_source = os.path.join('assets', 'template', '批量导出文字内容模版.xlsx')
        if not os.path.exists(template_source):
             messagebox.showerror('错误', '找不到模板文件！')
             return

        save_path = filedialog.asksaveasfilename(
            title='保存模板',
            initialfile='批量导出文字内容模版.xlsx',
            defaultextension='.xlsx',
            filetypes=[('Excel 文件', '*.xlsx')]
        )
        if save_path:
            try:
                import shutil
                shutil.copy2(template_source, save_path)
                messagebox.showinfo('成功', f'模板已保存到:\n{save_path}')
                
                # 询问是否立即打开
                if messagebox.askyesno('提示', '是否立即打开模板文件？'):
                    # 尝试打开文件
                    if sys.platform == 'darwin':
                        subprocess.run(['open', save_path])
                    elif sys.platform == 'win32':
                        os.startfile(save_path)
                    else:
                        subprocess.run(['xdg-open', save_path])
                        
            except Exception as e:
                messagebox.showerror('错误', f'保存模板失败: {e}')

    def select_excel_file(self):
        """选择 Excel 文件"""
        file_path = filedialog.askopenfilename(
            title='选择 Excel 数据表',
            filetypes=[('Excel 文件', '*.xlsx'), ('Excel 97-2003', '*.xls')],
            initialdir=os.path.dirname(self.batch_text_dir) if self.batch_text_dir else None
        )
        if file_path:
            self.batch_text_dir = file_path 
            if hasattr(self, 'text_dir_label'):
                self.text_dir_label.config(text=os.path.basename(file_path))
            self.save_settings()
    
    def show_toast(self, message, duration=2000):
        """显示非阻塞的 Toast 提示"""
        toast = tk.Toplevel(self)
        toast.overrideredirect(True)  # 无边框
        
        # 计算位置（居中显示）
        window_width = self.winfo_width()
        window_height = self.winfo_height()
        window_x = self.winfo_rootx()
        window_y = self.winfo_rooty()
        
        # 创建内容
        frame = tk.Frame(toast, bg='#333333', padx=20, pady=10)
        frame.pack(fill=tk.BOTH, expand=True)
        
        tk.Label(frame, text=message, fg='white', bg='#333333', 
                 font=('SF Pro Text', 11)).pack()
        
        # 调整大小和位置
        toast.update_idletasks()
        toast_width = toast.winfo_width()
        toast_height = toast.winfo_height()
        
        x = window_x + (window_width - toast_width) // 2
        y = window_y + (window_height - toast_height) // 2 + 100 #稍微偏下
        
        toast.geometry(f"{toast_width}x{toast_height}+{x}+{y}")
        
        # 确保在最上层
        toast.attributes('-topmost', True)
        toast.lift()
        
        # 设置圆角效果（macOS特有，Windows可能不生效但也不报错）
        try:
            toast.attributes('-transparent', True) # 尝试透明
        except:
            pass
            
        # 自动关闭
        toast.after(duration, toast.destroy)

    def load_from_input_dir(self):
        """从输入目录加载图片"""
        if not self.batch_input_dir:
            self.show_toast('请先设置输入目录')
            return
        
        if not os.path.isdir(self.batch_input_dir):
            messagebox.showerror('错误', '输入目录不存在')
            return
        
        # 获取目录中所有图片
        extensions = ('.jpg', '.jpeg', '.png', '.bmp', '.gif')
        all_images = []
        for f in os.listdir(self.batch_input_dir):
            if f.lower().endswith(extensions):
                all_images.append(os.path.join(self.batch_input_dir, f))
        
        self.batch_images = all_images
        
        # 统计已处理（历史）和未处理
        # 注意：这里的 pending 是基于历史记录的，用于增量处理
        pending = [p for p in all_images if os.path.basename(p) not in self.processed_images]
        
        # 重置当前会话的“本次已处理”计数
        self.current_session_processed = 0
        
        self.batch_count_label.config(text=f'已加载: {len(all_images)} 张图片')
        if hasattr(self, 'batch_status_label'):
             # UI显示：待处理(增量) | 本次已处理
             self.batch_status_label.config(text=f'待处理: {len(pending)} 张 | 本次已处理: 0 张')
        
        if all_images:
            self.show_toast(f'成功加载 {len(all_images)} 张图片')
        else:
            messagebox.showwarning('提示', '目录中没有图片文件')
    
    def batch_upload(self):
        """批量上传图片"""
        file_paths = filedialog.askopenfilenames(
            title='批量选择图片',
            filetypes=[
                ('图片文件', '*.jpg *.jpeg *.png *.bmp *.gif'),
                ('所有文件', '*.*')
            ]
        )
        
        if file_paths:
            self.batch_images = list(file_paths)
            self.batch_count_label.config(text=f'已选择: {len(self.batch_images)} 张图片')
            messagebox.showinfo('成功', f'已选择 {len(self.batch_images)} 张图片')
    
    def get_random_color(self):
        """随机获取颜色 (马卡龙 + 多巴胺色系)"""
        import random
        from constants import MACARON_COLORS, DOPAMINE_COLORS
        return random.choice(MACARON_COLORS + DOPAMINE_COLORS)
    
    def get_random_highlight_color(self):
        """随机获取高亮颜色 (仅限亮色)"""
        import random
        from constants import BRIGHT_HIGHLIGHT_COLORS
        return random.choice(BRIGHT_HIGHLIGHT_COLORS)

    def get_random_line_style(self):
        """随机获取线条样式"""
        import random
        from constants import LINE_STYLES
        return random.choice(LINE_STYLES)['id']

    def get_random_pattern(self):
        """随机获取边框图案"""
        import random
        from constants import BORDER_PATTERNS
        # 排除 'none'
        patterns = [p['id'] for p in BORDER_PATTERNS if p['id'] != 'none']
        return random.choice(patterns) if patterns else 'dots'


    def open_directory(self, path, select_file=None):
        """打开目录，支持选中文件"""
        if not os.path.exists(path):
            self.show_toast(f"目录不存在: {path}")
            return
            
        import platform
        import subprocess
        
        system = platform.system()
        try:
            if system == 'Darwin':  # macOS
                if select_file and os.path.exists(select_file):
                    subprocess.run(['open', '-R', select_file])
                else:
                    subprocess.run(['open', path])
            elif system == 'Windows':  # Windows
                if select_file and os.path.exists(select_file):
                    subprocess.run(['explorer', '/select,', os.path.normpath(select_file)])
                else:
                    os.startfile(path)
            else:  # Linux
                subprocess.run(['xdg-open', path])
        except Exception as e:
            self.show_toast(f"无法打开目录: {e}")
            print(f"Open directory error: {e}")
            messagebox.showerror('错误', f'无法打开目录: {e}')

    def _load_text_mapping(self, source_path):
        """加载文字映射 (仅 Excel)，并回写更新时间"""
        mapping = {}
        sequential_list = []
        
        if not source_path or not os.path.isfile(source_path):
            return None, []
            
        # Excel 模式
        try:
            import openpyxl
            from datetime import datetime
            
            # 必须用 load_workbook 加载，不能只读，因为要回写
            wb = openpyxl.load_workbook(source_path, data_only=False)
            ws = wb.active
            
            # 标记是否有修改
            has_update = False
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # 遍历所有行 (跳过 header? 假设第一行可能是 Header)
            # 为了确保准确，我们读取所有行，但通常第一行是标题。
            # 如果第一行是 "文件名" "内容"，我们一般跳过。
            # 简单策略：遍历所有，如果匹配特定特征才处理
            
            rows = list(ws.iter_rows(min_row=1)) # 获取所有行对象
            
            for row_idx, row in enumerate(rows):
                # 获取值 (注意: row是单元格对象，不是值，因为 data_only=False)
                # 使用 value 属性
                val1 = row[0].value
                val2 = row[1].value if len(row) > 1 else None
                
                col1 = str(val1).strip() if val1 is not None else ""
                col2 = str(val2).strip() if val2 is not None else ""
                
                # 跳过空行
                if not col1 and not col2:
                    continue
                    
                # 检查是否是标题行 (简单的关键词检查)
                if row_idx == 0:
                     if '文件名' in col1 or '内容' in col2 or 'Filename' in col1:
                         # 这里的标题行可以写入 "更新时间" 到第三列
                         if len(row) > 2:
                             row[2].value = "最后读取时间"
                             has_update = True
                         else:
                             # 只有两列，无法写标题到第三列? openpyxl 会自动扩展吗？可以
                             ws.cell(row=row_idx+1, column=3, value="最后读取时间")
                             has_update = True
                         continue

                # 提取数据
                content = ""
                # 逻辑复用：
                if '.' in col1 and len(col1) > 3:
                     # A=Filename, B=Content
                     content = col2
                     mapping[col1] = content
                else:
                     # Sequence
                     content = col2 if col2 else col1
                     sequential_list.append(content)
                
                # 回写时间到第3列 (Column C)
                # 只有当确实读取了这一行数据时才写
                if content:
                    ws.cell(row=row_idx+1, column=3, value=current_time)
                    has_update = True

            if has_update:
                try:
                    wb.save(source_path)
                    print(f"[INFO] 已更新 Excel 时间戳: {source_path}")
                except Exception as e:
                    print(f"[ERROR] 无法回写 Excel: {e} (可能文件被占用)")
                    self.show_toast(f"无法更新Excel时间: 文件被占用?")
                    
        except Exception as e:
            print(f"读取 Excel 失败: {e}")
            self.show_toast(f"读取 Excel 失败: {e}")
                
        return mapping, sequential_list

    def batch_export(self):
        """批量导出图片"""
        if not self.batch_images:
            messagebox.showwarning('提示', '请先加载图片！')
            return
        
        # 使用记忆的输出目录或选择新目录
        output_dir = self.batch_output_dir
        if not output_dir or not os.path.isdir(output_dir):
            output_dir = filedialog.askdirectory(title='选择输出目录', initialdir=self.batch_output_dir or None)
            if output_dir:
                self.batch_output_dir = output_dir
                if hasattr(self, 'output_dir_label'):
                    self.output_dir_label.config(text=output_dir)
                self.save_settings()
        
        if not output_dir:
            return
        
        # 确定要处理的图片列表
        # 始终处理所有选中的图片 (假设文件名唯一或自动重命名)
        images_to_process = self.batch_images

        
        success_count = 0
        preset_width = self.current_size_preset['width']
        preset_height = self.current_size_preset['height']
        
        # 开始日志
        self.batch_log(f"═══ 开始批量处理 ═══")
        self.batch_log(f"待处理: {len(images_to_process)} 张图片")
        self.batch_log("模式: 默认全量处理 (自动覆盖)")
            
        self.batch_log(f"输出目录: {output_dir}")
        self.batch_log(f"输出尺寸: {preset_width}x{preset_height}")
        
        # 记录本次会话处理数
        self.current_session_processed = 0
        
        # [EXCEL] 预加载文字映射
        text_mapping = {}
        text_sequence = []
        if self.batch_use_text_dir.get() and self.batch_text_dir:
            text_mapping, text_sequence = self._load_text_mapping(self.batch_text_dir)
            if text_mapping:
                self.batch_log(f"已加载 Excel 映射: {len(text_mapping)} 条记录")
            if text_sequence:
                self.batch_log(f"已加载 Excel 列表: {len(text_sequence)} 条记录 (及 {len(text_mapping)} 条指定映射)")
        
        for idx, img_path in enumerate(images_to_process):
            filename = os.path.basename(img_path)
            self.batch_log(f"[{idx+1}/{len(images_to_process)}] 处理: {filename}")
            self.update() # 刷新UI
            
            try:
                # 1. 加载图片
                processor = ImageProcessor()
                processor.load_image(img_path)
                processor.set_canvas_size(preset_width, preset_height)
                processor.resize_to_canvas(maintain_ratio=True)
                
                # 2. 准备边框配置 (支持随机化)
                border_config = self.border_config.copy()
                
                if self.batch_random_color.get():
                    new_color = self.get_random_color()
                    border_config['color'] = new_color
                    # 如果启用了图案且其颜色也是白色/默认，可能也需要随机？
                    # 简单策略：如果随机颜色，且有图案，图案也用这个色？或者图案颜色保持？
                    # 用户需求：Random Border Color. 
                
                if self.batch_random_style.get():
                    border_config['line_style'] = self.get_random_line_style()
                    
                if self.batch_random_pattern.get():
                    border_config['pattern'] = self.get_random_pattern()
                    # 自动调整图案大小
                    border_config['pattern_size'] = max(4, int(border_config['width'] * 0.6))
                
                
                # [SCALE FIX] 提前计算分辨率缩放比例
                # 所有的视觉元素（背景图案、边框宽度、文字大小）都需要基于预览比例进行缩放
                display_width = self.canvas_widget.width
                preview_scale = 1.0
                if display_width > 0:
                    preview_scale = preset_width / display_width

                # 3. 生成复合图片 (背景)
                composite = CompositeImage(
                    preset_width,
                    preset_height,
                    bg_color=self.background_color
                )
                
                # 绘制背景图案 (应用缩放)
                scaled_pattern_size = int(self.background_pattern_size * preview_scale)
                composite.draw_background_pattern(
                    self.background_pattern,
                    self.background_pattern_color,
                    scaled_pattern_size
                )
                
                # [LOGGING] 记录参考参数
                log_details = []
                
                # 4. 添加主图片
                if self.batch_match_canvas.get():
                    # 获取示例图的相对几何信息
                    geom = self.canvas_widget.get_main_image_geometry()
                    if geom:
                        rel_x, rel_y, rel_w, rel_h = geom
                        # 计算当前预设下的目标区域
                        target_x = rel_x * preset_width
                        target_y = rel_y * preset_height
                        target_w = rel_w * preset_width
                        target_h = rel_h * preset_height
                        
                        # [ENHANCED] 计算比例与缩放
                        cur_img = processor.get_current_image()
                        img_ratio = cur_img.width / cur_img.height if cur_img.height > 0 else 1.0
                        box_ratio = target_w / target_h if target_h > 0 else 1.0
                        
                        # 估算相对画布的缩放比例 (以宽为例)
                        # 假设原始 fit 是 contain 满画布
                        default_fit_w = preset_width if img_ratio > (preset_width/preset_height) else (preset_height * img_ratio)
                        scale_factor = target_w / default_fit_w if default_fit_w > 0 else 1.0
                        
                        # [SMART ALIGN] 智能对齐判断
                        # 如果参考位置非常靠上 (比如前 5%)，则判定为顶部对齐
                        # 如果参考位置非常靠下 (底部 5%)，则判定为底部对齐
                        anchor = 'center'
                        if rel_y < 0.05:
                            anchor = 'n'
                        elif (rel_y + rel_h) > 0.95:
                            anchor = 's'
                            
                        # 如果高度非常接近 (Full Height)，对齐方式影响不大，但保持 Default
                        
                        composite.add_main_image_with_geometry(
                            cur_img, 
                            target_x, target_y, target_w, target_h,
                            anchor=anchor
                        )
                        
                        anchor_map = {'n': '顶部', 's': '底部', 'center': '居中'}
                        log_details.append(f"参考位置: {rel_x:.2f},{rel_y:.2f} 尺寸: {rel_w:.2f}x{rel_h:.2f} => 目标: {int(target_x)},{int(target_y)} {int(target_w)}x{int(target_h)}")
                        log_details.append(f"比例检查: 图片{img_ratio:.2f} vs 目标框{box_ratio:.2f} | 缩放倍率: {scale_factor:.2f}x | 对齐: {anchor_map.get(anchor)}")
                    else:
                        # 获取失败回退到默认
                        composite.add_main_image(processor.get_current_image(), fit_mode='contain')
                        log_details.append("参考位置获取失败，已回退到默认")
                else:
                    composite.add_main_image(processor.get_current_image(), fit_mode='contain')
                    log_details.append("位置模式: 默认(适应画布)")
                
                # 记录边框随机化结果
                if self.batch_random_color.get():
                    log_details.append(f"随机颜色: {border_config.get('color')}")
                if self.batch_random_style.get():
                    log_details.append(f"随机样式: {border_config.get('line_style')}")
                if self.batch_random_pattern.get():
                    log_details.append(f"随机图案: {border_config.get('pattern')}")
                
                # 打印日志
                if log_details:
                    self.batch_log(f"  参数: {'; '.join(log_details)}")

                # 5. 应用边框到复合图片
                
                # 复制配置并应用缩放
                scaled_border_config = border_config.copy()
                if preview_scale != 1.0:
                    scaled_border_config['width'] = int(border_config.get('width', 0) * preview_scale)
                    scaled_border_config['radius'] = int(border_config.get('radius', 0) * preview_scale)
                    
                    # 只有在 random pattern 未覆盖时才缩放，或者统一缩放
                    # 如果 random pattern 已经被设置为 width * 0.6，那 width 已经是原始的了，这里应该缩放
                    # 但如果在 random 块里是用原始 width 计算的 size，那这里也要缩放 size
                    scaled_border_config['pattern_size'] = int(border_config.get('pattern_size', 0) * preview_scale)
                    
                # 根据形状判断调用哪个方法
                if scaled_border_config.get('shape') in ('rounded_rect', 'circle', 'ellipse') or scaled_border_config.get('radius', 0) > 0:
                    composite.add_rounded_border(scaled_border_config)
                else:
                    composite.add_border(scaled_border_config)
                
                # 添加贴纸 (需要在新尺寸下重新计算坐标)
                display_width = self.canvas_widget.width
                display_height = self.canvas_widget.height
                # 避免除以零
                if display_width > 0 and display_height > 0:
                    scale = max(preset_width / display_width, preset_height / display_height)
                else:
                    scale = 1.0
                
                for sticker in self.canvas_widget.get_stickers():
                    # 计算缩放后的位置和大小
                    scaled_x = int(sticker['x'] * scale)
                    scaled_y = int(sticker['y'] * scale)
                    scaled_size = int(sticker['size'] * scale)
                    
                    # 简单的贴纸添加 (暂不使用复杂Emoji渲染以保证稳定性，或者复用逻辑)
                    # 复用之前的Emoji渲染逻辑
                    try:
                        base_size = 160
                        font = ImageFont.truetype("/System/Library/Fonts/Apple Color Emoji.ttc", base_size)
                        temp_size = base_size * 2
                        emoji_temp = Image.new('RGBA', (temp_size, temp_size), (0, 0, 0, 0))
                        emoji_draw = ImageDraw.Draw(emoji_temp)
                        emoji_draw.text((temp_size // 2, temp_size // 2), sticker['text'], font=font, anchor="mm", embedded_color=True)
                        bbox = emoji_temp.getbbox()
                        if bbox:
                            emoji_cropped = emoji_temp.crop(bbox)
                            emoji_resized = emoji_cropped.resize((scaled_size, scaled_size), Image.Resampling.LANCZOS)
                            paste_x = scaled_x - scaled_size // 2
                            paste_y = scaled_y - scaled_size // 2
                            if composite.canvas.mode != 'RGBA':
                                composite.canvas = composite.canvas.convert('RGBA')
                            composite.canvas.paste(emoji_resized, (paste_x, paste_y), emoji_resized)
                    except Exception as e:
                        # 降级处理
                        composite.add_sticker(sticker['text'], scaled_x, scaled_y, scaled_size)
                
                # 6. 添加文字层
                text_content = None
                
                # 获取当前编辑器中的文字内容作为基础/兜底
                editor_content = None
                if hasattr(self, 'current_text_layer') and self.current_text_layer:
                    editor_content = self.current_text_layer.content

                # 只要勾选了"批量文字" (batch_use_text_dir)，就尝试添加文字
                # 逻辑：Excel映射 -> Excel顺序 -> .txt 文件 -> default.txt -> 编辑器文字
                if self.batch_use_text_dir.get():
                    
                    # 1. 尝试 Excel/列表 映射
                    if text_mapping and filename in text_mapping:
                        text_content = text_mapping[filename]
                        self.batch_log(f"  文字: Excel 匹配 ({filename})")
                    elif text_sequence and idx < len(text_sequence):
                        text_content = text_sequence[idx]
                        self.batch_log(f"  文字: Excel 顺序 (第{idx+1}行)")
                    
                    # 2. 尝试从文件读取 (目录模式 - 已移除)
                    # elif self.batch_text_dir and os.path.isdir(self.batch_text_dir):
                    #    pass
                    
                    # 3. Fallback: 使用编辑器文字
                    if not text_content and editor_content:
                        text_content = editor_content
                        self.batch_log(f"  文字: 使用编辑器配置")
                
                # 应用文字层
                if text_content:
                    from image_processor import TextLayer
                    text_layer = None
                    
                    # 优先克隆当前图层 (保证样式完全一致)
                    if hasattr(self, 'current_text_layer') and self.current_text_layer:
                        # 使用 to_dict/from_dict 克隆
                        layer_data = self.current_text_layer.to_dict()
                        # 更新内容
                        layer_data['content'] = text_content
                        text_layer = TextLayer.from_dict(layer_data)
                    else:
                        # Fallback: 使用 Config 创建 (可能样式不全)
                        cfg = self.current_text_config
                        text_layer = TextLayer(
                            content=text_content,
                            font_size=cfg.get('font_size', 48),
                            color=cfg.get('color', '#FFFFFF'),
                            font_family=cfg.get('font_family', 'pingfang'),
                            align=cfg.get('align', 'center'),
                            position=cfg.get('position', 'bottom'),
                            margin=cfg.get('margin', 20),
                            shadow=cfg.get('shadow'),
                            stroke=cfg.get('stroke'),
                            highlight=cfg.get('highlight'),
                            bold=cfg.get('bold', False),
                            italic=cfg.get('italic', False),
                            underline=cfg.get('underline', False),
                            indent=cfg.get('indent', False)
                        )

                    # [RANDOM FONT] 随机字体样式
                    if self.batch_random_font_style.get():
                        import random
                        # 使用与 UI 下拉框一致的字体列表 (keys: pingfang, heiti, etc.)
                        valid_fonts = list(text_layer.FONT_NAMES.keys())
                        text_layer.font_family = random.choice(valid_fonts)
                        text_layer.bold = random.choice([True, False])
                        text_layer.italic = random.choice([True, False])
                        # [RANDOM COLOR] 随机字体颜色
                        # [RANDOM COLOR] 智能对比度配色
                        from constants import MACARON_COLORS, DOPAMINE_COLORS
                        all_colors = MACARON_COLORS + DOPAMINE_COLORS
                        
                        # 1. 计算背景亮度
                        bg_brightness = 255 # 默认白背景
                        try:
                            # 尝试从 composite canvas 获取平均亮度 (简单采样)
                            if composite and composite.canvas:
                                # 缩略图采样
                                thumb = composite.canvas.resize((50, 50))
                                if thumb.mode != 'RGB':
                                    thumb = thumb.convert('RGB')
                                from PIL import ImageStat
                                stat = ImageStat.Stat(thumb)
                                r, g, b = stat.mean
                                bg_brightness = (r * 299 + g * 587 + b * 114) / 1000
                        except Exception as e:
                            print(f"[DEBUG] Calc bg brightness failed: {e}")
                            
                        # 2. 根据背景亮度筛选文字颜色
                        candidates = []
                        if bg_brightness < 100: # 深色背景
                            # 强制选亮色文字 (>150)
                            for c in all_colors:
                                try:
                                    crgb = tuple(int(c.lstrip('#')[i:i+2], 16) for i in (0, 2, 4))
                                    cb = (crgb[0] * 299 + crgb[1] * 587 + crgb[2] * 114) / 1000
                                    if cb > 150: candidates.append(c)
                                except: pass
                            if not candidates: candidates = ['#FFFFFF'] # Fallback
                        elif bg_brightness > 180: # 浅色背景
                             # 倾向选深色文字 (<120)
                             for c in all_colors:
                                try:
                                    crgb = tuple(int(c.lstrip('#')[i:i+2], 16) for i in (0, 2, 4))
                                    cb = (crgb[0] * 299 + crgb[1] * 587 + crgb[2] * 114) / 1000
                                    if cb < 120: candidates.append(c)
                                except: pass
                             # 如果没找到足够深的颜色，就随便选一个，但后面强制开描边
                             if not candidates: candidates = all_colors
                        else:
                            # 中性背景，什么颜色都行，靠描边补救
                            candidates = all_colors
                            
                        text_layer.color = random.choice(candidates)
                        
                        # 3. [RANDOM STROKE] 智能描边 (确保最终对比度)
                        if text_layer.stroke and text_layer.stroke.get('enabled'):
                             # 计算选中文字颜色的亮度
                            try:
                                c = str(text_layer.color).lstrip('#')
                                if len(c) == 6:
                                    rgb = tuple(int(c[i:i+2], 16) for i in (0, 2, 4))
                                    txt_brightness = (rgb[0] * 299 + rgb[1] * 587 + rgb[2] * 114) / 1000
                                else:
                                    txt_brightness = 200
                            except:
                                txt_brightness = 200
                            
                            # 策略:
                            # 如果背景亮 -> 需要深色元素 (文字深 或 描边深)
                            # 如果背景暗 -> 需要亮色元素 (文字亮 或 描边亮)
                            
                            final_stroke_color = '#000000'
                            
                            if bg_brightness > 150: # 浅色背景
                                if txt_brightness > 150: # 文字也亮 (对比度差)
                                    # 强制深色描边
                                    dark_strokes = ['#000000', '#333333', '#1A1A1A', '#2F4F4F', '#8B4513', '#800000', '#191970', '#006400']
                                    final_stroke_color = random.choice(dark_strokes)
                                else:
                                    # 文字深，背景亮 -> 描边可以使用浅色(形成光晕)或对比色，这里选安全的白色或浅色
                                    light_strokes = ['#FFFFFF', '#F0F8FF', '#F5F5F5']
                                    final_stroke_color = random.choice(light_strokes)
                            
                            elif bg_brightness < 100: # 深色背景
                                if txt_brightness < 100: # 文字也暗
                                    # 强制亮色描边
                                    light_strokes = ['#FFFFFF', '#F0F8FF', '#F5F5F5', '#FFFACD', '#E0FFFF', '#FFC0CB', '#98FB98']
                                    final_stroke_color = random.choice(light_strokes)
                                else:
                                    # 文字亮，背景暗 -> 描边可以用深色增加点缀
                                    dark_strokes = ['#000000', '#333333'] 
                                    final_stroke_color = random.choice(dark_strokes)
                            
                            else: # 中性背景
                                # 对比文字亮度即可
                                if txt_brightness > 128:
                                    final_stroke_color = '#333333'
                                else:
                                    final_stroke_color = '#FFFFFF'
                                    
                            text_layer.stroke['color'] = final_stroke_color
                            
                            # 确保描边宽度可见
                            if text_layer.stroke.get('width', 0) < 3:
                                text_layer.stroke['width'] = 4
                        
                    # [RANDOM HIGHLIGHT] 随机文字高亮 (配合 NLP)
                    if self.batch_random_highlight.get():
                         # 使用 'random' 字符串，让 image_processor 内部为每个关键词随机分配颜色 (彩虹效果)
                         random_hl_color = 'random'
                         
                         # 1. 确保 highlight 结构存在
                         if not text_layer.highlight or isinstance(text_layer.highlight, bool):
                             text_layer.highlight = {'enabled': True, 'keywords': [], 'color': random_hl_color}
                         else:
                             text_layer.highlight['enabled'] = True
                             text_layer.highlight['color'] = random_hl_color
                         
                             text_layer.highlight['enabled'] = True
                             text_layer.highlight['color'] = random_hl_color
                         
                         # [NLP] always try NLP first
                         try:
                             import jieba.analyse
                             curr_text = text_layer.content
                             # 提取 Top 5
                             extracted = jieba.analyse.extract_tags(curr_text, topK=5)
                             if extracted:
                                 text_layer.highlight['keywords'] = extracted
                                 log_details.append(f"NLP关键词: {extracted}")
                         except ImportError:
                             pass
                         except Exception as e:
                             print(f"Jieba failed: {e}")
                         
                         current_keywords = text_layer.highlight.get('keywords', [])
                         if not current_keywords:
                             # 简单的正则: 提取长度 >= 2 的词 (中文或单词)
                             import re
                             # 匹配中文或英文单词
                             words = re.findall(r'[\u4e00-\u9fa5]{2,}|[a-zA-Z]{4,}', text_content)
                             if words:
                                 # 随机选几个
                                 import random
                                 count = min(3, len(words))
                                 fallback_keywords = random.sample(words, count)
                                 text_layer.highlight['keywords'] = fallback_keywords
                                 log_details.append(f"正则兜底: {fallback_keywords}")
                         
                         log_details.append(f"随机高亮: {random_hl_color}")
                        
                    # 计算有效边框宽度 (用于文字防遮挡)
                    # [FIX] 使用 scaled_border_config 而非 border_config，确保边框宽度已缩放
                    effective_border_width = 0
                    if composite and composite.width and scaled_border_config and scaled_border_config.get('id') != 'none':
                         effective_border_width = scaled_border_config.get('width', 0)
                         # 稍微多给一点余量 (也要缩放)
                         effective_border_width += int(10 * preview_scale)

                    composite.add_text_layer(text_layer, scale=preview_scale, border_width=effective_border_width)
                
                # 7. 保存
                # [UNIQUE] 生成唯一文件名 (时间戳 + 随机数)防止覆盖
                import time
                import random
                from datetime import datetime
                name, ext = os.path.splitext(filename)
                # 格式: 原文件名_年月日时分秒毫秒
                time_str = datetime.now().strftime('%Y%m%d%H%M%S%f')[:-3]
                unique_filename = f"{name}_{time_str}{ext}"
                
                save_path = os.path.join(output_dir, unique_filename)
                if composite.save(save_path):
                    self.batch_log(f"  └─ 成功: {unique_filename}")
                    success_count += 1
                    self.current_session_processed += 1
                    
                    # [AUTH] 扣除使用次数
                    # [AUTH] 扣除使用次数
                    allowed, msg = auth.increment_usage(1)
                    if not allowed:
                        self.batch_log(f"  [STOP] {msg}")
                        messagebox.showwarning("限制提示", msg)
                        break
                else:
                    self.batch_log(f"  └─ 失败: 保存出错")
            
            except Exception as e:
                self.batch_log(f"  └─ 错误: {str(e)}")
                import traceback
                traceback.print_exc()
        
        self.batch_log(f"═══ 处理完成 ═══")
        self.batch_log(f"成功: {success_count} / {len(images_to_process)}")
        self.update_batch_status_text()
        if messagebox.askyesno('完成', f'批量处理完成！\n成功: {success_count}\n失败: {len(images_to_process) - success_count}\n\n是否打开所在目录？'):
            self.open_directory(output_dir)

    def save_history(self, action_name="操作"):
        """保存历史记录"""
        import copy
        from datetime import datetime
        
        # 创建状态快照
        # 创建状态快照
        state = {
            'timestamp': datetime.now().strftime('%H:%M:%S'),
            'action': action_name,
            'border_config': copy.deepcopy(self.border_config),
            'background_color': self.background_color,
            'background_pattern': self.background_pattern,
            'background_pattern_color': self.background_pattern_color,
            'background_pattern_size': self.background_pattern_size,
            'image': self.image_processor.current_image.copy() if self.image_processor.current_image else None,
            'stickers': copy.deepcopy(self.canvas_widget.stickers) if hasattr(self.canvas_widget, 'stickers') else [],
            # 保存文字配置
            'text_config': {
                'content': self.text_content_entry.get('1.0', 'end-1c') if hasattr(self, 'text_content_entry') else '',
                'color': self.text_color_var.get() if hasattr(self, 'text_color_var') else '#333333',
                'font_family': self.font_family_var.get() if hasattr(self, 'font_family_var') else '苹方 (默认)',
                'font_size': self.font_size_scale.get() if hasattr(self, 'font_size_scale') else 48,
                'align': self.text_align_var.get() if hasattr(self, 'text_align_var') else 'center',
                'position': self.text_position_var.get() if hasattr(self, 'text_position_var') else 'bottom',
                'margin': self.text_margin_var.get() if hasattr(self, 'text_margin_var') else 20,
                'bold': self.text_bold_var.get() if hasattr(self, 'text_bold_var') else False,
                'italic': self.text_italic_var.get() if hasattr(self, 'text_italic_var') else False,
                'underline': self.text_underline_var.get() if hasattr(self, 'text_underline_var') else False,
                'shadow_enabled': self.text_shadow_var.get() if hasattr(self, 'text_shadow_var') else False,
                'stroke_enabled': self.text_stroke_var.get() if hasattr(self, 'text_stroke_var') else False,
                'stroke_width': self.stroke_width_var.get() if hasattr(self, 'stroke_width_var') else 2,
                'stroke_color': self.stroke_color_var.get() if hasattr(self, 'stroke_color_var') else '#000000',
                'highlight_enabled': self.highlight_enabled_var.get() if hasattr(self, 'highlight_enabled_var') else True
            }
        }
        
        # 如果不是在历史末尾，删除后面的记录
        if self.history_index < len(self.history_stack) - 1:
            self.history_stack = self.history_stack[:self.history_index + 1]
        
        # 添加新记录
        self.history_stack.append(state)
        self.history_index = len(self.history_stack) - 1
        
        # 限制历史记录数量
        if len(self.history_stack) > self.max_history:
            self.history_stack.pop(0)
            self.history_index -= 1
        
        # 更新历史记录UI（如果存在）
        if hasattr(self, 'history_listbox'):
            self.update_history_display()
    
    def undo(self):
        """撤销"""
        if self.history_index > 0:
            self.history_index -= 1
            self.restore_state(self.history_stack[self.history_index])
            if hasattr(self, 'history_listbox'):
                self.update_history_display()
        else:
            messagebox.showinfo('提示', '没有更多可撤销的操作')
    
    def redo(self):
        """重做"""
        if self.history_index < len(self.history_stack) - 1:
            self.history_index += 1
            self.restore_state(self.history_stack[self.history_index])
            if hasattr(self, 'history_listbox'):
                self.update_history_display()
        else:
            messagebox.showinfo('提示', '没有更多可重做的操作')
    
    def restore_state(self, state):
        """恢复到指定状态"""
        import copy
        
        # 恢复边框配置
        self.border_config = copy.deepcopy(state['border_config'])
        
        # 恢复背景配置
        self.background_color = state['background_color']
        self.background_pattern = state['background_pattern']
        self.background_pattern_color = state['background_pattern_color']
        self.background_pattern_size = state['background_pattern_size']
        
        # 恢复图片
        if state['image']:
            self.image_processor.current_image = state['image'].copy()
        
        # 清空并恢复贴纸
        self.canvas_widget.canvas.delete('sticker')
        self.canvas_widget.stickers = []
        if state.get('stickers'):
            for sticker_data in state['stickers']:
                # 重新创建贴纸
                visible = sticker_data.get('visible', True)
                sticker_id = self.canvas_widget.canvas.create_text(
                    sticker_data['x'], sticker_data['y'],
                    text=sticker_data['text'],
                    font=('Arial', sticker_data['size']),
                    fill='black',
                    tags='sticker',
                    state='normal' if visible else 'hidden'
                )
                self.canvas_widget.stickers.append({
                    'id': sticker_id,
                    'text': sticker_data['text'],
                    'x': sticker_data['x'],
                    'y': sticker_data['y'],
                    'size': sticker_data['size'],
                    'visible': visible
                })
        
        # 刷新画布
        self.refresh_canvas()
        
        # 恢复文字配置
        if 'text_config' in state:
            tc = state['text_config']
            
            # 恢复UI变量
            if hasattr(self, 'text_content_entry'):
                self.text_content_entry.delete('1.0', tk.END)
                self.text_content_entry.insert('1.0', tc.get('content', ''))
                
            if hasattr(self, 'text_color_var'): self.text_color_var.set(tc.get('color', '#333333'))
            if hasattr(self, 'font_family_var'): self.font_family_var.set(tc.get('font_family', ''))
            if hasattr(self, 'font_size_scale'): self.font_size_scale.set(tc.get('font_size', 48))
            if hasattr(self, 'text_align_var'): self.text_align_var.set(tc.get('align', 'center'))
            if hasattr(self, 'text_position_var'): self.text_position_var.set(tc.get('position', 'bottom'))
            if hasattr(self, 'text_margin_var'): self.text_margin_var.set(tc.get('margin', 20))
            if hasattr(self, 'text_bold_var'): self.text_bold_var.set(tc.get('bold', False))
            if hasattr(self, 'text_italic_var'): self.text_italic_var.set(tc.get('italic', False))
            if hasattr(self, 'text_underline_var'): self.text_underline_var.set(tc.get('underline', False))
            if hasattr(self, 'text_shadow_var'): self.text_shadow_var.set(tc.get('shadow_enabled', False))
            if hasattr(self, 'text_stroke_var'): self.text_stroke_var.set(tc.get('stroke_enabled', False))
            if hasattr(self, 'stroke_width_var'): self.stroke_width_var.set(tc.get('stroke_width', 2))
            if hasattr(self, 'stroke_color_var'): self.stroke_color_var.set(tc.get('stroke_color', '#000000'))
            if hasattr(self, 'highlight_enabled_var'): self.highlight_enabled_var.set(tc.get('highlight_enabled', True))
            
            # 触发重新渲染文字
            self._auto_apply_text()
    
    def restore_to_history(self, index):
        """恢复到指定历史记录"""
        if 0 <= index < len(self.history_stack):
            self.history_index = index
            self.restore_state(self.history_stack[index])
            self.update_history_display()
    
    def update_history_display(self):
        """更新历史记录列表显示"""
        if not hasattr(self, 'history_listbox'):
            return
        
        self.history_listbox.delete(0, tk.END)
        for i, state in enumerate(self.history_stack):
            prefix = "▶ " if i == self.history_index else "  "
            self.history_listbox.insert(tk.END, f"{prefix}{state['timestamp']} - {state['action']}")
    
    def create_history_tab(self, parent):
        """历史记录标签页"""
        # 标题
        tk.Label(
            parent, text='📝 操作记录', font=('SF Pro Display', 14, 'bold'),
            bg=COLORS['panel_bg'], fg=COLORS['text_primary'], anchor='w'
        ).pack(fill=tk.X, padx=16, pady=(16, 8))
        
        # 撤销/重做按钮区
        btn_frame = tk.Frame(parent, bg=COLORS['panel_bg'])
        btn_frame.pack(fill=tk.X, padx=16, pady=(0, 12))
        
        undo_btn = tk.Label(
            btn_frame, text='↶ 撤销', font=('SF Pro Text', 10),
            bg=COLORS['bg_tertiary'], fg=COLORS['text_primary'],
            padx=12, pady=6, cursor='hand2'
        )
        undo_btn.pack(side=tk.LEFT, padx=(0, 4))
        undo_btn.bind('<Button-1>', lambda e: self.undo())
        undo_btn.bind('<Enter>', lambda e: undo_btn.config(bg=COLORS['hover']))
        undo_btn.bind('<Leave>', lambda e: undo_btn.config(bg=COLORS['bg_tertiary']))
        
        redo_btn = tk.Label(
            btn_frame, text='↷ 重做', font=('SF Pro Text', 10),
            bg=COLORS['bg_tertiary'], fg=COLORS['text_primary'],
            padx=12, pady=6, cursor='hand2'
        )
        redo_btn.pack(side=tk.LEFT)
        redo_btn.bind('<Button-1>', lambda e: self.redo())
        redo_btn.bind('<Enter>', lambda e: redo_btn.config(bg=COLORS['hover']))
        redo_btn.bind('<Leave>', lambda e: redo_btn.config(bg=COLORS['bg_tertiary']))
        
        clear_btn = tk.Label(
            btn_frame, text='清空', font=('SF Pro Text', 10),
            bg=COLORS['bg_tertiary'], fg=COLORS['text_secondary'],
            padx=12, pady=6, cursor='hand2'
        )
        clear_btn.pack(side=tk.RIGHT)
        clear_btn.bind('<Button-1>', lambda e: self.clear_history())
        clear_btn.bind('<Enter>', lambda e: clear_btn.config(bg=COLORS['hover']))
        clear_btn.bind('<Leave>', lambda e: clear_btn.config(bg=COLORS['bg_tertiary']))
        
        # 说明文字
        tk.Label(
            parent, text='点击记录可恢复到该状态', font=('SF Pro Text', 10),
            bg=COLORS['panel_bg'], fg=COLORS['text_secondary'], anchor='w'
        ).pack(fill=tk.X, padx=16, pady=(0, 8))
        
        # 历史记录列表
        list_frame = tk.Frame(parent, bg=COLORS['bg_tertiary'], relief=tk.FLAT)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=16, pady=(0, 16))
        
        self.history_listbox = tk.Listbox(
            list_frame, font=('SF Pro Text', 10),
            bg=COLORS['bg_tertiary'], fg=COLORS['text_primary'],
            selectbackground=COLORS['accent'], selectforeground='white',
            relief=tk.FLAT, borderwidth=0, highlightthickness=0,
            activestyle='none'
        )
        self.history_listbox.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)
        
        # 绑定点击事件
        self.history_listbox.bind('<ButtonRelease-1>', self.on_history_select)
        
        # 滚动条
        scrollbar = tk.Scrollbar(list_frame, command=self.history_listbox.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.history_listbox.config(yscrollcommand=scrollbar.set)
        
        # 初始化显示
        self.update_history_display()
    
    def on_history_select(self, event):
        """历史记录选择事件"""
        selection = self.history_listbox.curselection()
        if selection:
            index = selection[0]
            self.restore_to_history(index)
    
    def clear_history(self):
        """清空历史记录"""
        if messagebox.askyesno('确认', '确定要清空所有历史记录吗？'):
            self.history_stack = []
            self.history_index = -1
            self.update_history_display()
            
    def create_layer_tab(self, parent):
        """图层标签页"""
        # 标题栏
        header = tk.Frame(parent, bg=COLORS['panel_bg'])
        header.pack(fill=tk.X, padx=16, pady=(16, 8))
        
        tk.Label(
            header, text='📚 图层管理', font=('SF Pro Display', 14, 'bold'),
            bg=COLORS['panel_bg'], fg=COLORS['text_primary'], anchor='w'
        ).pack(side=tk.LEFT)
        
        # 刷新按钮
        refresh_btn = tk.Label(
            header, text='⟳', font=('SF Pro Text', 14),
            bg=COLORS['panel_bg'], fg=COLORS['text_secondary'],
            cursor='hand2'
        )
        refresh_btn.pack(side=tk.RIGHT)
        refresh_btn.bind('<Button-1>', lambda e: self.update_layer_list())
        
        # 图层列表容器
        list_frame = tk.Frame(parent, bg=COLORS['bg_tertiary'], relief=tk.FLAT)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=16, pady=(0, 16))
        
        # 自定义列表显示
        self.layer_list_frame = tk.Frame(list_frame, bg=COLORS['bg_tertiary'])
        self.layer_list_frame.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)
        
        # 初始化右键菜单
        self.layer_context_menu = tk.Menu(self, tearoff=0)
        self.layer_context_menu.add_command(label="👁️ 显示/隐藏", command=self.toggle_layer_visibility)
        self.layer_context_menu.add_separator()
        self.layer_context_menu.add_command(label="🗑️ 删除", command=self.delete_layer_item)
        
        # 初始加载
        self.update_layer_list()

    def update_layer_list(self):
        """更新图层列表显示"""
        if not hasattr(self, 'layer_list_frame'):
            return
            
        # 清空现有列表
        for widget in self.layer_list_frame.winfo_children():
            widget.destroy()
            
        # 获取所有图层项 (从上到下: 边框 -> 贴纸(反序) -> 主图 -> 背景)
        layers = []
        
        # 1. 边框 (如果存在或隐藏中)
        if self.border_config.get('width', 0) > 0 or getattr(self, '_temp_hidden_border_width', None):
            is_visible = not getattr(self, '_temp_hidden_border_width', None)
            layers.append({'type': 'border', 'name': '🖼️ 边框', 'id': 'border', 'visible': is_visible})
            
        # 2. 贴纸 (反序)
        if hasattr(self.canvas_widget, 'stickers'):
            for i, sticker in enumerate(reversed(self.canvas_widget.stickers)):
                text = sticker['text'][:10] + ('...' if len(sticker['text']) > 10 else '')
                is_visible = sticker.get('visible', True)
                layers.append({
                    'type': 'sticker', 
                    'name': f'✨ 贴纸: {text}', 
                    'id': sticker['id'],
                    'index': len(self.canvas_widget.stickers) - 1 - i,
                    'visible': is_visible
                })
                
        # 3. 主图片
        if self.image_processor.current_image:
            is_visible = self.canvas_widget.canvas.itemcget('main_image', 'state') != 'hidden'
            layers.append({'type': 'image', 'name': '📷 主图片', 'id': 'main_image', 'visible': is_visible})
            
        # 4. 背景
        is_visible = self.canvas_widget.canvas.itemcget('background_image', 'state') != 'hidden'
        layers.append({'type': 'background', 'name': '🎨 背景', 'id': 'background', 'visible': is_visible})
        
        # 渲染列表
        for idx, layer in enumerate(layers):
            item_frame = tk.Frame(self.layer_list_frame, bg=COLORS['bg_tertiary'])
            item_frame.pack(fill=tk.X, pady=1)
            
            # 可见性按钮 (眼睛图标)
            eye_icon = "👁️" if layer['visible'] else "⭕" # 使用圈圈代表闭眼/隐藏，或可用 🔒
            eye_label = tk.Label(
                item_frame, text=eye_icon, font=('SF Pro Text', 10),
                bg=COLORS['bg_tertiary'], fg=COLORS['text_secondary'] if layer['visible'] else COLORS['text_tertiary'],
                width=3, cursor='hand2'
            )
            eye_label.pack(side=tk.LEFT, fill=tk.Y)
            
            name_label = tk.Label(
                item_frame, text=layer['name'], font=('SF Pro Text', 10),
                bg=COLORS['bg_tertiary'], fg=COLORS['text_primary'],
                anchor='w', padx=8, pady=6
            )
            name_label.pack(fill=tk.X, side=tk.LEFT, expand=True)
            
            # 绑定可见性切换
            # 更新toggle_layer_visibility以接受上下文参数，或者我们在点击时设置context_layer
            def on_eye_click(e, l=layer):
                self.context_layer = l
                self.toggle_layer_visibility()
                
            eye_label.bind('<Button-1>', on_eye_click)
            
            # 事件绑定
            handler = lambda e, l=layer, f=item_frame: self.on_layer_select(l, f)
            name_label.bind('<Button-1>', handler)
            item_frame.bind('<Button-1>', handler)
            
            # 右键菜单
            ctx_handler = lambda e, l=layer: self.show_layer_context_menu(e, l)
            name_label.bind('<Button-2>', ctx_handler)
            name_label.bind('<Button-3>', ctx_handler)
            name_label.bind('<Control-Button-1>', ctx_handler)
            item_frame.bind('<Button-2>', ctx_handler)
            item_frame.bind('<Button-3>', ctx_handler)
            item_frame.bind('<Control-Button-1>', ctx_handler)

            # Hover
            def on_enter(e, f=item_frame, l=name_label, el=eye_label, lid=layer.get('id')):
                if getattr(self, 'selected_layer_id', None) != lid:
                    col = COLORS['hover']
                    f.config(bg=col)
                    l.config(bg=col)
                    el.config(bg=col)
            
            def on_leave(e, f=item_frame, l=name_label, el=eye_label, lid=layer.get('id')):
                if getattr(self, 'selected_layer_id', None) != lid:
                    col = COLORS['bg_tertiary']
                    f.config(bg=col)
                    l.config(bg=col)
                    el.config(bg=col)
            
            name_label.bind('<Enter>', on_enter)
            name_label.bind('<Leave>', on_leave)
            eye_label.bind('<Enter>', on_enter)
            eye_label.bind('<Leave>', on_leave)
            item_frame.bind('<Enter>', on_enter)
            item_frame.bind('<Leave>', on_leave)

    def on_layer_select(self, layer, item_frame):
        """图层选中处理"""
        self.selected_layer = layer
        self.selected_layer_id = layer.get('id')
        
        for widget in self.layer_list_frame.winfo_children():
            widget.config(bg=COLORS['bg_tertiary'])
            for child in widget.winfo_children():
                child.config(bg=COLORS['bg_tertiary'])
        
        item_frame.config(bg=COLORS['accent'])
        for child in item_frame.winfo_children():
            child.config(bg=COLORS['accent'])
            
        if layer['type'] == 'sticker':
            self.canvas_widget.select_item(layer['id'])
        else:
            self.canvas_widget.selected_item = None
            self.canvas_widget.canvas.delete('handle')
            
    def show_layer_context_menu(self, event, layer):
        """显示图层右键菜单"""
        self.on_layer_select(layer, event.widget.master)
        self.context_layer = layer
        self.layer_context_menu.entryconfig("🗑️ 删除", state=tk.NORMAL if layer['type'] in ['sticker', 'image'] else tk.DISABLED)
        self.layer_context_menu.post(event.x_root, event.y_root)
        
    def toggle_layer_visibility(self):
        """切换图层可见性"""
        if not hasattr(self, 'context_layer'): return
        ltype = self.context_layer['type']
        
        if ltype == 'sticker':
            iid = self.context_layer['id']
            curr = self.canvas_widget.canvas.itemcget(iid, 'state')
            new_state = 'hidden' if curr!='hidden' else 'normal'
            self.canvas_widget.canvas.itemconfigure(iid, state=new_state)
            
            # 更新数据中的可见性状态
            for sticker in self.canvas_widget.stickers:
                if sticker['id'] == iid:
                    sticker['visible'] = (new_state == 'normal')
                    break
            self.save_history("切换图层可见性")
            
        elif ltype == 'image':
            iid = 'main_image'
            curr = self.canvas_widget.canvas.itemcget(iid, 'state')
            new_state = 'hidden' if curr!='hidden' else 'normal'
            self.canvas_widget.canvas.itemconfigure(iid, state=new_state)
            
        elif ltype == 'border':
             # 简单切换边框宽度
            if self.border_config.get('width', 0) > 0:
                self._temp_hidden_border_width = self.border_config.get('width')
                self.border_config['width'] = 0
            elif getattr(self, '_temp_hidden_border_width', None):
                self.border_config['width'] = self._temp_hidden_border_width
                self._temp_hidden_border_width = None
            self.refresh_canvas()
        
        elif ltype == 'background':
            iid = 'background_image'
            curr = self.canvas_widget.canvas.itemcget(iid, 'state')
            new_state = 'hidden' if curr!='hidden' else 'normal'
            self.canvas_widget.canvas.itemconfigure(iid, state=new_state)
            
        # 刷新列表显示状态
        self.update_layer_list()
            
    def delete_layer_item(self):
        """删除图层项"""
        if not hasattr(self, 'context_layer'): return
        ltype = self.context_layer['type']
        
        if ltype == 'sticker':
            # 需要在 canvas_widget 中实现根据 ID 删除
            # 目前只有 delete_selected_sticker (删除选中的)
            # 既然我们已经选宏了它 (on_layer_select), delete_selected_sticker 应该有效
            self.canvas_widget.delete_selected_sticker()
            self.update_layer_list()
            self.save_history("删除贴纸")
        elif ltype == 'image':
            if messagebox.askyesno("确认", "确定要清除主图片吗？"):
                self.image_processor.current_image = None
                self.refresh_canvas()
                self.update_layer_list()
                self.save_history("删除图片")
    
    def create_background_tab(self, parent):
        """背景/主题标签页"""
        # 创建滚动区域
        scroll_canvas = tk.Canvas(parent, bg=COLORS['panel_bg'], highlightthickness=0)
        scrollbar = tk.Scrollbar(parent, orient='vertical', command=scroll_canvas.yview)
        scroll_frame = tk.Frame(scroll_canvas, bg=COLORS['panel_bg'])
        
        scroll_frame.bind('<Configure>', lambda e: scroll_canvas.configure(scrollregion=scroll_canvas.bbox('all')))
        scroll_canvas.create_window((0, 0), window=scroll_frame, anchor='nw')
        scroll_canvas.configure(yscrollcommand=scrollbar.set)
        
        # 自定义颜色标题(预设主题已移到左侧面板)
        tk.Label(
            scroll_frame, text='🌈 自定义颜色', font=('SF Pro Display', 14, 'bold'),
            bg=COLORS['panel_bg'], fg=COLORS['text_primary'], anchor='w'
        ).pack(fill=tk.X, padx=16, pady=(0, 12))
        
        # 自定义颜色区域 - 新布局
        custom_color_container = tk.Frame(scroll_frame, bg=COLORS['panel_bg'])
        custom_color_container.pack(fill=tk.X, padx=16, pady=(0, 16))
        
        # 1. 预览
        self.bg_color_preview = tk.Canvas(
            custom_color_container, width=60, height=60,
            bg=self.background_color, highlightthickness=2,
            highlightbackground=COLORS['separator']
        )
        self.bg_color_preview.pack(side=tk.LEFT, padx=(0, 12))
        self.bg_color_preview.bind('<Button-1>', lambda e: self.choose_background_color())
        self.bg_color_preview.config(cursor='hand2')
        
        # 2. 信息
        info_frame = tk.Frame(custom_color_container, bg=COLORS['panel_bg'])
        info_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        self.bg_color_hex_label = tk.Label(
            info_frame, text=self.background_color, font=('SF Mono', 14, 'bold'),
            bg=COLORS['panel_bg'], fg=COLORS['text_primary']
        )
        self.bg_color_hex_label.pack(anchor='w')
        
        tk.Label(
            info_frame, text='点击预览或按钮选择', font=('SF Pro Text', 10),
            bg=COLORS['panel_bg'], fg=COLORS['text_secondary']
        ).pack(anchor='w', pady=(2, 0))
        
        # 3. 按钮 (放在右侧，小一点)
        bg_choose_btn = tk.Label(
            custom_color_container, text='🎯 选择',
            bg=COLORS['accent'], fg='white', font=('SF Pro Text', 11),
            padx=10, pady=6, cursor='hand2'
        )
        bg_choose_btn.pack(side=tk.RIGHT, padx=(8, 0))
        bg_choose_btn.bind('<Button-1>', lambda e: self.choose_background_color())
        def make_hover(b):
            b.bind('<Enter>', lambda e: b.config(bg=COLORS['accent_hover']))
            b.bind('<Leave>', lambda e: b.config(bg=COLORS['accent']))
        make_hover(bg_choose_btn)
        
        # 快速颜色选择
        quick_color_frame = tk.Frame(scroll_frame, bg=COLORS['panel_bg'])
        quick_color_frame.pack(fill=tk.X, padx=16, pady=(0, 16))
        
        # 使用Grid布局显示大量颜色
        for idx, color in enumerate(QUICK_COLORS):
            row = idx // 10  # 每行10个
            col = idx % 10
            
            color_btn = tk.Canvas(
                quick_color_frame, width=24, height=24,
                bg=color, highlightthickness=1,
                highlightbackground=COLORS['separator'], cursor='hand2'
            )
            color_btn.grid(row=row, column=col, padx=2, pady=2)
            color_btn.bind('<Button-1>', lambda e, c=color: self.set_background_color(c))
        
        # 分隔线
        tk.Frame(scroll_frame, height=1, bg=COLORS['separator']).pack(fill=tk.X, padx=16, pady=8)
        
        # 背景图案
        tk.Label(
            scroll_frame, text='✦ 背景图案', font=('SF Pro Display', 14, 'bold'),
            bg=COLORS['panel_bg'], fg=COLORS['text_primary'], anchor='w'
        ).pack(fill=tk.X, padx=16, pady=(12, 12))
        
        bg_pattern_grid = tk.Frame(scroll_frame, bg=COLORS['panel_bg'])
        bg_pattern_grid.pack(fill=tk.X, padx=16, pady=(0, 16))
        
        self.bg_pattern_buttons = {}
        for idx, pattern in enumerate(BACKGROUND_PATTERNS):
            is_selected = pattern['id'] == self.background_pattern
            btn = tk.Label(
                bg_pattern_grid, text=f"{pattern['icon']}\n{pattern['name']}",
                bg=COLORS['accent'] if is_selected else COLORS['bg_tertiary'],
                fg=COLORS['text_bright'] if is_selected else COLORS['text_primary'],
                font=('SF Pro Text', 10, 'bold') if is_selected else ('SF Pro Text', 10),
                width=6, pady=6, cursor='hand2'
            )
            btn.grid(row=idx // 3, column=idx % 3, padx=4, pady=4)
            btn.bind('<Button-1>', lambda e, p=pattern['id']: self.set_bg_pattern(p))
            self.bg_pattern_buttons[pattern['id']] = btn
        
        # 图案颜色和大小
        pattern_config_frame = tk.Frame(scroll_frame, bg=COLORS['panel_bg'])
        pattern_config_frame.pack(fill=tk.X, padx=16, pady=(0, 16))
        
        # 图案颜色
        tk.Label(
            pattern_config_frame, text='图案颜色',
            font=('SF Pro Text', 11), bg=COLORS['panel_bg'], fg=COLORS['text_secondary']
        ).pack(anchor='w')
        
        pattern_color_row = tk.Frame(pattern_config_frame, bg=COLORS['panel_bg'])
        pattern_color_row.pack(fill=tk.X, pady=(4, 8))
        
        self.bg_pattern_color_canvas = tk.Canvas(
            pattern_color_row, width=40, height=40,
            bg=self.background_pattern_color, highlightthickness=1,
            highlightbackground=COLORS['separator'], cursor='hand2'
        )
        self.bg_pattern_color_canvas.pack(side=tk.LEFT, padx=(0, 8))
        self.bg_pattern_color_canvas.bind('<Button-1>', lambda e: self.choose_bg_pattern_color())
        
        # 图案大小滑块
        tk.Label(
            pattern_config_frame, text='图案大小',
            font=('SF Pro Text', 11), bg=COLORS['panel_bg'], fg=COLORS['text_secondary']
        ).pack(anchor='w')
        
        size_frame = tk.Frame(pattern_config_frame, bg=COLORS['panel_bg'])
        size_frame.pack(fill=tk.X, pady=(4, 0))
        
        self.bg_pattern_size_scale = tk.Scale(
            size_frame, from_=5, to=30, orient=tk.HORIZONTAL,
            command=self.on_bg_pattern_size_change, bg=COLORS['panel_bg'],
            highlightthickness=0, troughcolor=COLORS['separator'],
            activebackground=COLORS['accent'], length=150
        )
        self.bg_pattern_size_scale.set(self.background_pattern_size)
        self.bg_pattern_size_scale.pack(side=tk.LEFT)
        
        self.bg_pattern_size_label = tk.Label(
            size_frame, text=f'{self.background_pattern_size}px',
            font=('SF Mono', 10), bg=COLORS['panel_bg'], fg=COLORS['accent']
        )
        self.bg_pattern_size_label.pack(side=tk.LEFT, padx=(8, 0))
        
        # 绑定滚动 - 所有子控件创建完成后再绑定
        self.bind_mousewheel(scroll_canvas)
        self.bind_mousewheel(scroll_frame, scroll_canvas)
        
        scroll_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    
    def create_border_tab(self, parent):
        """边框标签页 - 美观版"""
        # 创建滚动区域
        scroll_canvas = tk.Canvas(parent, bg=COLORS['panel_bg'], highlightthickness=0)
        scrollbar = tk.Scrollbar(parent, orient='vertical', command=scroll_canvas.yview)
        scroll_frame = tk.Frame(scroll_canvas, bg=COLORS['panel_bg'])
        
        scroll_frame.bind('<Configure>', lambda e: scroll_canvas.configure(scrollregion=scroll_canvas.bbox('all')))
        scroll_canvas.create_window((0, 0), window=scroll_frame, anchor='nw')
        scroll_canvas.configure(yscrollcommand=scrollbar.set)
        
        scroll_canvas.configure(yscrollcommand=scrollbar.set)
        
        # 1. 边框大小
        tk.Label(
            scroll_frame, text='📏 大小与圆角', font=('SF Pro Display', 14, 'bold'),
            bg=COLORS['panel_bg'], fg=COLORS['text_primary'], anchor='w'
        ).pack(fill=tk.X, padx=16, pady=(12, 12))
        
        size_frame = tk.Frame(scroll_frame, bg=COLORS['panel_bg'])
        size_frame.pack(fill=tk.X, padx=16, pady=(0, 16))
        
        # 宽度
        tk.Label(
            size_frame, text='宽度', font=('SF Pro Text', 11),
            bg=COLORS['panel_bg'], fg=COLORS['text_secondary']
        ).grid(row=0, column=0, sticky='w', pady=4)
        
        self.border_width_scale = tk.Scale(
            size_frame, from_=0, to=100, orient=tk.HORIZONTAL,
            command=self.on_border_width_change, bg=COLORS['panel_bg'],
            highlightthickness=0, troughcolor=COLORS['separator'],
            activebackground=COLORS['accent'], length=180
        )
        self.border_width_scale.set(self.border_config['width'])
        self.border_width_scale.grid(row=0, column=1, padx=8, sticky='ew')
        
        # 圆角
        tk.Label(
            size_frame, text='圆角', font=('SF Pro Text', 11),
            bg=COLORS['panel_bg'], fg=COLORS['text_secondary']
        ).grid(row=1, column=0, sticky='w', pady=12)
        
        self.border_radius_scale = tk.Scale(
            size_frame, from_=0, to=100, orient=tk.HORIZONTAL,
            command=self.on_border_radius_change, bg=COLORS['panel_bg'],
            highlightthickness=0, troughcolor=COLORS['separator'],
            activebackground=COLORS['accent'], length=180
        )
        self.border_radius_scale.set(self.border_config['radius'])
        self.border_radius_scale.grid(row=1, column=1, padx=8, sticky='ew')
        
        # 分隔线
        tk.Frame(scroll_frame, height=1, bg=COLORS['separator']).pack(fill=tk.X, padx=16, pady=8)
        
        # 3. 颜色
        tk.Label(
            scroll_frame, text='🎨 颜色', font=('SF Pro Display', 14, 'bold'),
            bg=COLORS['panel_bg'], fg=COLORS['text_primary'], anchor='w'
        ).pack(fill=tk.X, padx=16, pady=(12, 8))
        
        color_container = tk.Frame(scroll_frame, bg=COLORS['panel_bg'])
        color_container.pack(fill=tk.X, padx=16, pady=(0, 16))
        
        # 预览
        self.border_color_canvas = tk.Canvas(
            color_container, width=60, height=60,
            bg=self.border_config['color'], highlightthickness=2,
            highlightbackground=COLORS['separator'], cursor='hand2'
        )
        self.border_color_canvas.pack(side=tk.LEFT, padx=(0, 12))
        self.border_color_canvas.bind('<Button-1>', lambda e: self.choose_border_color())
        
        # 信息
        info_frame = tk.Frame(color_container, bg=COLORS['panel_bg'])
        info_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        self.border_color_hex_label = tk.Label(
            info_frame, text=self.border_config['color'], 
            font=('SF Mono', 14, 'bold'),
            bg=COLORS['panel_bg'], fg=COLORS['accent']
        )
        self.border_color_hex_label.pack(anchor='w')
        
        tk.Label(
            info_frame, text='点击预览或按钮选择', font=('SF Pro Text', 10),
            bg=COLORS['panel_bg'], fg=COLORS['text_secondary']
        ).pack(anchor='w', pady=(2, 0))
        
        # 按钮
        border_color_btn = tk.Label(
            color_container, text='🎯 选择',
            bg=COLORS['bg_tertiary'], fg=COLORS['text_primary'], font=('SF Pro Text', 11),
            padx=10, pady=6, cursor='hand2'
        )
        border_color_btn.pack(side=tk.RIGHT, padx=(8,0))
        border_color_btn.bind('<Button-1>', lambda e: self.choose_border_color())
        def make_hover(b):
            b.bind('<Enter>', lambda e: b.config(bg=COLORS['hover']))
            b.bind('<Leave>', lambda e: b.config(bg=COLORS['bg_tertiary']))
        make_hover(border_color_btn)
        
        # 快速颜色选择
        quick_color_frame = tk.Frame(scroll_frame, bg=COLORS['panel_bg'])
        quick_color_frame.pack(fill=tk.X, padx=16, pady=(0, 16))
        
        quick_colors = QUICK_COLORS
        
        for idx, color in enumerate(quick_colors):
            row = idx // 10
            col = idx % 10
            
            color_btn = tk.Canvas(
                quick_color_frame, width=24, height=24,
                bg=color, highlightthickness=1,
                highlightbackground=COLORS['separator'], cursor='hand2'
            )
            color_btn.grid(row=row, column=col, padx=2, pady=2)
            color_btn.bind('<Button-1>', lambda e, c=color: self.set_border_color_quick(c))
        
        # 分隔线
        tk.Frame(scroll_frame, height=1, bg=COLORS['separator']).pack(fill=tk.X, padx=16, pady=8)
        
        # 5. 线条样式
        tk.Label(
            scroll_frame, text='〰 线条样式', font=('SF Pro Display', 13, 'bold'),
            bg=COLORS['panel_bg'], fg=COLORS['text_primary'], anchor='w'
        ).pack(fill=tk.X, padx=16, pady=(12, 8))
        
        line_style_grid = tk.Frame(scroll_frame, bg=COLORS['panel_bg'])
        line_style_grid.pack(fill=tk.X, padx=16, pady=(0, 16))
        
        self.line_style_buttons = {}
        for idx, style in enumerate(BORDER_LINE_STYLES):
            is_selected = style['id'] == self.border_config.get('line_style', 'solid')
            btn = tk.Label(
                line_style_grid, text=f"{style['icon']}\n{style['name']}",
                bg=COLORS['accent'] if is_selected else COLORS['bg_tertiary'],
                fg=COLORS['text_bright'] if is_selected else COLORS['text_primary'],
                font=('SF Pro Text', 10, 'bold') if is_selected else ('SF Pro Text', 10),
                width=6, pady=6, cursor='hand2'
            )
            btn.grid(row=0, column=idx, padx=4)
            btn.bind('<Button-1>', lambda e, s=style['id']: self.set_border_line_style(s))
            self.line_style_buttons[style['id']] = btn
        
        # 分隔线
        tk.Frame(scroll_frame, height=1, bg=COLORS['separator']).pack(fill=tk.X, padx=16, pady=8)
        
        # 6. 边框图案
        tk.Label(
            scroll_frame, text='✦ 边框图案', font=('SF Pro Display', 13, 'bold'),
            bg=COLORS['panel_bg'], fg=COLORS['text_primary'], anchor='w'
        ).pack(fill=tk.X, padx=16, pady=(12, 8))
        
        pattern_grid = tk.Frame(scroll_frame, bg=COLORS['panel_bg'])
        pattern_grid.pack(fill=tk.X, padx=16, pady=(0, 16))
        
        self.border_pattern_buttons = {}
        for idx, pattern in enumerate(BORDER_PATTERNS):
            is_selected = pattern['id'] == self.border_config.get('pattern', 'none')
            btn = tk.Label(
                pattern_grid, text=f"{pattern['icon']}\n{pattern['name']}",
                bg=COLORS['accent'] if is_selected else COLORS['bg_tertiary'],
                fg=COLORS['text_bright'] if is_selected else COLORS['text_primary'],
                font=('SF Pro Text', 10, 'bold') if is_selected else ('SF Pro Text', 10),
                width=6, pady=6, cursor='hand2'
            )
            
            # 两排布局 (每排5个)
            row = idx // 5
            col = idx % 5
            btn.grid(row=row, column=col, padx=4, pady=4)
            
            btn.bind('<Button-1>', lambda e, p=pattern['id']: self.set_border_pattern(p))
            self.border_pattern_buttons[pattern['id']] = btn
        
        # 分隔线
        tk.Frame(scroll_frame, height=1, bg=COLORS['separator']).pack(fill=tk.X, padx=16, pady=8)
        
        # 7. 操作按钮
        btn_frame = tk.Frame(scroll_frame, bg=COLORS['panel_bg'])
        btn_frame.pack(fill=tk.X, padx=16, pady=(12, 20))
        
        apply_btn = tk.Label(
            btn_frame, text='✓ 应用边框',
            bg=COLORS['accent'], fg='white', font=('SF Pro Text', 12, 'bold'),
            pady=10, cursor='hand2'
        )
        apply_btn.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8))
        apply_btn.bind('<Button-1>', lambda e: self.apply_custom_border())
        apply_btn.bind('<Enter>', lambda e: apply_btn.config(bg=COLORS['accent_hover']))
        apply_btn.bind('<Leave>', lambda e: apply_btn.config(bg=COLORS['accent']))
        
        clear_btn = tk.Label(
            btn_frame, text='✕ 清除边框',
            bg=COLORS['danger'], fg='white', font=('SF Pro Text', 12, 'bold'),
            pady=10, cursor='hand2'
        )
        clear_btn.pack(side=tk.LEFT, fill=tk.X, expand=True)
        clear_btn.bind('<Button-1>', lambda e: self.clear_border())
        clear_btn.bind('<Enter>', lambda e: clear_btn.config(bg='#FF6B6B'))
        clear_btn.bind('<Leave>', lambda e: clear_btn.config(bg=COLORS['danger']))
        
        # 绑定滚动 - 所有子控件创建完成后再绑定
        self.bind_mousewheel(scroll_canvas)
        self.bind_mousewheel(scroll_frame, scroll_canvas)
        
        scroll_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    
    def update_border_preview(self):
        """更新边框预览"""
        if not hasattr(self, 'border_preview_canvas'):
            return
        
        canvas = self.border_preview_canvas
        canvas.delete('all')
        
        w, h = 200, 120
        shape = self.border_config['shape']
        width = min(self.border_config['width'], 10)  # 预览中限制粗细
        radius = min(self.border_config['radius'], 25)
        color = self.border_config['color']
        
        if width <= 0:
            return
        
        # 绘制边框预览
        if shape == 'rectangle':
            for i in range(width):
                canvas.create_rectangle(
                    10 + i, 10 + i, w - 10 - i, h - 10 - i,
                    outline=color
                )
        elif shape == 'rounded_rect':
            # 圆角矩形预览
            self.draw_preview_rounded_rect(canvas, 10, 10, w - 10, h - 10, radius, width, color)
        elif shape in ('circle', 'ellipse'):
            for i in range(width):
                canvas.create_oval(
                    10 + i, 10 + i, w - 10 - i, h - 10 - i,
                    outline=color
                )
    
    def draw_preview_rounded_rect(self, canvas, x1, y1, x2, y2, radius, width, color):
        """在预览画布上绘制圆角矩形"""
        r = min(radius, min(x2-x1, y2-y1) // 4)
        
        for i in range(width):
            cx1, cy1 = x1 + i, y1 + i
            cx2, cy2 = x2 - i, y2 - i
            
            if r <= 0:
                canvas.create_rectangle(cx1, cy1, cx2, cy2, outline=color)
                continue
            
            # 四条直线
            canvas.create_line(cx1 + r, cy1, cx2 - r, cy1, fill=color)
            canvas.create_line(cx1 + r, cy2, cx2 - r, cy2, fill=color)
            canvas.create_line(cx1, cy1 + r, cx1, cy2 - r, fill=color)
            canvas.create_line(cx2, cy1 + r, cx2, cy2 - r, fill=color)
            
            # 四个圆角
            canvas.create_arc(cx1, cy1, cx1 + 2*r, cy1 + 2*r, 
                            start=90, extent=90, style='arc', outline=color)
            canvas.create_arc(cx2 - 2*r, cy1, cx2, cy1 + 2*r, 
                            start=0, extent=90, style='arc', outline=color)
            canvas.create_arc(cx1, cy2 - 2*r, cx1 + 2*r, cy2, 
                            start=180, extent=90, style='arc', outline=color)
            canvas.create_arc(cx2 - 2*r, cy2 - 2*r, cx2, cy2, 
                            start=270, extent=90, style='arc', outline=color)
    
    def set_border_line_style(self, style_id):
        """设置边框线条样式"""
        self.border_config['line_style'] = style_id
        # 更新按钮选中状态
        if hasattr(self, 'line_style_buttons'):
            for sid, btn in self.line_style_buttons.items():
                if sid == style_id:
                    btn.config(bg=COLORS['accent'], fg=COLORS['text_bright'], font=('SF Pro Text', 10, 'bold'))
                else:
                    btn.config(bg=COLORS['bg_tertiary'], fg=COLORS['text_primary'], font=('SF Pro Text', 10))
        self.update_border_preview()
        self.apply_border_realtime()
    
    def set_border_pattern(self, pattern_id):
        """设置边框图案"""
        self.border_config['pattern'] = pattern_id
        # 切换图案时也确保尺寸正确
        current_width = self.border_config.get('width', 10)
        self.border_config['pattern_size'] = max(4, int(current_width * 0.6))
        # 更新按钮选中状态
        if hasattr(self, 'border_pattern_buttons'):
            for pid, btn in self.border_pattern_buttons.items():
                if pid == pattern_id:
                    btn.config(bg=COLORS['accent'], fg=COLORS['text_bright'], font=('SF Pro Text', 10, 'bold'))
                else:
                    btn.config(bg=COLORS['bg_tertiary'], fg=COLORS['text_primary'], font=('SF Pro Text', 10))
        self.update_border_preview()
        self.apply_border_realtime()
    
    def set_bg_pattern(self, pattern_id):
        """设置背景图案"""
        self.background_pattern = pattern_id
        # 更新按钮选中状态
        if hasattr(self, 'bg_pattern_buttons'):
            for pid, btn in self.bg_pattern_buttons.items():
                if pid == pattern_id:
                    btn.config(bg=COLORS['accent'], fg=COLORS['text_bright'], font=('SF Pro Text', 10, 'bold'))
                else:
                    btn.config(bg=COLORS['bg_tertiary'], fg=COLORS['text_primary'], font=('SF Pro Text', 10))
        # 应用背景图案
        self.canvas_widget.set_background_pattern(
            self.background_pattern,
            self.background_color,
            self.background_pattern_color,
            self.background_pattern_size
        )
    
    def choose_bg_pattern_color(self):
        """选择背景图案颜色"""
        def on_color_selected(color):
            self.background_pattern_color = color
            if hasattr(self, 'bg_pattern_color_canvas'):
                self.bg_pattern_color_canvas.config(bg=color)
            # 应用背景图案
            self.canvas_widget.set_background_pattern(
                self.background_pattern,
                self.background_color,
                self.background_pattern_color,
                self.background_pattern_size
            )
        
        def on_realtime_preview(color):
            """实时预览图案颜色"""
            if hasattr(self, 'bg_pattern_color_canvas'):
                self.bg_pattern_color_canvas.config(bg=color)
            # 实时应用背景图案
            self.canvas_widget.set_background_pattern(
                self.background_pattern,
                self.background_color,
                color,
                self.background_pattern_size
            )
        
        ColorWheelPicker(self, self.background_pattern_color, on_color_selected, on_realtime_preview)
    
    def on_bg_pattern_size_change(self, value):
        """背景图案大小改变"""
        self.background_pattern_size = int(float(value))
        if hasattr(self, 'bg_pattern_size_label'):
            self.bg_pattern_size_label.config(text=f'{self.background_pattern_size}px')
        # 应用背景图案
        self.canvas_widget.set_background_pattern(
            self.background_pattern,
            self.background_color,
            self.background_pattern_color,
            self.background_pattern_size
        )
    
    def clear_border(self):
        """清除边框"""
        self.canvas_widget.canvas.delete('border')
        self.canvas_widget.canvas.delete('border_image')
        # 重置边框配置为默认值（但不设为0，以便重新设置）
        self.border_config['width'] = 10
        self.border_config['radius'] = 0
        self.border_config['shape'] = 'rectangle'
        self.border_config['color'] = '#007AFF'
        
        # 更新滑块和按钮状态
        if hasattr(self, 'border_width_scale'):
            self.border_width_scale.set(10)
        if hasattr(self, 'border_radius_scale'):
            self.border_radius_scale.set(0)
        if hasattr(self, 'border_width_value'):
            self.border_width_value.config(text="10px")
        if hasattr(self, 'border_radius_value'):
            self.border_radius_value.config(text="0px")
        if hasattr(self, 'border_color_canvas'):
            self.border_color_canvas.config(bg='#007AFF')
        if hasattr(self, 'border_color_hex_label'):
            self.border_color_hex_label.config(text='#007AFF')
        
        # 更新形状按钮
        if hasattr(self, 'border_shape_buttons'):
            for sid, btn in self.border_shape_buttons.items():
                if sid == 'rectangle':
                    btn.config(
                        bg=COLORS['accent'], fg=COLORS['selected_text'],
                        font=('SF Pro Text', 10, 'bold'),
                        highlightthickness=2, highlightbackground=COLORS['accent']
                    )
                else:
                    btn.config(
                        bg=COLORS['panel_bg'], fg=COLORS['text_secondary'],
                        font=('SF Pro Text', 10),
                        highlightthickness=1, highlightbackground=COLORS['separator']
                    )
        
        # 更新预览
        self.update_border_preview()
        print("✓ 边框已清除，可重新设置")
    
    def set_border_shape(self, shape_id):
        """设置边框形状"""
        self.border_config['shape'] = shape_id
        for sid, btn in self.border_shape_buttons.items():
            if sid == shape_id:
                btn.config(
                    bg=COLORS['accent'], fg=COLORS['text_bright'],
                    font=('SF Pro Text', 10, 'bold')
                )
            else:
                btn.config(
                    bg=COLORS['bg_tertiary'], fg=COLORS['text_primary'],
                    font=('SF Pro Text', 10)
                )
        self.update_border_preview()
        self.apply_border_realtime()  # 实时应用
    
    def on_border_width_change(self, value):
        """边框粗细改变"""
        width = int(float(value))
        self.border_config['width'] = width
        if hasattr(self, 'border_width_value'):
            self.border_width_value.config(text=f"{width}px")
        self.update_border_preview()
        self.apply_border_realtime()  # 实时应用
    
    def on_border_radius_change(self, value):
        """圆角改变"""
        radius = int(float(value))
        self.border_config['radius'] = radius
        if hasattr(self, 'border_radius_value'):
            self.border_radius_value.config(text=f"{radius}px")
        self.update_border_preview()
        self.apply_border_realtime()  # 实时应用
    
    def choose_border_color(self):
        """选择边框颜色 - 使用颜色圆盘"""
        def on_color_selected(color):
            self.border_config['color'] = color
            if hasattr(self, 'border_color_canvas'):
                self.border_color_canvas.config(bg=color)
            if hasattr(self, 'border_color_hex_label'):
                self.border_color_hex_label.config(text=color)
            self.update_border_preview()
        
        def on_realtime_preview(color):
            """实时预览边框颜色"""
            self.border_config['color'] = color
            if hasattr(self, 'border_color_canvas'):
                self.border_color_canvas.config(bg=color)
            if hasattr(self, 'border_color_hex_label'):
                self.border_color_hex_label.config(text=color)
            self.update_border_preview()
            self.apply_border_realtime()
        
        ColorWheelPicker(self, self.border_config['color'], on_color_selected, on_realtime_preview)
    
    def set_border_color_quick(self, color):
        """快速设置边框颜色"""
        self.border_config['color'] = color
        if hasattr(self, 'border_color_canvas'):
            self.border_color_canvas.config(bg=color)
        if hasattr(self, 'border_color_hex_label'):
            self.border_color_hex_label.config(text=color)
        self.update_border_preview()
        self.apply_border_realtime()  # 实时应用
    
    def apply_border_realtime(self):
        """实时应用边框到画布"""
        if self.border_config['width'] > 0:
            self.canvas_widget.apply_custom_border(self.border_config)
            self.save_history("修改边框")
    
    def apply_custom_border(self):
        """应用自定义边框"""
        print(f"✓ 应用边框: {self.border_config}")
        self.canvas_widget.apply_custom_border(self.border_config)

    def create_sticker_tab(self, parent):
        """贴纸标签页"""
        sticker_label = tk.Label(
            parent, text='点击添加贴纸', font=('SF Pro Display', 13, 'bold'),
            bg=COLORS['panel_bg'], fg=COLORS['text_primary'], anchor='w'
        )
        sticker_label.pack(fill=tk.X, padx=16, pady=(16, 8))
        
        sticker_grid = tk.Frame(parent, bg=COLORS['panel_bg'])
        sticker_grid.pack(fill=tk.X, padx=12, pady=(0, 16))
        
        for idx, sticker in enumerate(STICKER_LIST):
            row = idx // 4
            col = idx % 4
            
            # 使用Label替代Button
            if sticker['id'] in self.sticker_images:
                btn = tk.Label(
                    sticker_grid, image=self.sticker_images[sticker['id']],
                    bg=COLORS['bg_tertiary'], cursor='hand2'
                )
            else:
                btn = tk.Label(
                    sticker_grid, text=sticker['emoji'], font=('Apple Color Emoji', 28),
                    bg=COLORS['bg_tertiary'], width=2, height=1, cursor='hand2'
                )
            btn.grid(row=row, column=col, padx=4, pady=4)
            btn.bind('<Button-1>', lambda e, s=sticker: self.add_sticker(s))
            btn.bind('<Enter>', lambda e, b=btn: b.config(bg=COLORS['hover']))
            btn.bind('<Leave>', lambda e, b=btn: b.config(bg=COLORS['bg_tertiary']))
    
    def set_background_color(self, color):
        """设置背景颜色"""
        self.background_color = color
        self.canvas_widget.set_background_color(color)
        # 更新预览Canvas
        if hasattr(self, 'bg_color_preview'):
            self.bg_color_preview.config(bg=color)
        # 更新颜色值标签
        if hasattr(self, 'bg_color_hex_label'):
            self.bg_color_hex_label.config(text=color)
        # 更新选中效果 - 高亮当前选中的颜色
        if hasattr(self, 'bg_color_canvases'):
            for bg_id, canvas in self.bg_color_canvases.items():
                # 找到匹配的预设颜色
                is_selected = False
                for bg_preset in DEFAULT_BACKGROUNDS:
                    if bg_preset['id'] == bg_id and bg_preset['color'] == color:
                        is_selected = True
                        break
                # 设置高亮边框
                if is_selected:
                    canvas.config(highlightbackground='#007AFF', highlightthickness=3)
                else:
                    canvas.config(highlightbackground='#E5E5EA', highlightthickness=2)
        print(f"✓ 背景颜色: {color}")
        self.save_history("修改背景")
    
    def choose_background_color(self):
        """选择背景颜色 - 使用颜色圆盘"""
        def on_color_selected(color):
            self.set_background_color(color)
        
        def on_realtime_preview(color):
            """实时预览背景颜色"""
            self.set_background_color(color)
        
        ColorWheelPicker(self, self.background_color, on_color_selected, on_realtime_preview)
    
    def upload_background_image(self):
        """上传背景图片"""
        file_path = filedialog.askopenfilename(
            title='选择背景图片',
            filetypes=[('图片文件', '*.jpg *.jpeg *.png *.bmp'), ('所有文件', '*.*')]
        )
        if file_path:
            try:
                img = Image.open(file_path)
                self.background_image = img
                self.canvas_widget.set_background_image(img)
                print(f"✓ 背景图片已设置")
            except Exception as e:
                messagebox.showerror('错误', f'加载图片失败: {e}')
    
    def clear_background_image(self):
        """清除背景图片"""
        self.background_image = None
        self.canvas_widget.set_background_color(self.background_color)
        print("✓ 背景图片已清除")
    
    def generate_theme_thumbnail(self, theme, size=50):
        """生成主题缩略图"""
        from PIL import Image, ImageDraw, ImageTk
        
        # 创建缩略图画布
        img = Image.new('RGB', (size, size), theme.get('background_color', '#FFFFFF'))
        draw = ImageDraw.Draw(img)
        
        # 绘制背景图案（简化版）
        pattern = theme.get('background_pattern', 'none')
        pattern_color = theme.get('background_pattern_color', '#E0E0E0')
        if pattern == 'grid':
            spacing = 10
            for x in range(0, size, spacing):
                draw.line([(x, 0), (x, size)], fill=pattern_color, width=1)
            for y in range(0, size, spacing):
                draw.line([(0, y), (size, y)], fill=pattern_color, width=1)
        elif pattern == 'dots':
            spacing = 8
            for x in range(spacing//2, size, spacing):
                for y in range(spacing//2, size, spacing):
                    draw.ellipse([x-1, y-1, x+1, y+1], fill=pattern_color)
        elif pattern == 'stripe':
            spacing = 6
            for i in range(-size, size, spacing):
                draw.line([(i, 0), (i + size, size)], fill=pattern_color, width=1)
        
        # 绘制边框
        border_config = theme.get('border_config', {})
        border_width = min(border_config.get('width', 0) // 3, 5)  # 缩小边框
        if border_width > 0:
            border_color = border_config.get('color', '#000000')
            radius = min(border_config.get('radius', 0) // 4, 8)
            if radius > 0:
                draw.rounded_rectangle([0, 0, size-1, size-1], radius=radius, outline=border_color, width=border_width)
            else:
                for i in range(border_width):
                    draw.rectangle([i, i, size-1-i, size-1-i], outline=border_color)
        
        # 转换为 PhotoImage
        return ImageTk.PhotoImage(img)
        
    def get_current_theme_state(self):
        """获取当前主题状态"""
        return {
            'background_color': self.background_color,
            'background_pattern': self.background_pattern,
            'background_pattern_color': self.background_pattern_color,
            'background_pattern_size': self.background_pattern_size,
            'border_config': self.border_config.copy(),
            'stickers': self.canvas_widget.get_stickers()
        }

    def apply_theme_state(self, state):
        """应用主题状态"""
        # 应用背景
        self.set_background_color(state['background_color'])
        self.set_bg_pattern(state['background_pattern'])
        self.background_pattern_color = state['background_pattern_color']
        self.background_pattern_size = state['background_pattern_size']
        self.canvas_widget.set_background_pattern(
            self.background_pattern, 
            self.background_color, 
            self.background_pattern_color, 
            self.background_pattern_size
        )
        if hasattr(self, 'bg_pattern_color_canvas'):
            self.bg_pattern_color_canvas.config(bg=self.background_pattern_color)
        if hasattr(self, 'bg_pattern_size_scale'):
            self.bg_pattern_size_scale.set(self.background_pattern_size)
            self.bg_pattern_size_label.config(text=f'{self.background_pattern_size}px')
        
        # 应用边框
        self.border_config = state['border_config'].copy()
        self.canvas_widget.apply_custom_border(self.border_config)
        # 更新边框UI状态
        self.selected_border_color = self.border_config['color']
        if hasattr(self, 'border_width_scale'):
            self.border_width_scale.set(self.border_config['width'])
        if hasattr(self, 'border_radius_scale'):
            self.border_radius_scale.set(self.border_config['radius'])
        if hasattr(self, 'border_color_canvas'):
            self.border_color_canvas.config(bg=self.border_config['color'])
        if hasattr(self, 'border_color_hex_label'):
            self.border_color_hex_label.config(text=self.border_config['color'])
        if hasattr(self, 'update_border_preview'):
            self.update_border_preview()
        
        # 应用贴纸
        self.canvas_widget.delete_selected_sticker()
        for sticker in self.canvas_widget.stickers:
            self.canvas_widget.canvas.delete(sticker['id'])
        self.canvas_widget.stickers = []
        
        for sticker_data in state['stickers']:
            s_id = self.canvas_widget.canvas.create_text(
                sticker_data['x'], sticker_data['y'],
                text=sticker_data['text'],
                font=('Arial', sticker_data['size']),
                fill='black',
                tags='sticker'
            )
            new_s = sticker_data.copy()
            new_s['id'] = s_id
            self.canvas_widget.stickers.append(new_s)
            
    def save_preset_theme(self, index=None, silent=False):
        """保存当前为预设主题
        
        Args:
            index: 保存到的索引位置（目前未使用）
            silent: 如果为True，则不显示成功提示
        """
        state = self.get_current_theme_state()
        
        if len(self.preset_themes) >= 8:
            if not messagebox.askyesno("提示", "预设已满(8个)，保存新预设将覆盖最早的预设，是否继续？"):
                return
            self.preset_themes.pop(0)
            self.preset_themes.append(state)
        else:
            self.preset_themes.append(state)
            
        self.save_settings() # 保存设置 (包含预设)
        self.update_preset_theme_display()
        self.update_left_preset_display()
        
        if not silent:
            messagebox.showinfo("成功", "主题已保存！")

    def apply_preset_theme(self, index):
        """应用预设主题"""
        if 0 <= index < len(self.preset_themes):
            self.apply_theme_state(self.preset_themes[index])
            
    def update_preset_theme_display(self):
        """更新预设主题显示区域"""
        if hasattr(self, 'preset_grid_frame'):
            for widget in self.preset_grid_frame.winfo_children():
                widget.destroy()
            
            # 清理旧的缩略图引用
            if not hasattr(self, 'preset_thumbnails'):
                self.preset_thumbnails = []
            self.preset_thumbnails.clear()
            
            for i in range(8):
                row = i // 4
                col = i % 4
                
                container = tk.Frame(self.preset_grid_frame, bg=COLORS['panel_bg'])
                container.grid(row=row, column=col, padx=4, pady=4)
                
                if i < len(self.preset_themes):
                    # 生成缩略图
                    theme = self.preset_themes[i]
                    thumbnail = self.generate_theme_thumbnail(theme, size=50)
                    self.preset_thumbnails.append(thumbnail)
                    
                    btn = tk.Label(
                        container,
                        image=thumbnail,
                        bg=COLORS['bg_tertiary'],
                        cursor='hand2',
                        relief=tk.FLAT,
                        bd=2
                    )
                    btn.pack()
                    btn.bind('<Button-1>', lambda e, idx=i: self.apply_preset_theme(idx))
                    # Hover effect
                    def make_hover(b):
                        b.bind('<Enter>', lambda e: b.config(bg=COLORS['hover']))
                        b.bind('<Leave>', lambda e: b.config(bg=COLORS['bg_tertiary']))
                    make_hover(btn)
                else:
                    btn = tk.Label(
                        container,
                        text="＋",
                        bg=COLORS['bg_secondary'],
                        fg=COLORS['text_secondary'],
                        font=('SF Pro Text', 14),
                        width=5, height=2,
                        cursor='hand2'
                    )
                    btn.pack()
                    btn.bind('<Button-1>', lambda e: self.save_preset_theme())
                    # Hover effect
                    def make_hover(b):
                        b.bind('<Enter>', lambda e: b.config(bg=COLORS['hover']))
                        b.bind('<Leave>', lambda e: b.config(bg=COLORS['bg_secondary']))
                    make_hover(btn)

    def update_left_preset_display(self):
        """更新左侧面板的预设主题显示"""
        if not hasattr(self, 'left_preset_grid'):
            return
            
        # 清空现有按钮
        for widget in self.left_preset_grid.winfo_children():
            widget.destroy()
        
        # 清理旧的缩略图引用
        if not hasattr(self, 'left_preset_thumbnails'):
            self.left_preset_thumbnails = []
        self.left_preset_thumbnails.clear()
        
        # 创建2列4行的按钮网格
        for i in range(8):
            row = i // 2
            col = i % 2
            
            if i < len(self.preset_themes):
                # 生成缩略图
                theme = self.preset_themes[i]
                thumbnail = self.generate_theme_thumbnail(theme, size=40)
                self.left_preset_thumbnails.append(thumbnail)
                
                # 已保存的预设 - 使用缩略图
                btn = tk.Label(
                    self.left_preset_grid,
                    image=thumbnail,
                    bg=COLORS['bg_tertiary'],
                    cursor='hand2',
                    relief=tk.FLAT,
                    bd=1
                )
                btn.grid(row=row, column=col, padx=2, pady=2, sticky='ew')
                btn.bind('<Button-1>', lambda e, idx=i: self.apply_preset_theme(idx))
                
                def make_hover(b):
                    b.bind('<Enter>', lambda e: b.config(bg=COLORS['hover']))
                    b.bind('<Leave>', lambda e: b.config(bg=COLORS['bg_tertiary']))
                make_hover(btn)
            else:
                # 空槽位 - 点击保存新预设
                btn = tk.Label(
                    self.left_preset_grid,
                    text="＋",
                    bg=COLORS['bg_secondary'],
                    fg=COLORS['text_secondary'],
                    font=('SF Pro Text', 12),
                    width=4, height=2,
                    cursor='hand2'
                )
                btn.grid(row=row, column=col, padx=2, pady=2, sticky='ew')
                btn.bind('<Button-1>', lambda e: self.save_preset_theme())
                
                def make_hover(b):
                    b.bind('<Enter>', lambda e: b.config(bg=COLORS['hover']))
                    b.bind('<Leave>', lambda e: b.config(bg=COLORS['bg_secondary']))
                make_hover(btn)
        
        # 配置列权重使按钮均匀分布
        self.left_preset_grid.columnconfigure(0, weight=1)
        self.left_preset_grid.columnconfigure(1, weight=1)
